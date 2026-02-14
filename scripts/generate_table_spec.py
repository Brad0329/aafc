"""
AAFC Django 테이블 상세 명세서 생성 스크립트
기존 주니어_테이블_명세서_20260209.xlsx 와 동일한 포맷으로 생성
"""
import os
import sys
import django

# Django 설정
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.local')
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.apps import apps


def get_field_info(field):
    """Django 필드에서 명세서 정보 추출"""
    field_name = field.column if hasattr(field, 'column') else field.name

    # 타입
    internal_type = field.get_internal_type()
    type_map = {
        'AutoField': 'SERIAL',
        'BigAutoField': 'BIGSERIAL',
        'CharField': 'VARCHAR',
        'TextField': 'TEXT',
        'IntegerField': 'INTEGER',
        'BigIntegerField': 'BIGINT',
        'SmallIntegerField': 'SMALLINT',
        'FloatField': 'FLOAT',
        'DecimalField': 'DECIMAL',
        'BooleanField': 'BOOLEAN',
        'DateField': 'DATE',
        'DateTimeField': 'TIMESTAMP',
        'TimeField': 'TIME',
        'EmailField': 'VARCHAR',
        'URLField': 'VARCHAR',
        'FileField': 'VARCHAR',
        'ImageField': 'VARCHAR',
        'ForeignKey': 'INTEGER',
        'OneToOneField': 'INTEGER',
    }
    db_type = type_map.get(internal_type, internal_type.upper())

    # Length
    length = ''
    if hasattr(field, 'max_length') and field.max_length:
        length = str(field.max_length)

    # Key
    key = ''
    if field.primary_key:
        key = 'PK'
    elif hasattr(field, 'unique') and field.unique and not field.primary_key:
        key = 'UQ'
    elif hasattr(field, 'related_model') and field.related_model:
        key = 'FK'

    # Default
    default = ''
    if field.has_default():
        d = field.default
        if d is not None and d != '':
            if callable(d):
                default = 'auto'
            else:
                default = str(d)

    # 비고 (verbose_name + 추가 정보)
    verbose = str(field.verbose_name) if field.verbose_name else ''
    notes_parts = []
    if verbose and verbose != field.name:
        notes_parts.append(verbose)

    if hasattr(field, 'related_model') and field.related_model:
        ref_table = field.related_model._meta.db_table
        to_field = getattr(field, 'to_fields', [None])[0] if hasattr(field, 'to_fields') else None
        if to_field:
            notes_parts.append(f'FK→{ref_table}({to_field})')
        else:
            notes_parts.append(f'FK→{ref_table}')

    if hasattr(field, 'choices') and field.choices:
        choices_str = ', '.join([f'{c[0]}={c[1]}' for c in field.choices])
        if len(choices_str) <= 80:
            notes_parts.append(f'[{choices_str}]')

    null_info = []
    if field.null:
        null_info.append('NULL')
    if field.blank:
        null_info.append('blank')
    if null_info:
        notes_parts.append(', '.join(null_info))

    notes = ' | '.join(notes_parts)

    return {
        'column': field_name,
        'type': db_type,
        'length': length,
        'key': key,
        'default': default,
        'notes': notes,
    }


def generate_spec():
    wb = Workbook()
    ws = wb.active
    ws.title = '테이블상세'

    # 스타일 정의
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)

    # 열 너비
    ws.column_dimensions['A'].width = 5    # No
    ws.column_dimensions['B'].width = 18   # 관리영역
    ws.column_dimensions['C'].width = 25   # 관리항목
    ws.column_dimensions['D'].width = 25   # 컬럼명
    ws.column_dimensions['E'].width = 15   # 타입
    ws.column_dimensions['F'].width = 10   # Length
    ws.column_dimensions['G'].width = 8    # Key
    ws.column_dimensions['H'].width = 12   # Default
    ws.column_dimensions['I'].width = 60   # 비고

    # 앱 순서 및 한글 영역명
    app_labels = {
        'accounts': '회원관리',
        'common': '공통코드',
        'courses': '강좌/구장/코치',
        'enrollment': '수강신청',
        'payments': '결제',
        'board': '게시판',
        'consult': '상담',
        'shop': '쇼핑몰',
        'points': '포인트',
        'notifications': '알림/SMS',
    }

    app_order = ['accounts', 'common', 'courses', 'enrollment', 'payments',
                 'board', 'consult', 'shop', 'points', 'notifications']

    row = 1

    # 문서 제목
    ws.cell(row=row, column=1, value='AAFC Django 테이블 상세 명세서')
    ws.cell(row=row, column=1).font = Font(bold=True, size=16)
    row += 1
    ws.cell(row=row, column=1, value='생성일: 2026-02-14 / Django + PostgreSQL')
    ws.cell(row=row, column=1).font = Font(size=10, italic=True)
    row += 2

    table_no = 0

    for app_label in app_order:
        try:
            app_config = apps.get_app_config(app_label)
        except LookupError:
            continue

        models = app_config.get_models()
        area_name = app_labels.get(app_label, app_label)

        for model in models:
            table_no += 1
            db_table = model._meta.db_table
            model_name = model.__name__
            verbose_name = str(model._meta.verbose_name) if model._meta.verbose_name else model_name

            # 테이블명 헤더 행 (노란 배경)
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.fill = yellow_fill
                cell.border = thin_border
            ws.cell(row=row, column=1, value=table_no).font = title_font
            ws.cell(row=row, column=1).fill = yellow_fill
            ws.cell(row=row, column=2, value=f'{db_table}').font = title_font
            ws.cell(row=row, column=2).fill = yellow_fill
            ws.cell(row=row, column=5, value=f'Model: {model_name}').font = Font(bold=True, size=11)
            ws.cell(row=row, column=5).fill = yellow_fill
            ws.cell(row=row, column=8, value=f'{verbose_name}').font = Font(size=11)
            ws.cell(row=row, column=8).fill = yellow_fill
            row += 1

            # 컬럼 헤더 행 (회색 배경)
            headers = ['No', '관리영역', '관리항목', '컬럼명', '타입', 'Length', 'Key', 'Default', '비고']
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.fill = gray_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row += 1

            # 필드 행
            fields = model._meta.get_fields()
            field_no = 0
            for field in fields:
                # ManyToMany, reverse relations 제외
                if field.many_to_many or field.one_to_many:
                    continue
                if not hasattr(field, 'column'):
                    continue

                field_no += 1
                info = get_field_info(field)

                values = [
                    field_no,
                    area_name,
                    verbose_name,
                    info['column'],
                    info['type'],
                    info['length'],
                    info['key'],
                    info['default'],
                    info['notes'],
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx in (1, 6, 7):
                        cell.alignment = Alignment(horizontal='center')

                row += 1

            # unique_together 정보
            if model._meta.unique_together:
                for ut in model._meta.unique_together:
                    cell = ws.cell(row=row, column=4, value=f'UNIQUE TOGETHER: ({", ".join(ut)})')
                    cell.font = Font(italic=True, color='0000AA')
                    cell.border = thin_border
                    row += 1

            # 빈 줄
            row += 1

    # 요약 시트
    ws2 = wb.create_sheet('테이블요약')
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 40
    ws2.column_dimensions['F'].width = 15

    ws2.cell(row=1, column=1, value='AAFC Django 테이블 요약')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=16)

    summary_headers = ['No', '앱(영역)', 'DB 테이블명', 'Model 클래스', '설명', '비고']
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill = gray_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    summary_row = 4
    table_no = 0
    for app_label in app_order:
        try:
            app_config = apps.get_app_config(app_label)
        except LookupError:
            continue
        area_name = app_labels.get(app_label, app_label)
        for model in app_config.get_models():
            table_no += 1
            db_table = model._meta.db_table
            model_name = model.__name__
            verbose_name = str(model._meta.verbose_name) if model._meta.verbose_name else ''

            field_count = sum(1 for f in model._meta.get_fields()
                            if not f.many_to_many and not f.one_to_many and hasattr(f, 'column'))

            values = [table_no, area_name, db_table, model_name, verbose_name, f'{field_count}개 필드']
            for col_idx, val in enumerate(values, 1):
                cell = ws2.cell(row=summary_row, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center')
            summary_row += 1

    # 저장
    output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                               'AAFC_Django_테이블_명세서_20260214.xlsx')
    wb.save(output_path)
    print(f'테이블 명세서 생성 완료: {output_path}')
    print(f'총 {table_no}개 테이블')


if __name__ == '__main__':
    generate_spec()
