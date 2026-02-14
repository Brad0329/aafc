from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum, Count
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import (
    DailyTotalData, DailyCoachData, DailyCoachDataMonth, MonthlyData,
)
from apps.enrollment.models import Attendance


def staff_required(view_func):
    """staff 권한 체크 데코레이터"""
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_staff:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('관리자 권한이 필요합니다.')
        return view_func(request, *args, **kwargs)
    return login_required(_wrapped)


# ─── 엑셀 공통 스타일 ───
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _apply_header_style(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


# ─── 1. 전체 DATA ───
@staff_required
def total_data_view(request):
    qs = DailyTotalData.objects.all()

    proc_dt = request.GET.get('proc_dt', '')
    sta_name = request.GET.get('sta_name', '')
    pay_stats = request.GET.get('pay_stats', '')
    pay_method = request.GET.get('pay_method', '')
    course_ym = request.GET.get('course_ym', '')
    keyword = request.GET.get('keyword', '')

    if proc_dt:
        qs = qs.filter(proc_dt__startswith=proc_dt)
    if sta_name:
        qs = qs.filter(sta_name__icontains=sta_name)
    if pay_stats:
        qs = qs.filter(pay_stats=pay_stats)
    if pay_method:
        qs = qs.filter(pay_method=pay_method)
    if course_ym:
        qs = qs.filter(course_ym=course_ym)
    if keyword:
        from django.db.models import Q
        qs = qs.filter(Q(member_name__icontains=keyword) | Q(child_name__icontains=keyword))

    qs = qs.order_by('-id')

    # 요약 통계
    stats = qs.aggregate(
        total_count=Count('id'),
        total_pay=Sum('pay_price'),
    )

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    # 필터 옵션
    pay_stats_list = DailyTotalData.objects.values_list('pay_stats', flat=True).distinct().order_by('pay_stats')
    pay_method_list = DailyTotalData.objects.values_list('pay_method', flat=True).distinct().order_by('pay_method')

    return render(request, 'reports/total_data.html', {
        'page_obj': page_obj,
        'stats': stats,
        'proc_dt': proc_dt,
        'sta_name': sta_name,
        'pay_stats': pay_stats,
        'pay_method': pay_method,
        'course_ym': course_ym,
        'keyword': keyword,
        'pay_stats_list': pay_stats_list,
        'pay_method_list': pay_method_list,
    })


@staff_required
def total_data_excel(request):
    qs = DailyTotalData.objects.all()

    proc_dt = request.GET.get('proc_dt', '')
    sta_name = request.GET.get('sta_name', '')
    pay_stats = request.GET.get('pay_stats', '')
    pay_method = request.GET.get('pay_method', '')
    course_ym = request.GET.get('course_ym', '')

    if proc_dt:
        qs = qs.filter(proc_dt__startswith=proc_dt)
    if sta_name:
        qs = qs.filter(sta_name__icontains=sta_name)
    if pay_stats:
        qs = qs.filter(pay_stats=pay_stats)
    if pay_method:
        qs = qs.filter(pay_method=pay_method)
    if course_ym:
        qs = qs.filter(course_ym=course_ym)

    qs = qs.order_by('-id')[:10000]  # 최대 1만건

    wb = Workbook()
    ws = wb.active
    ws.title = '전체DATA'

    headers = ['처리일', '학부모명', '자녀명', '연락처', '구장', '강좌',
               '코치', '결제상태', '결제방법', '결제금액', '수강년월', '신청구분']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    for i, item in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=item.proc_dt)
        ws.cell(row=i, column=2, value=item.member_name)
        ws.cell(row=i, column=3, value=item.child_name)
        ws.cell(row=i, column=4, value=item.mhtel)
        ws.cell(row=i, column=5, value=item.sta_name)
        ws.cell(row=i, column=6, value=item.lecture_title)
        ws.cell(row=i, column=7, value=item.coach_name)
        ws.cell(row=i, column=8, value=item.pay_stats)
        ws.cell(row=i, column=9, value=item.pay_method)
        ws.cell(row=i, column=10, value=item.pay_price)
        ws.cell(row=i, column=11, value=item.course_ym)
        ws.cell(row=i, column=12, value=item.apply_gubun)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="total_data.xlsx"'
    wb.save(response)
    return response


# ─── 2. 코치별 통계 ───
@staff_required
def coach_stats_view(request):
    qs = DailyCoachDataMonth.objects.all()

    course_ym = request.GET.get('course_ym', '')
    coach_name = request.GET.get('coach_name', '')
    sta_code = request.GET.get('sta_code', '')

    if course_ym:
        qs = qs.filter(course_ym=course_ym)
    if coach_name:
        qs = qs.filter(coach_name__icontains=coach_name)
    if sta_code:
        qs = qs.filter(sta_code=sta_code)

    # 코치별 집계
    coach_summary = qs.values('course_ym', 'coach_name').annotate(
        total_cnt=Count('id'),
        sum_cl=Sum('cl_cnt'),
        sum_m1001=Sum('m1001_price'),
        sum_m1002=Sum('m1002_price'),
        sum_m1003=Sum('m1003_price'),
        sum_m2001=Sum('m2001_price'),
        sum_m2002=Sum('m2002_price'),
    ).order_by('-course_ym', 'coach_name')

    paginator = Paginator(coach_summary, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'reports/coach_stats.html', {
        'page_obj': page_obj,
        'course_ym': course_ym,
        'coach_name': coach_name,
        'sta_code': sta_code,
    })


@staff_required
def coach_stats_excel(request):
    qs = DailyCoachDataMonth.objects.all()

    course_ym = request.GET.get('course_ym', '')
    coach_name = request.GET.get('coach_name', '')

    if course_ym:
        qs = qs.filter(course_ym=course_ym)
    if coach_name:
        qs = qs.filter(coach_name__icontains=coach_name)

    coach_summary = qs.values('course_ym', 'coach_name').annotate(
        total_cnt=Count('id'),
        sum_cl=Sum('cl_cnt'),
        sum_m1001=Sum('m1001_price'),
        sum_m1002=Sum('m1002_price'),
        sum_m1003=Sum('m1003_price'),
        sum_m2001=Sum('m2001_price'),
        sum_m2002=Sum('m2002_price'),
    ).order_by('-course_ym', 'coach_name')[:5000]

    wb = Workbook()
    ws = wb.active
    ws.title = '코치별통계'

    headers = ['수강년월', '코치', '회원수', '수업횟수', '수업료', '프로모션', '상품비', '교육용품1', '교육용품2']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    for i, item in enumerate(coach_summary, 2):
        ws.cell(row=i, column=1, value=item['course_ym'])
        ws.cell(row=i, column=2, value=item['coach_name'])
        ws.cell(row=i, column=3, value=item['total_cnt'])
        ws.cell(row=i, column=4, value=item['sum_cl'])
        ws.cell(row=i, column=5, value=item['sum_m1001'])
        ws.cell(row=i, column=6, value=item['sum_m1002'])
        ws.cell(row=i, column=7, value=item['sum_m1003'])
        ws.cell(row=i, column=8, value=item['sum_m2001'])
        ws.cell(row=i, column=9, value=item['sum_m2002'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="coach_stats.xlsx"'
    wb.save(response)
    return response


# ─── 3. 출석 통계 ───
@staff_required
def attendance_report_view(request):
    qs = Attendance.objects.all()

    app_month = request.GET.get('app_month', '')
    sta_code = request.GET.get('sta_code', '')
    lecture_code = request.GET.get('lecture_code', '')
    child_id = request.GET.get('child_id', '')

    if app_month:
        qs = qs.filter(app_month=app_month)
    if sta_code:
        qs = qs.filter(sta_code=sta_code)
    if lecture_code:
        qs = qs.filter(lecture_code=lecture_code)
    if child_id:
        qs = qs.filter(child_id__icontains=child_id)

    qs = qs.order_by('-attendance_dt', '-id')

    # 출석 구분별 통계
    gbn_stats = qs.values('attendance_gbn').annotate(cnt=Count('id')).order_by('attendance_gbn')

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'reports/attendance_report.html', {
        'page_obj': page_obj,
        'gbn_stats': gbn_stats,
        'app_month': app_month,
        'sta_code': sta_code,
        'lecture_code': lecture_code,
        'child_id': child_id,
    })


@staff_required
def attendance_report_excel(request):
    qs = Attendance.objects.all()

    app_month = request.GET.get('app_month', '')
    sta_code = request.GET.get('sta_code', '')
    lecture_code = request.GET.get('lecture_code', '')

    if app_month:
        qs = qs.filter(app_month=app_month)
    if sta_code:
        qs = qs.filter(sta_code=sta_code)
    if lecture_code:
        qs = qs.filter(lecture_code=lecture_code)

    qs = qs.order_by('-attendance_dt', '-id')[:10000]

    wb = Workbook()
    ws = wb.active
    ws.title = '출석현황'

    headers = ['출석일', '구장코드', '강좌코드', '자녀ID', '출석구분', '비고', '적용월', '완료']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    gbn_map = {'Y': '출석', 'N': '결석', 'A': '보강', 'R': '우천취소', 'D': '수업연기', 'E': '출결제외'}

    for i, item in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=item.attendance_dt)
        ws.cell(row=i, column=2, value=item.sta_code)
        ws.cell(row=i, column=3, value=item.lecture_code)
        ws.cell(row=i, column=4, value=item.child_id)
        ws.cell(row=i, column=5, value=gbn_map.get(item.attendance_gbn, item.attendance_gbn))
        ws.cell(row=i, column=6, value=item.attendance_desc)
        ws.cell(row=i, column=7, value=item.app_month)
        ws.cell(row=i, column=8, value=item.complete_yn)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    wb.save(response)
    return response


# ─── 4. 월별 구장 통계 ───
@staff_required
def monthly_stats_view(request):
    qs = MonthlyData.objects.all()

    proc_dt = request.GET.get('proc_dt', '')
    sta_name = request.GET.get('sta_name', '')

    if proc_dt:
        qs = qs.filter(proc_dt__startswith=proc_dt)
    if sta_name:
        qs = qs.filter(sta_name__icontains=sta_name)

    qs = qs.order_by('-proc_dt', 'sta_name')

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'reports/monthly_stats.html', {
        'page_obj': page_obj,
        'proc_dt': proc_dt,
        'sta_name': sta_name,
    })


@staff_required
def monthly_stats_excel(request):
    qs = MonthlyData.objects.all()

    proc_dt = request.GET.get('proc_dt', '')
    sta_name = request.GET.get('sta_name', '')

    if proc_dt:
        qs = qs.filter(proc_dt__startswith=proc_dt)
    if sta_name:
        qs = qs.filter(sta_name__icontains=sta_name)

    qs = qs.order_by('-proc_dt', 'sta_name')[:5000]

    wb = Workbook()
    ws = wb.active
    ws.title = '월별통계'

    headers = ['처리일', '코드', '구장', '회원수', '목표', '총수업',
               '신규체험', '신규유료', '갱신체험', '갱신유료', '재입단', '전체', '수강', '수강체험', '수강유료']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    for i, item in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=item.proc_dt)
        ws.cell(row=i, column=2, value=item.code_desc)
        ws.cell(row=i, column=3, value=item.sta_name)
        ws.cell(row=i, column=4, value=item.m_cnt)
        ws.cell(row=i, column=5, value=item.goal_cnt)
        ws.cell(row=i, column=6, value=item.tocl)
        ws.cell(row=i, column=7, value=item.newT_appl_cnt)
        ws.cell(row=i, column=8, value=item.newF_appl_cnt)
        ws.cell(row=i, column=9, value=item.renewT_appl_cnt)
        ws.cell(row=i, column=10, value=item.renewF_appl_cnt)
        ws.cell(row=i, column=11, value=item.again_appl_cnt)
        ws.cell(row=i, column=12, value=item.stats_tot_cnt)
        ws.cell(row=i, column=13, value=item.stats_ln_cnt)
        ws.cell(row=i, column=14, value=item.stats_lnT_cnt)
        ws.cell(row=i, column=15, value=item.stats_lnF_cnt)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="monthly_stats.xlsx"'
    wb.save(response)
    return response
