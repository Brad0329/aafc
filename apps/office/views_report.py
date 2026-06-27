"""Phase 8-7: REPORT 관리 뷰 (권한코드 'R')"""
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connection
from django.db.models import Q, Sum, Count, Case, When, Value, IntegerField, CharField, F
from django.db.models.functions import Coalesce, Substr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.core.paginator import Paginator
from .decorators import office_login_required, office_permission_required
from apps.reports.models import MonthlyData, DailyTotalData, DailyCoachDataNew
from apps.enrollment.models import (
    Enrollment, EnrollmentCourse, EnrollmentBill, Attendance,
)
from apps.courses.models import Stadium, Coach, Lecture, StadiumCoach
from apps.accounts.models import Member, MemberChild
from apps.common.models import CodeValue
from apps.shop.models import Order, OrderItem, OrderItemOption
from apps.payments.models import PaymentKCP


# ── 공통 헬퍼 ──────────────────────────────────────────

PAY_METHOD_MAP = {
    'CARD': '카드', 'ACCT': '무통장입금', 'VACCT': '가상계좌', 'CASH': '현금',
    'R': '실시간계좌이체', 'EMART': '이마트', 'LOTTE': '은평롯데몰',
    'MUCU': '다문화입금', 'MCDO': '맥도날드입금', 'SP': '방학특강',
    'SDM': '서대문구', 'SPBA': '스포츠바우처', 'GRPY': '현장결제',
    'CMS': 'CMS', 'TA': '태화복지관', 'ESNC': '동수원엔씨백화점',
    'ZEROPAY': '제로페이', 'BENEFIT': '장학혜택', 'CAU': '중대부초입금',
    'YDP': '영등포입금',
}

PAY_STATS_MAP = {
    'PY': '결제완료', 'PP': '결제대기', 'PN': '결제취소',
    'PZ': '결제대기취소', 'PQ': '입금확인대기', 'UN': '미결제',
}

LECTURE_STATS_MAP = {
    'LY': '수강확정', 'LP': '수강예정', 'LN': '퇴단',
    'PN': '일시중지', 'LS': '중도취소',
}

APPLY_GUBUN_MAP = {
    'NEW': '신규입단', 'RENEW': '재입단', 'AGAIN': '재수강', 'RE': '재등록',
}

CSR_METHODS = {'MUCU', 'MCDO', 'SP', 'SDM', 'TA', 'CAU', 'YDP'}


def _get_pay_method_display(code):
    return PAY_METHOD_MAP.get(code or '', code or '')


def _get_pay_stats_display(code):
    return PAY_STATS_MAP.get(code or '', code or '')


def _get_lecture_stats_display(code):
    return LECTURE_STATS_MAP.get(code or '', code or '')


def _get_apply_gubun_display(code):
    return APPLY_GUBUN_MAP.get(code or '', code or '')


def _default_ym():
    """기본 조회월 YYYYMM"""
    today = date.today()
    if today.day > 22:
        d = today + relativedelta(months=1)
        return d.strftime('%Y%m')
    return today.strftime('%Y%m')


def _month_choices_ym():
    """YYYYMM 형식 월 목록 (현재월 ~ 202401, 최신순)"""
    result = []
    cur = date.today().replace(day=1)
    end = date(2024, 1, 1)
    while cur >= end:
        result.append(cur.strftime('%Y%m'))
        cur = (cur - timedelta(days=1)).replace(day=1)
    return result


def _month_choices_dash():
    """YYYY-MM 형식 월 목록 (현재월 ~ 2024-01, 최신순)"""
    result = []
    cur = date.today().replace(day=1)
    end = date(2024, 1, 1)
    while cur >= end:
        result.append(cur.strftime('%Y-%m'))
        cur = (cur - timedelta(days=1)).replace(day=1)
    return result


def _year_choices():
    """YYYY 형식 년도 목록 (현재년 ~ 2024, 최신순)"""
    return [str(y) for y in range(date.today().year, 2023, -1)]


def _query_bill_prorated(where_clause, params, group_select, group_by):
    """EnrollmentBill 월할 금액 조회 (bill_amt / lec_period).

    EnrollmentBill은 수강 전체 금액이므로 lec_period로 나눠서 월별 금액으로 변환.
    bill_code 1003(결제금액차감), 1007(주2회이상할인), 1009(차량이용료).

    Returns: dict mapping group_key → {'m1003_b': int, 'm1007_b': int, 'm1009_b': int}
    """
    sql = f"""
    SELECT {group_select},
           SUM(ROUND(COALESCE(bs.b1003, 0)::numeric / NULLIF(em.lec_period, 0), 0))::int AS m1003_b,
           SUM(ROUND(COALESCE(bs.b1007, 0)::numeric / NULLIF(em.lec_period, 0), 0))::int AS m1007_b,
           SUM(ROUND(COALESCE(bs.b1009, 0)::numeric / NULLIF(em.lec_period, 0), 0))::int AS m1009_b
    FROM (
        SELECT DISTINCT e.id, e.lec_period,
               TO_CHAR(ec.course_ym, 'YYYYMM') AS course_ym,
               CASE WHEN COALESCE(e.pay_method, '') = '' THEN 'XXXX' ELSE e.pay_method END AS pay_method,
               CASE WHEN e.pay_method IN ('CARD','R','VACCT') THEN 'YES' ELSE 'NO' END AS kcp_yn
        FROM enrollment_enrollment e
        JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
        JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
        JOIN courses_coach co ON l.coach_id = co.id
        WHERE {where_clause}
          AND e.pay_stats = 'PY' AND e.del_chk = 'N' AND ec.course_stats = 'LY'
    ) em
    LEFT JOIN (
        SELECT no_seq,
               SUM(CASE WHEN bill_code='1003' THEN bill_amt ELSE 0 END) AS b1003,
               SUM(CASE WHEN bill_code='1007' THEN bill_amt ELSE 0 END) AS b1007,
               SUM(CASE WHEN bill_code='1009' THEN bill_amt ELSE 0 END) AS b1009
        FROM enrollment_enrollmentbill
        WHERE bill_code IN ('1003','1007','1009')
        GROUP BY no_seq
    ) bs ON em.id = bs.no_seq
    GROUP BY {group_by}
    """
    result = {}
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        cols = [col[0] for col in cursor.description]
        for row in cursor.fetchall():
            d = dict(zip(cols, row))
            # group_by 컬럼 수에 따라 키 생성 (m1003_b/m1007_b/m1009_b 제외)
            key_cols = [c for c in cols if c not in ('m1003_b', 'm1007_b', 'm1009_b')]
            key = tuple(d[c] for c in key_cols) if len(key_cols) > 1 else d[key_cols[0]]
            result[key] = {
                'm1003_b': int(d['m1003_b'] or 0),
                'm1007_b': int(d['m1007_b'] or 0),
                'm1009_b': int(d['m1009_b'] or 0),
            }
    return result


def _excel_response(wb, filename):
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def _header_fill():
    return PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


def _thin_border():
    side = Side(style='thin')
    return Border(top=side, bottom=side, left=side, right=side)


# ── 1. 주간회의자료 ──────────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_weekly(request):
    """주간회의자료 (weekly_report_3) - MonthlyData 기반"""
    search_date = request.GET.get('search_date', '')
    if not search_date:
        latest = MonthlyData.objects.order_by('-proc_dt').values_list('proc_dt', flat=True).first()
        search_date = (latest or _default_ym())[:6]

    # 3개월: 전전월, 전월, 기준월
    try:
        base = datetime.strptime(search_date[:6], '%Y%m')
    except (ValueError, TypeError):
        base = datetime.today().replace(day=1)
    m0 = base.strftime('%Y%m')
    m1 = (base - relativedelta(months=1)).strftime('%Y%m')
    m2 = (base - relativedelta(months=2)).strftime('%Y%m')

    exclude_sta = ['관악AI-11센터', '신도림로꼬구장']

    # 각 월별 실제 proc_dt max 값 (헤더 표시용)
    def _get_max_proc_dt(prefix):
        v = MonthlyData.objects.filter(proc_dt__startswith=prefix).order_by('-proc_dt').values_list('proc_dt', flat=True).first()
        return v or prefix

    proc_dt_m0 = _get_max_proc_dt(m0)
    proc_dt_m1 = _get_max_proc_dt(m1)
    proc_dt_m2 = _get_max_proc_dt(m2)

    def _get_month_data(exact_proc_dt):
        """해당 월의 max(proc_dt) 하루치 데이터만 조회 (ASP 원본과 동일)"""
        qs = MonthlyData.objects.filter(
            proc_dt=exact_proc_dt
        ).exclude(sta_name__in=exclude_sta)
        rows = {}
        for r in qs:
            key = (r.code_desc or '', r.sta_name or '')
            if key not in rows:
                rows[key] = {
                    'code_desc': r.code_desc, 'sta_name': r.sta_name,
                    'goal_cnt': 0, 'm_cnt': 0, 'tocl': 0,
                    'new_cl_cnt': 0, 'again_cl_cnt': 0, 'stats_lnT_cnt': 0,
                }
            d = rows[key]
            d['goal_cnt'] += r.goal_cnt or 0
            d['m_cnt'] += r.m_cnt or 0
            d['tocl'] += r.tocl or 0
            d['new_cl_cnt'] += r.newT_appl_cnt or 0
            d['again_cl_cnt'] += (r.newF_appl_cnt or 0) + (r.renewT_appl_cnt or 0) + \
                                  (r.renewF_appl_cnt or 0) + (r.again_appl_cnt or 0)
            d['stats_lnT_cnt'] += r.stats_lnT_cnt or 0
        return rows

    data2 = _get_month_data(proc_dt_m2)
    data1 = _get_month_data(proc_dt_m1)
    data0 = _get_month_data(proc_dt_m0)

    all_keys = sorted(set(list(data2.keys()) + list(data1.keys()) + list(data0.keys())))

    rows = []
    code_subtotals = {}
    grand_total = {f'{p}_{m}': 0 for p in ['goal', 'mcnt', 'tocl', 'new', 'again', 'end']
                   for m in ['m2', 'm1', 'm0']}

    prev_code = None
    for key in all_keys:
        code_desc, sta_name = key
        d2 = data2.get(key, {})
        d1 = data1.get(key, {})
        d0 = data0.get(key, {})

        row = {
            'code_desc': code_desc, 'sta_name': sta_name,
            'goal_m2': d2.get('goal_cnt', 0), 'mcnt_m2': d2.get('m_cnt', 0),
            'tocl_m2': d2.get('tocl', 0), 'new_m2': d2.get('new_cl_cnt', 0),
            'again_m2': d2.get('again_cl_cnt', 0), 'end_m2': d2.get('stats_lnT_cnt', 0),
            'goal_m1': d1.get('goal_cnt', 0), 'mcnt_m1': d1.get('m_cnt', 0),
            'tocl_m1': d1.get('tocl', 0), 'new_m1': d1.get('new_cl_cnt', 0),
            'again_m1': d1.get('again_cl_cnt', 0), 'end_m1': d1.get('stats_lnT_cnt', 0),
            'goal_m0': d0.get('goal_cnt', 0), 'mcnt_m0': d0.get('m_cnt', 0),
            'tocl_m0': d0.get('tocl', 0), 'new_m0': d0.get('new_cl_cnt', 0),
            'again_m0': d0.get('again_cl_cnt', 0), 'end_m0': d0.get('stats_lnT_cnt', 0),
            'is_subtotal': False, 'is_total': False,
        }
        row['mcnt_diff_m1'] = row['mcnt_m1'] - row['mcnt_m2']
        row['tocl_diff_m1'] = row['tocl_m1'] - row['tocl_m2']
        row['mcnt_diff_m0'] = row['mcnt_m0'] - row['mcnt_m1']
        row['tocl_diff_m0'] = row['tocl_m0'] - row['tocl_m1']

        if code_desc != prev_code and prev_code is not None and prev_code in code_subtotals:
            sub = code_subtotals[prev_code]
            sub['is_subtotal'] = True
            sub['mcnt_diff_m1'] = sub['mcnt_m1'] - sub['mcnt_m2']
            sub['tocl_diff_m1'] = sub['tocl_m1'] - sub['tocl_m2']
            sub['mcnt_diff_m0'] = sub['mcnt_m0'] - sub['mcnt_m1']
            sub['tocl_diff_m0'] = sub['tocl_m0'] - sub['tocl_m1']
            rows.append(sub)

        if code_desc not in code_subtotals:
            code_subtotals[code_desc] = {
                'code_desc': code_desc, 'sta_name': '소계',
                **{f'{p}_{m}': 0 for p in ['goal', 'mcnt', 'tocl', 'new', 'again', 'end']
                   for m in ['m2', 'm1', 'm0']},
                'is_subtotal': False, 'is_total': False,
            }
        for p in ['goal', 'mcnt', 'tocl', 'new', 'again', 'end']:
            for m in ['m2', 'm1', 'm0']:
                k = f'{p}_{m}'
                code_subtotals[code_desc][k] += row[k]
                grand_total[k] += row[k]

        rows.append(row)
        prev_code = code_desc

    if prev_code and prev_code in code_subtotals:
        sub = code_subtotals[prev_code]
        sub['is_subtotal'] = True
        sub['mcnt_diff_m1'] = sub['mcnt_m1'] - sub['mcnt_m2']
        sub['tocl_diff_m1'] = sub['tocl_m1'] - sub['tocl_m2']
        sub['mcnt_diff_m0'] = sub['mcnt_m0'] - sub['mcnt_m1']
        sub['tocl_diff_m0'] = sub['tocl_m0'] - sub['tocl_m1']
        rows.append(sub)

    grand_total['code_desc'] = 'AAFC'
    grand_total['sta_name'] = '합계'
    grand_total['is_subtotal'] = False
    grand_total['is_total'] = True
    grand_total['mcnt_diff_m1'] = grand_total['mcnt_m1'] - grand_total['mcnt_m2']
    grand_total['tocl_diff_m1'] = grand_total['tocl_m1'] - grand_total['tocl_m2']
    grand_total['mcnt_diff_m0'] = grand_total['mcnt_m0'] - grand_total['mcnt_m1']
    grand_total['tocl_diff_m0'] = grand_total['tocl_m0'] - grand_total['tocl_m1']
    rows.append(grand_total)

    return render(request, 'ba_office/lfreport/weekly_report.html', {
        'search_date': search_date, 'rows': rows,
        'month_m2': m2, 'month_m1': m1, 'month_m0': m0,
        'proc_dt_m2': proc_dt_m2, 'proc_dt_m1': proc_dt_m1, 'proc_dt_m0': proc_dt_m0,
        'month_list': _month_choices_ym(),
    })


# ── 2. (회원)전체_DATA ──────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_total_data(request):
    """(회원)전체_DATA - 라이브 JOIN 쿼리"""
    sch = request.GET.get('sch_lecture_month', '')
    if not sch:
        sch = _default_ym()

    rows = _get_total_data(sch)

    return render(request, 'ba_office/lfreport/total_data.html', {
        'sch_lecture_month': sch, 'rows': rows, 'total_count': len(rows),
        'month_list': _month_choices_ym(),
    })


@office_login_required
@office_permission_required('R')
def report_total_data_excel(request):
    """(회원)전체_DATA Excel"""
    sch = request.GET.get('sch_lecture_month', '')
    if not sch:
        return HttpResponse('조회월을 지정해주세요.')

    rows = _get_total_data(sch)
    wb = Workbook()
    ws = wb.active
    ws.title = '회원전체DATA'
    headers = ['순번', 'NO_SEQ', '부모아이디', '부모명', '자녀아이디', '핸드폰번호',
               '자녀명', '카드번호', '입단구분', '구장', '클래스코드', '클래스명',
               '담당코치', '수업주기', '수강기간', '수강상태', '결제금액', '수업료',
               '월수강금액', '교육용품비', '셔틀이용료', '결제상태', '결제방법', '결제일자',
               '취소일자', '취소코드', '취소사유', '시작월', '종료월',
               '수강월', '수강금액', '등록자', '등록일자', '출석', '결석', '취소',
               '쇼핑몰건수', '쇼핑몰금액']
    hfill = _header_fill()
    border = _thin_border()
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hfill
        cell.border = border
        cell.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['num'], r['no_seq'], r['member_id'], r['member_name'],
                r['child_id'], r['mhtel'], r['child_name'], r['card_num'],
                r['apply_gubun_nm'], r['sta_name'], r['lecture_code'],
                r['lecture_title'], r['coach_name'],
                f"주{r['lec_cycle']}회", f"{r['lec_period']}개월",
                r['lecture_stats_nm'], r['pay_price'], r['lec_price'],
                r['lec_course_ym_amt'], r['join_price'], r['shuttle_price'],
                r['pay_stats_nm'], r['pay_method_nm'], r['pay_dt_str'],
                r['cancel_date_str'], r['cancel_code'], r['cancel_desc'],
                r['start_dt'], r['end_dt'], r['course_ym_str'],
                r['course_ym_amt'], r['insert_id'], r['insert_dt_str'],
                r['attend_y'], r['attend_n'], r['attend_c'],
                r['shop_cnt'], r['shop_price']]
        for col_idx, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=col_idx, value=v).border = border
    return _excel_response(wb, f'total_data_{sch}.xlsx')


def _get_total_data(sch_month):
    """(회원)전체_DATA 실시간 조회 - raw SQL"""
    sql = """
    SELECT ROW_NUMBER() OVER(ORDER BY s.sta_name, e.child_id) AS rnum,
           e.id AS no_seq, e.member_id, m.name AS member_name,
           e.child_id, m.phone AS mhtel,
           mc.name AS child_name, mc.card_num,
           e.apply_gubun, s.sta_name, ec.lecture_code,
           l.lecture_title, co.coach_name,
           e.lec_cycle, e.lec_period, e.lecture_stats,
           CASE WHEN e.pay_stats = 'PY' AND e.pay_method NOT IN ('MUCU') THEN e.pay_price ELSE 0 END AS pay_price,
           COALESCE(bill_sum.lec_price, 0) AS lec_price,
           COALESCE(bill_sum.join_price, 0) AS join_price,
           COALESCE(shuttle.shuttle_price, 0) AS shuttle_price,
           ec.course_ym_amt AS lec_course_ym_amt,
           e.pay_stats, e.pay_method,
           e.pay_dt, e.cancel_date, e.cancel_code, e.cancel_desc,
           e.start_dt, e.end_dt,
           ec.course_ym, ec.course_ym_amt,
           e.insert_id, e.insert_dt,
           COALESCE(att.attend_y, 0) AS attend_y,
           COALESCE(att.attend_n, 0) AS attend_n,
           COALESCE(att.attend_c, 0) AS attend_c,
           COALESCE(shop.shop_cnt, 0) AS shop_cnt,
           COALESCE(shop.shop_price, 0) AS shop_price
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    INNER JOIN accounts_member m ON e.member_id = m.username
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    INNER JOIN (
        SELECT no_seq,
               SUM(CASE WHEN LEFT(bill_code, 2) = '10' THEN bill_amt ELSE 0 END) AS lec_price,
               SUM(CASE WHEN LEFT(bill_code, 2) = '20' THEN bill_amt ELSE 0 END) AS join_price
        FROM enrollment_enrollmentbill GROUP BY no_seq
    ) bill_sum ON e.id = bill_sum.no_seq
    LEFT JOIN (
        SELECT no_seq, SUM(course_ym_amt) AS shuttle_price
        FROM enrollment_enrollmentcourse WHERE bill_code = '1009'
        GROUP BY no_seq
    ) shuttle ON e.id = shuttle.no_seq
    LEFT JOIN (
        SELECT child_id,
               COUNT(CASE WHEN attendance_gbn IN ('Y','A') THEN 1 END) AS attend_y,
               COUNT(CASE WHEN attendance_gbn = 'N' THEN 1 END) AS attend_n,
               COUNT(CASE WHEN attendance_gbn IN ('R','D','E') THEN 1 END) AS attend_c
        FROM enrollment_attendance
        WHERE attendance_dt LIKE %s
        GROUP BY child_id
    ) att ON e.child_id = att.child_id
    LEFT JOIN (
        SELECT person_uid, COUNT(*) AS shop_cnt,
               COALESCE(SUM(total_price), 0) AS shop_price
        FROM shop_order
        WHERE is_finish = 'T' AND is_confirm = 'T'
          AND TO_CHAR(reg_date, 'YYYYMM') = %s
        GROUP BY person_uid
    ) shop ON m.username = shop.person_uid
    WHERE ec.bill_code = '1001'
      AND TO_CHAR(ec.course_ym, 'YYYYMM') = %s
    ORDER BY s.sta_name, e.child_id
    LIMIT 10000
    """
    att_prefix = sch_month[:4] + '-' + sch_month[4:6] + '%'
    with connection.cursor() as cursor:
        cursor.execute(sql, [att_prefix, sch_month, sch_month])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    data = []
    for r in result:
        pay_dt = r['pay_dt']
        cancel_date = r['cancel_date']
        insert_dt = r['insert_dt']
        course_ym = r['course_ym']
        data.append({
            **r,
            'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'cancel_date_str': cancel_date.strftime('%Y-%m-%d') if cancel_date else '',
            'insert_dt_str': insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
            'course_ym_str': course_ym.strftime('%Y-%m') if course_ym else '',
            'num': r['rnum'],
        })
    return data


# ── 4. (결제월)전체_DATA ─────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_sale_list(request):
    """(결제월)전체_DATA"""
    sch = request.GET.get('sch_lecture_month', '')
    sch_pay = request.GET.get('sch_pay_stats', '')
    if not sch:
        sch = _default_ym()

    # 요약: 해당 연도의 월별 결제상태별 합계
    cur_year = sch[:4]
    summary = _get_sale_summary(cur_year)

    # 상세 리스트
    rows = _get_sale_list_data(sch, sch_pay)

    return render(request, 'ba_office/lfreport/sale_list.html', {
        'sch_lecture_month': sch, 'sch_pay_stats': sch_pay,
        'summary': summary, 'rows': rows, 'total_count': len(rows),
        'month_list': _month_choices_ym(),
    })


def _get_sale_summary(cur_year):
    """결제월 요약 - 월별 결제상태별 금액 (KST 기준)"""
    sql = """
    SELECT stand_dt,
           SUM(CASE WHEN pay_stats = 'PY' THEN pay_price ELSE 0 END) AS py_price,
           SUM(CASE WHEN pay_stats = 'PP' THEN pay_price ELSE 0 END) AS pp_price,
           SUM(CASE WHEN pay_stats = 'PQ' THEN pay_price ELSE 0 END) AS pq_price,
           SUM(CASE WHEN pay_stats = 'PN' THEN pay_price ELSE 0 END) AS pn_price,
           SUM(CASE WHEN pay_stats = 'PZ' THEN pay_price ELSE 0 END) AS pz_price
    FROM (
        SELECT pay_stats, pay_price,
               COALESCE(
                   TO_CHAR(pay_dt AT TIME ZONE 'Asia/Seoul', 'YYYYMM'),
                   TO_CHAR(insert_dt AT TIME ZONE 'Asia/Seoul', 'YYYYMM')
               ) AS stand_dt
        FROM enrollment_enrollment
        WHERE del_chk = 'N'
          AND COALESCE(
                  TO_CHAR(pay_dt AT TIME ZONE 'Asia/Seoul', 'YYYY'),
                  TO_CHAR(insert_dt AT TIME ZONE 'Asia/Seoul', 'YYYY')
              ) = %s
    ) a
    GROUP BY stand_dt
    ORDER BY stand_dt
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [cur_year])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _get_sale_list_data(sch_month, sch_pay_stats=''):
    """(결제월)전체_DATA 상세"""
    sql = """
    SELECT ROW_NUMBER() OVER(ORDER BY s.sta_name, e.child_id) AS rnum,
           e.member_id, m.name AS member_name, e.child_id,
           m.phone AS mhtel, mc.name AS child_name,
           e.apply_gubun, e.lec_cycle, e.lec_period,
           e.lecture_stats, e.pay_price, e.pay_stats, e.pay_method,
           e.pay_dt, e.cancel_date, e.cancel_code, e.cancel_desc,
           e.start_dt, e.end_dt
    FROM enrollment_enrollment e
    INNER JOIN (
        SELECT no_seq, MAX(lecture_code) AS lecture_code
        FROM enrollment_enrollmentcourse
        WHERE bill_code = '1001'
          AND TO_CHAR(course_ym, 'YYYYMM') = %s
          AND course_stats IN ('LY', 'LP', 'PN')
        GROUP BY no_seq
    ) ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN accounts_member m ON e.member_id = m.username
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    WHERE 1=1
    """
    params = [sch_month]
    if sch_pay_stats:
        sql += " AND e.pay_stats = %s"
        params.append(sch_pay_stats)
    sql += " ORDER BY s.sta_name, e.child_id LIMIT 10000"

    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    data = []
    for r in result:
        pay_dt = r['pay_dt']
        cancel_date = r['cancel_date']
        data.append({
            **r,
            'num': r['rnum'],
            'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'cancel_date_str': cancel_date.strftime('%Y-%m-%d') if cancel_date else '',
        })
    return data


@office_login_required
@office_permission_required('R')
def report_sale_list_excel(request):
    """(결제월)전체_DATA Excel"""
    sch = request.GET.get('sch_lecture_month', '')
    sch_pay = request.GET.get('sch_pay_stats', '')
    if not sch:
        return HttpResponse('조회월을 지정해주세요.')
    rows = _get_sale_list_data(sch, sch_pay)
    wb = Workbook()
    ws = wb.active
    ws.title = '결제월전체DATA'
    headers = ['순번', '부모아이디', '부모명', '자녀아이디', '핸드폰번호', '자녀명',
               '입단구분', '수업주기', '수강기간', '수강상태', '결제금액',
               '결제상태', '결제방법', '결제일자', '취소일자', '취소코드',
               '취소사유', '시작월', '종료월']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['num'], r['member_id'], r['member_name'], r['child_id'],
                r['mhtel'], r['child_name'], r['apply_gubun_nm'],
                f"주{r['lec_cycle']}회", f"{r['lec_period']}개월",
                r['lecture_stats_nm'], r['pay_price'], r['pay_stats_nm'],
                r['pay_method_nm'], r['pay_dt_str'], r['cancel_date_str'],
                r['cancel_code'], r['cancel_desc'], r['start_dt'], r['end_dt']]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'sale_list_{sch}.xlsx')


# ── 5. (결제일)전체_DATA ─────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_sale_day_list(request):
    """(결제일)전체_DATA"""
    stdt = request.GET.get('sch_lecture_stdt', '')
    eddt = request.GET.get('sch_lecture_eddt', '')
    if not stdt:
        today = date.today()
        stdt = today.strftime('%Y%m%d')
        eddt = stdt

    rows = _get_sale_day_data(stdt, eddt)

    # 종료월별 요약
    summary = {}
    for r in rows:
        end = r['end_dt'] or ''
        if end not in summary:
            summary[end] = 0
        summary[end] += r['pay_price'] or 0
    summary_list = sorted(summary.items())

    return render(request, 'ba_office/lfreport/sale_day_list.html', {
        'sch_lecture_stdt': stdt, 'sch_lecture_eddt': eddt,
        'rows': rows, 'total_count': len(rows), 'summary': summary_list,
    })


def _get_sale_day_data(stdt, eddt):
    """결제일 기준 데이터"""
    st = f"{stdt[:4]}-{stdt[4:6]}-{stdt[6:8]}"
    ed = f"{eddt[:4]}-{eddt[4:6]}-{eddt[6:8]}"
    sql = """
    SELECT ROW_NUMBER() OVER(ORDER BY s.sta_name, e.child_id) AS rnum,
           e.member_id, m.name AS member_name, e.child_id,
           m.phone AS mhtel, mc.name AS child_name,
           mc.birth AS child_birth,
           e.apply_gubun, e.lec_cycle, e.lec_period,
           e.lecture_stats, e.pay_price, e.pay_stats, e.pay_method,
           e.pay_dt, e.start_dt, e.end_dt, s.sta_name
    FROM enrollment_enrollment e
    INNER JOIN (
        SELECT no_seq, MAX(lecture_code) AS lecture_code
        FROM enrollment_enrollmentcourse
        WHERE bill_code = '1001' AND course_stats = 'LY'
        GROUP BY no_seq
    ) ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN accounts_member m ON e.member_id = m.username
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    WHERE e.pay_stats = 'PY'
      AND e.pay_dt::date BETWEEN %s AND %s
    ORDER BY s.sta_name, e.child_id
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [st, ed])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    today = date.today()
    data = []
    for r in result:
        birth = r.get('child_birth') or ''
        if birth and len(birth) >= 4:
            try:
                age = today.year - int(birth[:4])
                if age < 0:
                    age = 0
            except (ValueError, TypeError):
                age = 99
        else:
            age = 99
        tax_type = '비과세' if age < 13 else '과세'
        pay_dt = r['pay_dt']
        birth_str = f'{birth[:4]}-{birth[4:6]}-{birth[6:8]}' if birth and len(birth) >= 8 else ''
        data.append({
            **r,
            'num': r['rnum'],
            'age': age, 'tax_type': tax_type,
            'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'birth_str': birth_str,
        })
    return data


@office_login_required
@office_permission_required('R')
def report_sale_day_list_excel(request):
    """(결제일)전체_DATA Excel"""
    stdt = request.GET.get('sch_lecture_stdt', '')
    eddt = request.GET.get('sch_lecture_eddt', '')
    if not stdt:
        return HttpResponse('조회일을 지정해주세요.')
    rows = _get_sale_day_data(stdt, eddt)
    wb = Workbook()
    ws = wb.active
    ws.title = '결제일전체DATA'
    headers = ['순번', '부모아이디', '부모명', '자녀아이디', '핸드폰번호', '자녀명',
               '입단구분', '수업주기', '수강기간', '수강상태', '결제금액',
               '결제상태', '결제방법', '결제일자', '시작월', '종료월',
               '생년월일', '나이', '과세여부', '구장']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['num'], r['member_id'], r['member_name'], r['child_id'],
                r['mhtel'], r['child_name'], r['apply_gubun_nm'],
                f"주{r['lec_cycle']}회", f"{r['lec_period']}개월",
                r['lecture_stats_nm'], r['pay_price'], r['pay_stats_nm'],
                r['pay_method_nm'], r['pay_dt_str'], r['start_dt'], r['end_dt'],
                r['birth_str'], r['age'], r['tax_type'], r['sta_name']]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'sale_day_list_{stdt}.xlsx')


# ── 6. (회원)구장별 현재원 ───────────────────────────────

@office_login_required
@office_permission_required('R')
def report_now_data(request):
    """(회원)구장별 현재원"""
    selsta = request.GET.get('selsta_code', '')
    lecture_dt = request.GET.get('lecture_dt', '')
    if not lecture_dt:
        lecture_dt = _default_ym()

    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    rows = []
    if selsta:
        rows = _get_now_data(selsta, lecture_dt)

    return render(request, 'ba_office/lfreport/now_data.html', {
        'selsta_code': selsta, 'lecture_dt': lecture_dt,
        'stadiums': stadiums, 'rows': rows, 'total_count': len(rows),
        'month_list': _month_choices_ym(),
    })


def _get_now_data(sta_code, lecture_dt):
    """구장별 현재원 - 3-UNION(신규수강원/재수강/CSR)"""
    csr_methods_sql = "','".join(CSR_METHODS)
    sql = f"""
    SELECT * FROM (
        SELECT ROW_NUMBER() OVER(ORDER BY child_id) AS rnum, Y.* FROM (
            SELECT e.id AS no_seq, e.child_id, mc.name AS child_name,
                   mc.school, mc.gender, mc.birth AS child_birth,
                   m.phone AS mhtel, m.email,
                   e.apply_gubun, s.sta_name, s.sta_code,
                   cv.code_desc, ec.lecture_code, l.lecture_title,
                   co.coach_name, e.lecture_stats,
                   CASE WHEN e.pay_stats='PY' AND e.pay_method NOT IN ('{csr_methods_sql}')
                        THEN e.pay_price ELSE 0 END AS pay_price,
                   COALESCE(bs.lec_price,0) AS lec_price,
                   COALESCE(bs.join_price,0) AS join_price,
                   COALESCE(bs.total_amt,0) AS total_amt,
                   e.pay_stats, e.pay_method,
                   e.pay_dt, e.lec_period, e.lec_cycle,
                   e.start_dt, e.end_dt,
                   ec.course_ym, ec.course_ym_amt,
                   e.insert_id, e.insert_dt,
                   m.address1, m.address2,
                   CASE WHEN e.apply_gubun IN ('NEW','RENEW') AND e.pay_dt IS NOT NULL
                        AND e.lecture_stats IN ('LY','LP','PN')
                        AND e.pay_method NOT IN ('{csr_methods_sql}')
                        THEN '신규수강원'
                        WHEN e.apply_gubun = 'AGAIN'
                        AND e.lecture_stats IN ('LY','LP','PN')
                        AND e.pay_method NOT IN ('{csr_methods_sql}')
                        THEN '재수강'
                        WHEN e.lecture_stats IN ('LY','LP','PN')
                        AND e.pay_method IN ('{csr_methods_sql}')
                        THEN 'CSR'
                        ELSE NULL END AS data_gbn
            FROM enrollment_enrollment e
            INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
            INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
            INNER JOIN courses_coach co ON l.coach_id = co.id
            INNER JOIN courses_stadium s ON l.stadium_id = s.id
            INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
            INNER JOIN accounts_member m ON e.member_id = m.username
            INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
            INNER JOIN (
                SELECT no_seq,
                       SUM(CASE WHEN LEFT(bill_code,2)='10' THEN bill_amt ELSE 0 END) AS lec_price,
                       SUM(CASE WHEN LEFT(bill_code,2)='20' THEN bill_amt ELSE 0 END) AS join_price,
                       SUM(bill_amt) AS total_amt
                FROM enrollment_enrollmentbill GROUP BY no_seq
            ) bs ON e.id = bs.no_seq
            WHERE ec.bill_code = '1001'
              AND TO_CHAR(ec.course_ym, 'YYYYMM') = (
                  SELECT MIN(TO_CHAR(h.course_ym, 'YYYYMM'))
                  FROM enrollment_enrollmentcourse h
                  INNER JOIN enrollment_enrollment i ON h.no_seq = i.id
                  WHERE h.bill_code = '1001'
                    AND TO_CHAR(h.course_ym, 'YYYYMM') = %s
                    AND i.child_id = e.child_id
              )
              AND s.sta_code = %s
        ) Y WHERE Y.data_gbn IS NOT NULL
    ) Z ORDER BY Z.child_id
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [lecture_dt, sta_code])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    today = date.today()
    data = []
    for r in result:
        birth = r.get('child_birth') or ''
        if birth and len(birth) >= 4:
            try:
                age = today.year - int(birth[:4]) + 1
            except (ValueError, TypeError):
                age = 0
        else:
            age = 0
        gender_nm = '여자' if r.get('gender') == 'F' else '남자'
        is_mucu = r.get('pay_method') in CSR_METHODS
        pay_dt = r.get('pay_dt')
        insert_dt = r.get('insert_dt')
        course_ym = r.get('course_ym')
        birth_str = f'{birth[:4]}-{birth[4:6]}-{birth[6:8]}' if birth and len(birth) >= 8 else ''
        data.append({
            **r,
            'num': r['rnum'],
            'age': age, 'gender_nm': gender_nm,
            'birth_str': birth_str,
            'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'insert_dt_str': insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
            'course_ym_str': course_ym.strftime('%Y-%m') if course_ym else '',
            'display_pay_price': 0 if is_mucu else (r.get('pay_price') or 0),
            'display_lec_price': 0 if is_mucu else (r.get('lec_price') or 0),
            'display_join_price': 0 if is_mucu else (r.get('join_price') or 0),
            'display_total_amt': 0 if is_mucu else (r.get('total_amt') or 0),
        })
    return data


@office_login_required
@office_permission_required('R')
def report_now_data_excel(request):
    """(회원)구장별 현재원 Excel"""
    selsta = request.GET.get('selsta_code', '')
    lecture_dt = request.GET.get('lecture_dt', '')
    if not selsta or not lecture_dt:
        return HttpResponse('구장과 조회월을 지정해주세요.')
    rows = _get_now_data(selsta, lecture_dt)
    wb = Workbook()
    ws = wb.active
    ws.title = '구장별현재원'
    headers = ['순번', '자녀아이디', '자녀명', '학교', '성별', '생년월일', '나이',
               '전화번호', '이메일', '입단구분', '구장명', '권역', '클래스코드',
               '클래스명', '코치명', '수강상태', '결제금액', '총금액', '수업료',
               '교육용품비', '결제상태', '결제방법', '결제일자', '수강기간', '수업주기',
               '시작월', '종료월', '수강월', '수강월수강료', '등록자ID', '등록일자',
               '데이터구분', '주소1', '주소2']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['num'], r['child_id'], r['child_name'], r.get('school', ''),
                r['gender_nm'], r['birth_str'], r['age'], r['mhtel'], r.get('email', ''),
                r['apply_gubun_nm'], r['sta_name'], r.get('code_desc', ''),
                r['lecture_code'], r['lecture_title'], r['coach_name'],
                r['lecture_stats_nm'], r['display_pay_price'], r['display_total_amt'],
                r['display_lec_price'], r['display_join_price'],
                r['pay_stats_nm'], r['pay_method_nm'], r['pay_dt_str'],
                f"{r['lec_period']}개월", f"주{r['lec_cycle']}회",
                r['start_dt'], r['end_dt'], r['course_ym_str'],
                r.get('course_ym_amt', 0), r['insert_id'], r['insert_dt_str'],
                r.get('data_gbn', ''), r.get('address1', ''), r.get('address2', '')]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'now_data_{lecture_dt}.xlsx')


# ── 7. 조회월 수강 통계 ─────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_now_statics_1(request):
    """조회월 수강 통계 (WITH ROLLUP 대체)"""
    lecture_dt = request.GET.get('lecture_dt', '')
    if not lecture_dt:
        lecture_dt = _default_ym()

    sql = """
    SELECT cv.code_desc AS codegrp, cv2.code_name, s.sta_name,
           l.lecture_title, l.lecture_code,
           COALESCE(s.kapa_tot, 0) AS stu_cnt,
           COUNT(CASE WHEN e.lecture_stats = 'LY' THEN 1 END) AS sugang_yes,
           COUNT(CASE WHEN e.lecture_stats = 'LP' THEN 1 END) AS sugang_wait,
           COUNT(CASE WHEN e.lecture_stats = 'PN' THEN 1 END) AS sugang_pause,
           COUNT(CASE WHEN e.lecture_stats = 'LN' THEN 1 END) AS sugang_end
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
    LEFT JOIN common_codevalue cv2 ON s.local_code = cv2.subcode AND cv2.grpcode = 'LOCD'
    WHERE ec.bill_code = '1001'
      AND ec.course_stats IN ('LY', 'LP', 'LN', 'PN')
      AND TO_CHAR(ec.course_ym, 'YYYYMM') = %s
    GROUP BY cv.code_desc, cv2.code_name, s.sta_name,
             l.lecture_title, l.lecture_code, s.kapa_tot
    ORDER BY cv.code_desc, s.sta_name, l.lecture_code
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [lecture_dt])
        columns = [col[0] for col in cursor.description]
        raw_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Python 으로 소계/합계 계산
    rows = []
    subtotals = {}
    grand = {'stu_cnt': 0, 'sugang_yes': 0, 'sugang_wait': 0,
             'sugang_pause': 0, 'sugang_end': 0}
    prev_codegrp = None
    prev_sta = None

    for r in raw_rows:
        cg = r['codegrp'] or ''
        sn = r['sta_name'] or ''

        if prev_codegrp is not None and cg != prev_codegrp:
            if prev_codegrp in subtotals:
                rows.append({**subtotals[prev_codegrp], 'row_type': 'subtotal_code'})

        if cg not in subtotals:
            subtotals[cg] = {
                'codegrp': cg, 'code_name': '', 'sta_name': '소계',
                'lecture_title': '', 'stu_cnt': 0, 'sugang_yes': 0,
                'sugang_wait': 0, 'sugang_pause': 0, 'sugang_end': 0,
            }

        row = {**r, 'row_type': 'data'}
        rows.append(row)

        for k in ['stu_cnt', 'sugang_yes', 'sugang_wait', 'sugang_pause', 'sugang_end']:
            subtotals[cg][k] += r[k] or 0
            grand[k] += r[k] or 0

        prev_codegrp = cg

    if prev_codegrp and prev_codegrp in subtotals:
        rows.append({**subtotals[prev_codegrp], 'row_type': 'subtotal_code'})

    rows.append({
        'codegrp': '합계', 'code_name': '', 'sta_name': '',
        'lecture_title': '', 'row_type': 'total', **grand,
    })

    return render(request, 'ba_office/lfreport/now_statics_1.html', {
        'lecture_dt': lecture_dt, 'rows': rows, 'total_count': len(raw_rows),
        'month_list': _month_choices_ym(),
    })


# ── 8. 신규(재)입단 리스트 ───────────────────────────────

@office_login_required
@office_permission_required('R')
def report_new_student(request):
    """신규(재)입단 리스트"""
    p_start_dt = request.GET.get('p_start_dt', '')
    if not p_start_dt:
        p_start_dt = _default_ym()

    rows = _get_new_student_data(p_start_dt)

    return render(request, 'ba_office/lfreport/new_student.html', {
        'p_start_dt': p_start_dt, 'rows': rows, 'total_count': len(rows),
    })


def _get_new_student_data(p_start_dt):
    """신규/재입단 리스트 데이터"""
    sql = """
    SELECT * FROM (
        SELECT ROW_NUMBER() OVER(ORDER BY e.pay_dt, s.sta_name, e.child_id) AS rnum,
               e.id AS no_seq, cv.code_desc, s.sta_name, l.lecture_title,
               e.child_id, mc.name AS child_name,
               e.apply_gubun, e.lecture_stats, e.pay_price, e.pay_stats, e.pay_method,
               e.pay_dt, e.cancel_date, e.cancel_desc,
               e.lec_period, e.lec_cycle, e.start_dt, e.end_dt,
               ec.course_ym, e.insert_id, e.insert_dt,
               COALESCE(bs.lec_price, 0) AS lec_price,
               COALESCE(bs.join_price, 0) AS join_price,
               CASE WHEN e.apply_gubun = 'NEW'
                    AND TO_CHAR(e.pay_dt, 'YYYYMM') = %s THEN 'Y' ELSE 'N' END AS is_new,
               CASE WHEN e.apply_gubun = 'RENEW'
                    AND TO_CHAR(e.pay_dt, 'YYYYMM') = %s THEN 'Y' ELSE 'N' END AS is_renew
        FROM enrollment_enrollment e
        INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
        INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
        INNER JOIN courses_stadium s ON l.stadium_id = s.id
        INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
        INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
        INNER JOIN (
            SELECT no_seq,
                   SUM(CASE WHEN LEFT(bill_code,2)='10' THEN bill_amt ELSE 0 END) AS lec_price,
                   SUM(CASE WHEN LEFT(bill_code,2)='20' THEN bill_amt ELSE 0 END) AS join_price
            FROM enrollment_enrollmentbill GROUP BY no_seq
        ) bs ON e.id = bs.no_seq
        WHERE ec.bill_code = '1001'
          AND TO_CHAR(ec.course_ym, 'YYYYMM') = (
              SELECT MIN(TO_CHAR(h.course_ym, 'YYYYMM'))
              FROM enrollment_enrollmentcourse h
              INNER JOIN enrollment_enrollment i ON h.no_seq = i.id
              WHERE h.bill_code = '1001'
                AND TO_CHAR(h.course_ym, 'YYYYMM') >= %s
                AND i.child_id = e.child_id
          )
    ) X
    WHERE X.is_new = 'Y' OR X.is_renew = 'Y'
    ORDER BY X.rnum
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [p_start_dt, p_start_dt, p_start_dt])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    data = []
    for r in result:
        is_mucu = (r.get('pay_method') or '') in CSR_METHODS
        pay_dt = r.get('pay_dt')
        cancel_date = r.get('cancel_date')
        insert_dt = r.get('insert_dt')
        course_ym = r.get('course_ym')
        data.append({
            **r,
            'num': r['rnum'],
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'cancel_date_str': cancel_date.strftime('%Y-%m-%d') if cancel_date else '',
            'insert_dt_str': insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
            'course_ym_str': course_ym.strftime('%Y-%m') if course_ym else '',
            'display_pay_price': 0 if is_mucu else (r.get('pay_price') or 0),
            'display_lec_price': 0 if is_mucu else (r.get('lec_price') or 0),
            'display_join_price': 0 if is_mucu else (r.get('join_price') or 0),
        })
    return data


@office_login_required
@office_permission_required('R')
def report_new_student_excel(request):
    """신규(재)입단 리스트 Excel"""
    p = request.GET.get('p_start_dt', '')
    if not p:
        return HttpResponse('검색월을 지정해주세요.')
    rows = _get_new_student_data(p)
    wb = Workbook()
    ws = wb.active
    ws.title = '신규재입단'
    headers = ['순번', 'NO_SEQ', '필드', '구장', '클래스', '자녀아이디', '자녀명',
               '신규입단', '재입단', '수강상태', '결제금액', '수업료', '교육용품비',
               '결제상태', '결제방법', '결제일자', '취소일자', '취소사유',
               '수강기간', '수업주기', '시작월', '종료월', '수강월', '등록자', '등록일자']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['num'], r['no_seq'], r.get('code_desc', ''), r['sta_name'],
                r['lecture_title'], r['child_id'], r['child_name'],
                r['is_new'], r['is_renew'], r['lecture_stats_nm'],
                r['display_pay_price'], r['display_lec_price'], r['display_join_price'],
                r['pay_stats_nm'], r['pay_method_nm'], r['pay_dt_str'],
                r['cancel_date_str'], r.get('cancel_desc', ''),
                f"{r['lec_period']}개월", f"주{r['lec_cycle']}회",
                r['start_dt'], r['end_dt'], r['course_ym_str'],
                r['insert_id'], r['insert_dt_str']]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'new_student_{p}.xlsx')


# ── 9. 퇴단자 리스트 ────────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_end_student(request):
    """퇴단자 리스트"""
    p_start_dt = request.GET.get('p_start_dt', '')
    if not p_start_dt:
        p_start_dt = _default_ym()
    rows = _get_end_student_data(p_start_dt)
    return render(request, 'ba_office/lfreport/end_student.html', {
        'p_start_dt': p_start_dt, 'rows': rows, 'total_count': len(rows),
    })


def _get_end_student_data(p_start_dt):
    """퇴단자 데이터"""
    sql = """
    SELECT * FROM (
        SELECT ROW_NUMBER() OVER(ORDER BY e.cancel_date, s.sta_name, e.child_id) AS rnum,
               e.id AS no_seq, cv.code_desc, s.sta_name, l.lecture_title, co.coach_name,
               e.child_id, mc.name AS child_name, mc.birth AS child_birth,
               m.phone AS mhtel,
               e.apply_gubun, e.lecture_stats, e.pay_price, e.pay_stats, e.pay_method,
               e.pay_dt, e.cancel_date, e.cancel_code, e.cancel_desc,
               e.lec_period, e.lec_cycle, e.start_dt, e.end_dt,
               ec.course_ym, e.insert_id, e.insert_dt,
               COALESCE(bs.lec_price, 0) AS lec_price,
               COALESCE(bs.join_price, 0) AS join_price,
               CASE WHEN e.lecture_stats = 'LN'
                    AND TO_CHAR(e.cancel_date, 'YYYYMM') = %s THEN 'Y' ELSE 'N' END AS is_end,
               CASE WHEN e.apply_gubun IN ('NEW','RENEW') AND e.pay_dt IS NULL
                    AND e.lecture_stats = 'LN'
                    AND TO_CHAR(e.cancel_date, 'YYYYMM') = %s THEN 'Y' ELSE 'N' END AS is_end_exclude
        FROM enrollment_enrollment e
        INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
        INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
        INNER JOIN courses_coach co ON l.coach_id = co.id
        INNER JOIN courses_stadium s ON l.stadium_id = s.id AND s.use_gbn = 'Y'
        INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
        INNER JOIN accounts_member m ON e.member_id = m.username
        INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
        INNER JOIN (
            SELECT no_seq,
                   SUM(CASE WHEN LEFT(bill_code,2)='10' THEN bill_amt ELSE 0 END) AS lec_price,
                   SUM(CASE WHEN LEFT(bill_code,2)='20' THEN bill_amt ELSE 0 END) AS join_price
            FROM enrollment_enrollmentbill GROUP BY no_seq
        ) bs ON e.id = bs.no_seq
        WHERE ec.bill_code = '1001'
          AND TO_CHAR(ec.course_ym, 'YYYYMM') = (
              SELECT MIN(TO_CHAR(h.course_ym, 'YYYYMM'))
              FROM enrollment_enrollmentcourse h
              INNER JOIN enrollment_enrollment i ON h.no_seq = i.id
              WHERE h.bill_code = '1001'
                AND TO_CHAR(h.course_ym, 'YYYYMM') >= %s
                AND i.child_id = e.child_id
          )
    ) X
    WHERE X.is_end = 'Y' AND X.is_end_exclude = 'N'
    ORDER BY X.rnum, X.cancel_date DESC, X.sta_name, X.child_id
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [p_start_dt, p_start_dt, p_start_dt])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # 퇴단유형 코드맵
    resn_codes = dict(
        CodeValue.objects.filter(group_id='RESN', del_chk='N').values_list('subcode', 'code_name')
    )

    today = date.today()
    data = []
    for idx, r in enumerate(result, 1):
        birth = r.get('child_birth') or ''
        age = (today.year - int(birth[:4]) + 1) if birth and len(birth) >= 4 else 0
        cancel_code = str(r.get('cancel_code') or '')
        cancel_type = resn_codes.get(int(cancel_code), '') if cancel_code.isdigit() else ''
        is_mucu = (r.get('pay_method') or '') in CSR_METHODS
        pay_dt = r.get('pay_dt')
        cancel_date = r.get('cancel_date')
        insert_dt = r.get('insert_dt')
        course_ym = r.get('course_ym')
        data.append({
            **r,
            'num': idx, 'age': age,
            'cancel_type': cancel_type,
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'cancel_date_str': cancel_date.strftime('%Y-%m-%d') if cancel_date else '',
            'insert_dt_str': insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
            'course_ym_str': course_ym.strftime('%Y-%m') if course_ym else '',
            'display_pay_price': 0 if is_mucu else (r.get('pay_price') or 0),
            'display_lec_price': 0 if is_mucu else (r.get('lec_price') or 0),
            'display_join_price': 0 if is_mucu else (r.get('join_price') or 0),
        })
    return data


@office_login_required
@office_permission_required('R')
def report_end_student_excel(request):
    """퇴단자 리스트 Excel"""
    p = request.GET.get('p_start_dt', '')
    if not p:
        return HttpResponse('검색월을 지정해주세요.')
    rows = _get_end_student_data(p)
    wb = Workbook()
    ws = wb.active
    ws.title = '퇴단자리스트'
    headers = ['순번', 'NO_SEQ', '필드', '구장', '클래스', '코치명', '자녀아이디',
               '자녀명', '나이', '전화번호', '퇴단일자', '퇴단코드', '퇴단유형',
               '퇴단사유', '퇴단일자', '수강상태', '총결제금액', '수업료', '교육용품비',
               '결제상태', '결제방법', '결제일자', '수강기간', '수업주가',
               '시작월', '종료월', '수강월', '등록아이디', '등록일자']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        course_ym_raw = (r.get('course_ym_str') or '').replace('-', '')
        vals = [r['num'], r['no_seq'], r.get('code_desc', ''), r['sta_name'],
                r['lecture_title'], r['coach_name'], r['child_id'], r['child_name'],
                r['age'], r['mhtel'], r['cancel_date_str'], r.get('cancel_code', ''),
                r['cancel_type'], r.get('cancel_desc', ''),
                r['cancel_date_str'],
                r['lecture_stats_nm'],
                f"{r['display_pay_price']:,}원" if r['display_pay_price'] else '0원',
                f"{r['display_lec_price']:,}원" if r['display_lec_price'] else '0원',
                f"{r['display_join_price']:,}원" if r['display_join_price'] else '0원',
                r['pay_stats_nm'], r['pay_method_nm'], r['pay_dt_str'],
                r['lec_period'], r['lec_cycle'],
                r['start_dt'], r['end_dt'], course_ym_raw,
                r['insert_id'], r['insert_dt_str']]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'end_student_{p}.xlsx')


# ── 10. 퇴단자 통계 ─────────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_anal_end(request):
    """퇴단자 통계"""
    p_start_dt = request.GET.get('p_start_dt', '')
    if not p_start_dt:
        p_start_dt = _default_ym()

    end_rows = _get_end_student_data(p_start_dt)

    # cancel_code별 구장별 집계
    stats = {}
    cancel_codes = {
        66: '단순변심', 67: '이사', 68: '일시휴식', 69: '불만족',
        70: '가입오류', 81: '구장이동', 108: '날씨', 109: '여행연수',
        110: '부상', 111: '스케줄', 112: '장기연체', 113: 'CSR종료', 116: '기타',
    }
    for r in end_rows:
        key = (r.get('code_desc', ''), r.get('sta_name', ''))
        if key not in stats:
            stats[key] = {c: 0 for c in cancel_codes}
            stats[key]['total'] = 0
            stats[key]['code_desc'] = key[0]
            stats[key]['sta_name'] = key[1]
        cc = r.get('cancel_code')
        try:
            cc_int = int(cc) if cc else 0
        except (ValueError, TypeError):
            cc_int = 0
        stats[key]['total'] += 1
        if cc_int in cancel_codes:
            stats[key][cc_int] += 1
        else:
            stats[key][116] += 1  # 기타

    rows = sorted(stats.values(), key=lambda x: (x['code_desc'], x['sta_name']))

    # 합계
    totals = {c: 0 for c in cancel_codes}
    totals['total'] = 0
    for r in rows:
        totals['total'] += r['total']
        for c in cancel_codes:
            totals[c] += r[c]

    return render(request, 'ba_office/lfreport/anal_end.html', {
        'p_start_dt': p_start_dt, 'rows': rows, 'totals': totals,
        'cancel_codes': cancel_codes,
    })


@office_login_required
@office_permission_required('R')
def report_anal_end_excel(request):
    """퇴단자 통계 Excel"""
    p = request.GET.get('p_start_dt', '')
    if not p:
        return HttpResponse('검색월을 지정해주세요.')
    end_rows = _get_end_student_data(p)
    cancel_codes = {
        66: '단순변심', 67: '이사', 68: '일시휴식', 69: '불만족',
        70: '가입오류', 81: '구장이동', 108: '날씨', 109: '여행연수',
        110: '부상', 111: '스케줄', 112: '장기연체', 113: 'CSR종료', 116: '기타',
    }
    stats = {}
    for r in end_rows:
        key = (r.get('code_desc', ''), r.get('sta_name', ''))
        if key not in stats:
            stats[key] = {c: 0 for c in cancel_codes}
            stats[key]['total'] = 0
            stats[key]['code_desc'] = key[0]
            stats[key]['sta_name'] = key[1]
        cc = r.get('cancel_code')
        try:
            cc_int = int(cc) if cc else 0
        except (ValueError, TypeError):
            cc_int = 0
        stats[key]['total'] += 1
        if cc_int in cancel_codes:
            stats[key][cc_int] += 1
        else:
            stats[key][116] += 1
    rows = sorted(stats.values(), key=lambda x: (x['code_desc'], x['sta_name']))
    totals = {c: 0 for c in cancel_codes}
    totals['total'] = 0
    for r in rows:
        totals['total'] += r['total']
        for c in cancel_codes:
            totals[c] += r[c]
    wb = Workbook()
    ws = wb.active
    ws.title = '퇴단자통계'
    headers = ['필드', '구장', '총인원'] + list(cancel_codes.values())
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['code_desc'], r['sta_name'], r['total']] + [r[c] for c in cancel_codes]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    ri = len(rows) + 2
    ws.cell(row=ri, column=1, value='총계').font = Font(bold=True)
    ws.cell(row=ri, column=3, value=totals['total'])
    for ci, c in enumerate(cancel_codes, 4):
        ws.cell(row=ri, column=ci, value=totals[c])
    return _excel_response(wb, f'anal_end_{p}.xlsx')


# ── 11. 조회월 현재원 (검색+AJAX) ───────────────────────

@office_login_required
@office_permission_required('R')
def report_now_statics_2(request):
    """조회월 현재원 - 검색 폼"""
    lecture_dt = request.GET.get('sel_lecture_dt', '')
    if not lecture_dt:
        lecture_dt = _default_ym()

    # 권역(필드) 목록
    code_descs = list(
        CodeValue.objects.filter(group_id='LOCD', del_chk='N')
        .values_list('code_desc', flat=True).distinct().order_by('code_desc')
    )

    return render(request, 'ba_office/lfreport/now_statics_2.html', {
        'sel_lecture_dt': lecture_dt,
        'code_descs': code_descs,
    })


@office_login_required
@office_permission_required('R')
def report_now_statics_2_load(request):
    """조회월 현재원 - AJAX 데이터 로드"""
    lecture_dt = request.GET.get('sel_lecture_dt', '')
    group_code = request.GET.get('group_code', '')
    local_code = request.GET.get('local_code', '')
    sta_code = request.GET.get('sta_code', '')
    lecture_code = request.GET.get('lecture_code', '')
    apply_gubun = request.GET.get('apply_gubun', '')
    lecture_stats = request.GET.get('lecture_stats', '')
    pay_stats = request.GET.get('pay_stats', '')

    if not lecture_dt:
        lecture_dt = _default_ym()

    sql = """
    SELECT e.id AS no_seq, e.child_id, mc.name AS child_name,
           mc.school, mc.birth AS child_birth,
           m.phone AS mhtel, m.email, m.address1, m.address2,
           e.apply_gubun, s.sta_name, cv.code_desc,
           l.lecture_title, e.lecture_stats,
           CASE WHEN e.pay_method = 'MUCU' THEN 0 ELSE e.pay_price END AS pay_price,
           COALESCE(bs.total_amt, 0) AS total_amt,
           CASE WHEN e.pay_method = 'MUCU' THEN 0 ELSE COALESCE(bs.lec_price, 0) END AS lec_price,
           CASE WHEN e.pay_method = 'MUCU' THEN 0 ELSE COALESCE(bs.join_price, 0) END AS join_price,
           e.pay_stats, e.pay_method, e.pay_dt,
           e.lec_period, e.lec_cycle, e.start_dt, e.end_dt,
           ec.course_ym, ec.course_ym_amt, e.insert_id, e.insert_dt
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
    INNER JOIN accounts_member m ON e.member_id = m.username
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    INNER JOIN (
        SELECT no_seq,
               SUM(CASE WHEN LEFT(bill_code,2)='10' THEN bill_amt ELSE 0 END) AS lec_price,
               SUM(CASE WHEN LEFT(bill_code,2)='20' THEN bill_amt ELSE 0 END) AS join_price,
               SUM(bill_amt) AS total_amt
        FROM enrollment_enrollmentbill GROUP BY no_seq
    ) bs ON e.id = bs.no_seq
    WHERE ec.bill_code = '1001'
      AND TO_CHAR(ec.course_ym, 'YYYYMM') = %s
    """
    params = [lecture_dt]

    if group_code:
        sql += " AND cv.code_desc = %s"
        params.append(group_code)
    if local_code:
        sql += " AND s.local_code = %s"
        params.append(int(local_code))
    if sta_code:
        sql += " AND s.sta_code = %s"
        params.append(int(sta_code))
    if lecture_code:
        sql += " AND ec.lecture_code = %s"
        params.append(int(lecture_code))
    if apply_gubun:
        sql += " AND e.apply_gubun = %s"
        params.append(apply_gubun)
    if lecture_stats:
        sql += " AND e.lecture_stats = %s"
        params.append(lecture_stats)
    if pay_stats:
        sql += " AND e.pay_stats = %s"
        params.append(pay_stats)

    sql += " ORDER BY s.local_code, s.sta_code, mc.name LIMIT 10000"

    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    today = date.today()
    data = []
    for i, r in enumerate(result, 1):
        birth = r.get('child_birth') or ''
        age = (today.year - int(birth[:4]) + 1) if birth and len(birth) >= 4 else 0
        pay_dt = r.get('pay_dt')
        insert_dt = r.get('insert_dt')
        course_ym = r.get('course_ym')
        data.append({
            **r, 'num': i, 'age': age,
            'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'insert_dt_str': insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
            'course_ym_str': course_ym.strftime('%Y-%m') if course_ym else '',
        })

    return render(request, 'ba_office/lfreport/now_statics_2_data.html', {
        'rows': data, 'total_count': len(data),
    })


def ajax_report_local(request):
    """REPORT > AJAX 지역 목록 (필드/code_desc 기준)"""
    from django.http import JsonResponse
    code_desc = request.GET.get('code_desc', '')
    qs = CodeValue.objects.filter(group_id='LOCD', del_chk='N')
    if code_desc:
        qs = qs.filter(code_desc=code_desc)
    items = list(qs.order_by('code_desc', 'code_order').values('subcode', 'code_name', 'code_desc'))
    return JsonResponse({'locals': items})


def ajax_report_stadium(request):
    """REPORT > AJAX 구장 목록"""
    local_code = request.GET.get('local_code', '')
    qs = Stadium.objects.filter(use_gbn='Y')
    if local_code:
        qs = qs.filter(local_code=int(local_code))
    from django.http import JsonResponse
    return JsonResponse({
        'stadiums': list(qs.values('sta_code', 'sta_name').order_by('sta_name'))
    })


def ajax_report_course(request):
    """REPORT > AJAX 과정(강좌) 목록 (구장 기준)"""
    from django.http import JsonResponse
    sta_code = request.GET.get('sta_code', '')
    qs = Lecture.objects.all()
    if sta_code:
        qs = qs.filter(stadium__sta_code=int(sta_code))
    qs = qs.order_by('-use_gbn', 'lecture_day', 'class_gbn', 'lecture_time')
    items = []
    for lec in qs.values('lecture_code', 'lecture_title', 'use_gbn'):
        title = lec['lecture_title']
        if lec['use_gbn'] == 'N':
            title = '(종료)' + title
        items.append({'lecture_code': lec['lecture_code'], 'lecture_title': title})
    return JsonResponse({'courses': items})


# ── 12. 쇼핑몰 주문현황 ────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_order_list(request):
    """쇼핑몰 주문현황"""
    stdt = request.GET.get('p_start_dt', '')
    eddt = request.GET.get('p_end_dt', '')
    goods_title = request.GET.get('goods_title', '')
    if not stdt:
        today = date.today()
        stdt = (today - timedelta(days=30)).strftime('%Y%m%d')
        eddt = today.strftime('%Y%m%d')

    rows = _get_order_list_data(stdt, eddt, goods_title)

    return render(request, 'ba_office/lfreport/order_list.html', {
        'p_start_dt': stdt, 'p_end_dt': eddt, 'goods_title': goods_title,
        'rows': rows, 'total_count': len(rows),
    })


def _get_order_list_data(stdt, eddt, goods_title=''):
    """쇼핑몰 주문 데이터"""
    st = f"{stdt[:4]}-{stdt[4:6]}-{stdt[6:8]}"
    ed = f"{eddt[:4]}-{eddt[4:6]}-{eddt[6:8]}"

    qs = Order.objects.filter(
        insert_dt__date__gte=st, insert_dt__date__lte=ed,
    ).select_related('member').prefetch_related('items', 'items__options').order_by('-insert_dt')

    if goods_title:
        qs = qs.filter(items__goods_title__icontains=goods_title).distinct()

    data = []
    for i, o in enumerate(qs[:5000], 1):
        for item in o.items.all():
            opts = list(item.options.all().order_by('sort'))
            opt_list = [(op.title, op.item) for op in opts]
            data.append({
                'num': i, 'uid': o.id, 'order_no': o.order_no,
                'goods_title': item.goods_title,
                'price': (item.price * item.ea) + item.option_price,
                'insert_dt': o.insert_dt.strftime('%Y-%m-%d') if o.insert_dt else '',
                'pay_dt': o.pay_dt.strftime('%Y-%m-%d') if o.pay_dt else '',
                'payway': o.payway or '',
                'state': o.get_state_display() if hasattr(o, 'get_state_display') else o.state,
                'is_cancel': '취소' if o.state == 402 else '',
                'cancel_dt': o.cancel_dt.strftime('%Y-%m-%d') if hasattr(o, 'cancel_dt') and o.cancel_dt else '',
                'ord_name': o.ord_name or '',
                'ord_phone': o.ord_phone or '',
                'ord_memo': o.ord_memo or '',
                'rcv_name': o.rcv_name or '',
                'rcv_phone': o.rcv_phone or '',
                'rcv_address': f"{o.rcv_address1 or ''} {o.rcv_address2 or ''}",
                'options': opt_list,
                'child_id': o.child_id or '',
                'ea': item.ea,
            })
    return data


@office_login_required
@office_permission_required('R')
def report_order_list_excel(request):
    """쇼핑몰 주문현황 Excel"""
    stdt = request.GET.get('p_start_dt', '')
    eddt = request.GET.get('p_end_dt', '')
    goods_title = request.GET.get('goods_title', '')
    if not stdt:
        return HttpResponse('조회일을 지정해주세요.')
    rows = _get_order_list_data(stdt, eddt, goods_title)
    wb = Workbook()
    ws = wb.active
    ws.title = '쇼핑몰주문'
    headers = ['순번', '주문번호', '상품명', '가격', '주문일자', '결제일자',
               '결제방법', '취소여부', '주문자', '주문자연락처', '주문메모',
               '수령인', '수령인연락처', '수령인주소', '자녀아이디']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(rows, 1):
        vals = [r['num'], r['order_no'], r['goods_title'], r['price'],
                r['insert_dt'], r['pay_dt'], r['payway'], r['is_cancel'],
                r['ord_name'], r['ord_phone'], r['ord_memo'],
                r['rcv_name'], r['rcv_phone'], r['rcv_address'], r['child_id']]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'order_list_{stdt}_{eddt}.xlsx')


@office_login_required
@office_permission_required('R')
def report_order_list_dedup(request):
    """쇼핑몰 주문현황(중복제거)"""
    stdt = request.GET.get('p_start_dt', '')
    eddt = request.GET.get('p_end_dt', '')
    goods_title = request.GET.get('goods_title', '')
    if not stdt:
        today = date.today()
        stdt = (today - timedelta(days=30)).strftime('%Y%m%d')
        eddt = today.strftime('%Y%m%d')

    rows = _get_order_list_data(stdt, eddt, goods_title)
    # 중복제거: child_id당 1건만
    seen = set()
    dedup = []
    for r in rows:
        cid = r.get('child_id', '')
        if cid and cid in seen:
            continue
        if cid:
            seen.add(cid)
        dedup.append(r)

    return render(request, 'ba_office/lfreport/order_list.html', {
        'p_start_dt': stdt, 'p_end_dt': eddt, 'goods_title': goods_title,
        'rows': dedup, 'total_count': len(dedup), 'is_dedup': True,
    })


@office_login_required
@office_permission_required('R')
def report_order_list_dedup_excel(request):
    """쇼핑몰 주문현황(중복제거) Excel"""
    stdt = request.GET.get('p_start_dt', '')
    eddt = request.GET.get('p_end_dt', '')
    goods_title = request.GET.get('goods_title', '')
    if not stdt:
        return HttpResponse('조회일을 지정해주세요.')
    rows = _get_order_list_data(stdt, eddt, goods_title)
    seen = set()
    dedup = []
    for r in rows:
        cid = r.get('child_id', '')
        if cid and cid in seen:
            continue
        if cid:
            seen.add(cid)
        dedup.append(r)
    wb = Workbook()
    ws = wb.active
    ws.title = '쇼핑몰주문(중복제거)'
    headers = ['순번', '주문번호', '상품명', '가격', '주문일자', '결제일자',
               '결제방법', '취소여부', '주문자', '주문자연락처', '자녀아이디']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(dedup, 1):
        vals = [i, r['order_no'], r['goods_title'], r['price'],
                r['insert_dt'], r['pay_dt'], r['payway'], r['is_cancel'],
                r['ord_name'], r['ord_phone'], r['child_id']]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'order_list_dedup_{stdt}_{eddt}.xlsx')


# ── 13. 미납자관리 ──────────────────────────────────────

def _delay_base_where():
    """미납자 공통 WHERE 절"""
    return """
    WHERE ec.bill_code = '1001'
      AND TO_CHAR(ec.course_ym, 'YYYYMMDD') <= %s
      AND e.lecture_stats IN ('LY', 'LP', 'PN')
      AND e.pay_stats IN ('PP', 'PQ')
      AND cv.grpcode = 'LOCD'
    """


@office_login_required
@office_permission_required('R')
def report_delay_data(request):
    """미납자관리"""
    search_date = request.GET.get('search_date', '')
    p_st_code = request.GET.get('p_st_code', '')
    if not search_date:
        search_date = date.today().strftime('%Y%m%d')

    # 구장 목록
    stadiums = list(
        Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
        .values_list('sta_code', 'sta_name')
    )

    # 통계 테이블 (구장별 신규/재입단/재수강 집계 with ROLLUP)
    stats_sql = """
    SELECT cv.code_desc,
           s.sta_name,
           SUM(CASE WHEN e.apply_gubun = 'NEW' THEN 1 ELSE 0 END) AS new_cnt,
           SUM(CASE WHEN e.apply_gubun = 'RENEW' THEN 1 ELSE 0 END) AS renew_cnt,
           SUM(CASE WHEN e.apply_gubun = 'AGAIN' THEN 1 ELSE 0 END) AS again_cnt,
           SUM(CASE WHEN e.apply_gubun = 'NEW' THEN 1 ELSE 0 END) +
           SUM(CASE WHEN e.apply_gubun = 'RENEW' THEN 1 ELSE 0 END) +
           SUM(CASE WHEN e.apply_gubun = 'AGAIN' THEN 1 ELSE 0 END) AS total_cnt
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
    INNER JOIN accounts_member m ON e.member_id = m.username
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    """ + _delay_base_where() + """
    GROUP BY ROLLUP(cv.code_desc, s.sta_name)
    """
    with connection.cursor() as cursor:
        cursor.execute(stats_sql, [search_date])
        stats_cols = [col[0] for col in cursor.description]
        stats_rows = [dict(zip(stats_cols, row)) for row in cursor.fetchall()]

    # 상세 리스트 (구장 선택 시만)
    detail_rows = []
    if p_st_code:
        detail_sql = """
        SELECT m.name AS member_name,
               mc.name AS child_name,
               e.child_id,
               mc.card_num,
               m.phone AS mhtel,
               e.apply_gubun, s.sta_name,
               e.lec_period, e.lec_cycle, e.start_dt, e.end_dt,
               l.lecture_title,
               e.lecture_stats, e.pay_price, e.pay_stats, e.pay_method,
               e.pay_dt, e.id AS no_seq
        FROM enrollment_enrollment e
        INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
        INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
        INNER JOIN courses_stadium s ON l.stadium_id = s.id
        INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
        INNER JOIN accounts_member m ON e.member_id = m.username
        INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
        """ + _delay_base_where() + """
          AND s.sta_code = %s
        ORDER BY mc.name, e.id, l.lecture_title
        LIMIT 10000
        """
        with connection.cursor() as cursor:
            cursor.execute(detail_sql, [search_date, int(p_st_code)])
            cols = [col[0] for col in cursor.description]
            result = [dict(zip(cols, row)) for row in cursor.fetchall()]

        for r in result:
            pay_dt = r.get('pay_dt')
            detail_rows.append({
                **r,
                'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
                'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
                'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
                'pay_method_nm': _get_pay_method_display(r['pay_method']),
                'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            })

    return render(request, 'ba_office/lfreport/delay_data.html', {
        'search_date': search_date,
        'p_st_code': p_st_code,
        'stadiums': stadiums,
        'stats_rows': stats_rows,
        'rows': detail_rows,
    })


@office_login_required
@office_permission_required('R')
def report_delay_data_excel(request):
    """미납자관리 Excel"""
    search_date = request.GET.get('search_date', '')
    p_st_code = request.GET.get('p_st_code', '')
    if not search_date or not p_st_code:
        return HttpResponse('기준일자와 구장을 선택해주세요.')

    detail_sql = """
    SELECT m.name AS member_name,
           mc.name AS child_name,
           e.child_id,
           m.phone AS mhtel,
           e.apply_gubun, s.sta_name,
           e.lec_period, e.lec_cycle, e.start_dt, e.end_dt,
           l.lecture_title,
           e.lecture_stats, e.pay_price, e.pay_stats, e.pay_method,
           e.pay_dt, e.id AS no_seq
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN common_codevalue cv ON s.local_code = cv.subcode AND cv.grpcode = 'LOCD'
    INNER JOIN accounts_member m ON e.member_id = m.username
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    """ + _delay_base_where() + """
      AND s.sta_code = %s
    ORDER BY mc.name, e.id, l.lecture_title
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(detail_sql, [search_date, int(p_st_code)])
        cols = [col[0] for col in cursor.description]
        result = [dict(zip(cols, row)) for row in cursor.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = '미납자관리'
    headers = ['부모명', '자녀명', '자녀아이디', '전화번호', 'NO_SEQ',
               '입단구분', '수강상태', '수강기간', '수업주기',
               '시작월', '종료월', '결제금액', '결제상태', '클래스']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(result, 1):
        vals = [r['member_name'], r['child_name'], r['child_id'],
                r['mhtel'], r['no_seq'],
                _get_apply_gubun_display(r['apply_gubun']),
                _get_lecture_stats_display(r['lecture_stats']),
                f"{r['lec_period']}개월", f"주{r['lec_cycle']}회",
                r['start_dt'], r['end_dt'],
                r['pay_price'],
                _get_pay_stats_display(r['pay_stats']),
                r.get('lecture_title', '')]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'delay_data_{search_date}.xlsx')


# ── 14. 월별출결현황 ────────────────────────────────────

@office_login_required
@office_permission_required('R')
def report_attendance_month(request):
    """월별출결현황 - 일별 출결 건수 PIVOT"""
    lecture_dt = request.GET.get('lecture_dt', '')
    if not lecture_dt:
        lecture_dt = date.today().strftime('%Y%m')

    ym_prefix = f"{lecture_dt[:4]}-{lecture_dt[4:6]}"

    # ASP 원본과 동일: 일별(01~31) 출결 건수를 PIVOT
    day_cols = ',\n'.join([
        f"           SUM(CASE WHEN attd_dt = '{d:02d}' THEN sum_cnt ELSE 0 END) AS cnt_{d:02d}"
        for d in range(1, 32)
    ])
    sql = f"""
    SELECT s.sta_name, l.lecture_title,
{day_cols}
    FROM (
        SELECT sta_code, lecture_code,
               SUBSTRING(attendance_dt, 9, 2) AS attd_dt,
               COUNT(*) AS sum_cnt
        FROM enrollment_attendance
        WHERE attendance_dt LIKE %s
        GROUP BY sta_code, lecture_code, SUBSTRING(attendance_dt, 9, 2)
    ) sub
    INNER JOIN courses_stadium s ON sub.sta_code = s.sta_code
    INNER JOIN courses_lecture l ON sub.lecture_code = l.lecture_code
    GROUP BY s.sta_name, l.lecture_title
    ORDER BY s.sta_name, l.lecture_title
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [ym_prefix + '%'])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return render(request, 'ba_office/lfreport/attendance_month.html', {
        'lecture_dt': lecture_dt, 'rows': result,
    })


@office_login_required
@office_permission_required('R')
def report_attendance_month_excel(request):
    """월별출결현황 Excel - 일별 PIVOT"""
    lecture_dt = request.GET.get('lecture_dt', '')
    if not lecture_dt:
        return HttpResponse('조회월을 지정해주세요.')
    ym_prefix = f"{lecture_dt[:4]}-{lecture_dt[4:6]}"
    day_cols = ',\n'.join([
        f"           SUM(CASE WHEN attd_dt = '{d:02d}' THEN sum_cnt ELSE 0 END) AS cnt_{d:02d}"
        for d in range(1, 32)
    ])
    sql = f"""
    SELECT s.sta_name, l.lecture_title,
{day_cols}
    FROM (
        SELECT sta_code, lecture_code,
               SUBSTRING(attendance_dt, 9, 2) AS attd_dt,
               COUNT(*) AS sum_cnt
        FROM enrollment_attendance
        WHERE attendance_dt LIKE %s
        GROUP BY sta_code, lecture_code, SUBSTRING(attendance_dt, 9, 2)
    ) sub
    INNER JOIN courses_stadium s ON sub.sta_code = s.sta_code
    INNER JOIN courses_lecture l ON sub.lecture_code = l.lecture_code
    GROUP BY s.sta_name, l.lecture_title
    ORDER BY s.sta_name, l.lecture_title
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [ym_prefix + '%'])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]
    wb = Workbook()
    ws = wb.active
    ws.title = '월별출결'
    # ASP 원본과 동일: 구장, 수업, 01~31 (12 제외)
    day_nums = [d for d in range(1, 32) if d != 12]
    headers = ['구장', '수업'] + [f'{d:02d}' for d in day_nums]
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(result, 1):
        vals = [r['sta_name'], r['lecture_title']]
        for d in day_nums:
            v = r.get(f'cnt_{d:02d}', 0)
            vals.append(v if v else '')
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'attendance_month_{lecture_dt}.xlsx')


# ── 15. 연도별 구장별 현황 ──────────────────────────────

def _pv(me, ve, pf):
    """월별 PIVOT SUM 컬럼 생성 (01~12)"""
    return ', '.join([
        f"SUM(CASE WHEN {me} = {m} THEN {ve} ELSE 0 END) AS {pf}_{m:02d}"
        for m in range(1, 13)])


def _parse_cnt_amt(cursor, calc_avg=False):
    """cnt_01~12, amt_01~12 결과 파싱"""
    cols = [c[0] for c in cursor.description]
    rows, avg_vals, avg_sum = [], None, 0
    for raw in cursor.fetchall():
        d = dict(zip(cols, raw))
        lbl = d.get('lec_title')
        is_total = lbl is None
        name = '합계' if is_total else (lbl[1:] if lbl else '')
        cnts = [d.get(f'cnt_{m:02d}', 0) or 0 for m in range(1, 13)]
        amts = [d.get(f'amt_{m:02d}', 0) or 0 for m in range(1, 13)]
        cs, ams = sum(cnts), sum(amts)
        if is_total and calc_avg and cs > 0:
            avg_vals = [round(amts[i] / cnts[i]) if cnts[i] else 0 for i in range(12)]
            avg_sum = round(ams / cs)
        rows.append({'name': name, 'is_total': is_total,
                     'cnt': cnts, 'amt': amts, 'cnt_sum': cs, 'amt_sum': ams})
    return rows, avg_vals, avg_sum


def _parse_bill(cursor):
    """bill_label, amt_01~12 결과 파싱"""
    cols = [c[0] for c in cursor.description]
    rows = []
    for raw in cursor.fetchall():
        d = dict(zip(cols, raw))
        lbl = d.get('bill_label')
        is_total = lbl is None
        name = '합계' if is_total else (lbl[1:] if lbl else '')
        amts = [d.get(f'amt_{m:02d}', 0) or 0 for m in range(1, 13)]
        rows.append({'name': name, 'is_total': is_total, 'amt': amts, 'amt_sum': sum(amts)})
    return rows


def _parse_attd(cursor):
    """출석/결석 월별 쌍 파싱"""
    cols = [c[0] for c in cursor.description]
    rows = []
    for raw in cursor.fetchall():
        d = dict(zip(cols, raw))
        name = d['lec_title'][1:] if d.get('lec_title') else ''
        data = [(d.get(f'yattd_{m:02d}', 0) or 0, d.get(f'nattd_{m:02d}', 0) or 0)
                for m in range(1, 13)]
        rows.append({'name': name, 'data': data})
    return rows


@office_login_required
@office_permission_required('R')
def report_stadium_year(request):
    """연도별 구장별 현황 - ASP stadium_year_statics.asp"""
    syear = request.GET.get('syear', str(date.today().year))
    selst_code = request.GET.get('selst_code', '')

    stadiums = list(
        Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
        .values_list('sta_code', 'sta_name')
    )

    ctx = {
        'syear': syear, 'stadiums': stadiums, 'selst_code': selst_code,
        'sel_sta_name': '', 'coaches': [],
        'lec_rows': [], 'lec_avg': None, 'lec_avg_sum': 0,
        'join_rows': [], 'shop_row': None,
        'tot_rows': [], 'tot2_rows': [], 'tot3_rows': [],
        'attd_rows': [],
        'year_list': _year_choices(),
    }

    if not selst_code:
        return render(request, 'ba_office/lfreport/stadium_year.html', ctx)

    sta = int(selst_code)
    try:
        ctx['sel_sta_name'] = Stadium.objects.get(sta_code=sta).sta_name
    except Stadium.DoesNotExist:
        return render(request, 'ba_office/lfreport/stadium_year.html', ctx)

    byear = int(syear) - 1

    # 공통 SQL 조각
    EF = """FROM enrollment_enrollment e
            JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
            JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
            JOIN courses_stadium s ON l.stadium_id = s.id"""

    BL = """CASE ec.bill_code
                WHEN '1001' THEN 'A수업료' WHEN '1002' THEN 'B결제금액할인'
                WHEN '1003' THEN 'E수업료할인' WHEN '1007' THEN 'G주3회 할인'
                WHEN '2001' THEN 'K교육용품비' WHEN '2002' THEN 'L교육용품비할인'
                ELSE 'Z기타' END"""

    ME = "EXTRACT(MONTH FROM ec.course_ym)::integer"
    LT = "l.lecture_day::text || l.lecture_title"
    SM = "EXTRACT(MONTH FROM o.confirm_date)::integer"
    CID = f"""(SELECT DISTINCT e.child_id {EF}
               WHERE e.pay_stats = 'PY' AND s.sta_code = %s
                 AND EXTRACT(YEAR FROM ec.course_ym) = %s)"""
    AM = _pv("s_month", "amt", "amt")

    with connection.cursor() as cursor:
        # ── 1. 담당코치정보 ──
        cursor.execute("""
            SELECT c.coach_name, COALESCE(cv.code_name, '') AS code_name
            FROM courses_stadiumcoach sc
            JOIN courses_coach c ON sc.coach_id = c.id
            JOIN courses_stadium s ON sc.stadium_id = s.id
            LEFT JOIN common_codevalue cv
                ON c.coach_level ~ '^\\d+$'
                AND c.coach_level::integer = cv.subcode AND cv.grpcode = 'LEVL'
            WHERE c.use_gbn = 'Y' AND s.sta_code = %s
            ORDER BY cv.code_order
        """, [sta])
        ctx['coaches'] = [{'coach_name': r[0], 'code_name': r[1]} for r in cursor.fetchall()]

        # ── 2. 월별 수강 인원 (bill_code=1001) ──
        cursor.execute(f"""
            SELECT {LT} AS lec_title,
                   {_pv(ME, '1', 'cnt')}, {_pv(ME, 'ec.course_ym_amt', 'amt')}
            {EF}
            WHERE e.pay_stats = 'PY' AND s.sta_code = %s
              AND ec.bill_code = '1001' AND EXTRACT(YEAR FROM ec.course_ym) = %s
            GROUP BY ROLLUP({LT}) ORDER BY {LT} NULLS LAST
        """, [sta, syear])
        ctx['lec_rows'], ctx['lec_avg'], ctx['lec_avg_sum'] = _parse_cnt_amt(cursor, True)

        # ── 3. 교육용품비 (bill_code=2001) ──
        cursor.execute(f"""
            SELECT {LT} AS lec_title,
                   {_pv(ME, '1', 'cnt')}, {_pv(ME, 'ec.course_ym_amt', 'amt')}
            {EF}
            WHERE e.pay_stats = 'PY' AND s.sta_code = %s
              AND ec.bill_code = '2001' AND EXTRACT(YEAR FROM ec.course_ym) = %s
            GROUP BY ROLLUP({LT}) ORDER BY {LT} NULLS LAST
        """, [sta, syear])
        ctx['join_rows'], _, _ = _parse_cnt_amt(cursor)

        # ── 4. 쇼핑몰 이용현황 ──
        cursor.execute(f"""
            SELECT {_pv(SM, '1', 'cnt')}, {_pv(SM, 'o.total_order_price', 'amt')}
            FROM {CID} sub
            JOIN shop_order o ON sub.child_id = o.child_id
            WHERE o.is_finish = 'T' AND o.is_confirm = 'T'
              AND EXTRACT(YEAR FROM o.confirm_date) = %s
        """, [sta, syear, syear])
        row = cursor.fetchone()
        if row:
            cols = [c[0] for c in cursor.description]
            d = dict(zip(cols, row))
            cnts = [d.get(f'cnt_{m:02d}', 0) or 0 for m in range(1, 13)]
            amts = [d.get(f'amt_{m:02d}', 0) or 0 for m in range(1, 13)]
            ctx['shop_row'] = {'cnt': cnts, 'amt': amts,
                               'cnt_sum': sum(cnts), 'amt_sum': sum(amts)}

        # ── 5. 매출 수강월 기준 ──
        cursor.execute(f"""
            SELECT bill_label, {AM}
            FROM (
                SELECT {BL} AS bill_label, {ME} AS s_month, ec.course_ym_amt AS amt
                {EF}
                WHERE e.pay_stats = 'PY' AND s.sta_code = %s
                  AND EXTRACT(YEAR FROM ec.course_ym) = %s
                UNION ALL
                SELECT 'M쇼핑몰', {SM}, o.total_order_price
                FROM {CID} sub2
                JOIN shop_order o ON sub2.child_id = o.child_id
                WHERE o.is_finish = 'T' AND o.is_confirm = 'T'
                  AND EXTRACT(YEAR FROM o.confirm_date) = %s
            ) combined
            GROUP BY ROLLUP(bill_label) ORDER BY bill_label NULLS LAST
        """, [sta, syear, sta, syear, syear])
        ctx['tot_rows'] = _parse_bill(cursor)

        # ── 6. 매출 결제일 기준 ──
        # ASP: child_id 서브쿼리가 pay_dt/insert_dt 연도 기준 (CID는 course_ym 사용)
        PM = "EXTRACT(MONTH FROM COALESCE(e.pay_dt, e.insert_dt))::integer"
        CID6 = f"""(SELECT DISTINCT e.child_id {EF}
                    WHERE e.pay_stats = 'PY' AND s.sta_code = %s
                      AND EXTRACT(YEAR FROM COALESCE(e.pay_dt, e.insert_dt)) = %s)"""
        cursor.execute(f"""
            SELECT bill_label, {AM}
            FROM (
                SELECT {BL} AS bill_label, {PM} AS s_month, ec.course_ym_amt AS amt
                {EF}
                WHERE e.pay_stats = 'PY' AND s.sta_code = %s
                  AND EXTRACT(YEAR FROM COALESCE(e.pay_dt, e.insert_dt)) = %s
                UNION ALL
                SELECT 'M쇼핑몰', {SM}, o.total_order_price
                FROM {CID6} sub3
                JOIN shop_order o ON sub3.child_id = o.child_id
                WHERE o.is_finish = 'T' AND o.is_confirm = 'T'
                  AND EXTRACT(YEAR FROM o.confirm_date) = %s
            ) combined
            GROUP BY ROLLUP(bill_label) ORDER BY bill_label NULLS LAST
        """, [sta, syear, sta, syear, syear])
        ctx['tot2_rows'] = _parse_bill(cursor)

        # ── 7. 매출 정산월 기준 (25일 기준 월 조정) ──
        start_dt = f'{byear}-12-25 00:00:00'
        end_dt = f'{syear}-12-24 23:59:59'
        PDTI = "COALESCE(e.pay_dt, e.insert_dt)"
        cursor.execute(f"""
            SELECT bill_label, {AM}
            FROM (
                SELECT {BL} AS bill_label,
                       EXTRACT(MONTH FROM {PDTI})::integer AS s_month,
                       ec.course_ym_amt AS amt
                {EF}
                WHERE e.pay_stats = 'PY' AND s.sta_code = %s
                  AND e.pay_method NOT IN ('CARD','VACCT','R')
                  AND EXTRACT(YEAR FROM {PDTI}) = %s
                UNION ALL
                SELECT {BL} AS bill_label,
                       CASE WHEN EXTRACT(DAY FROM {PDTI}) < 25
                            THEN EXTRACT(MONTH FROM {PDTI})::integer
                            ELSE EXTRACT(MONTH FROM {PDTI} + INTERVAL '1 month')::integer
                       END AS s_month,
                       ec.course_ym_amt AS amt
                {EF}
                WHERE e.pay_stats = 'PY' AND s.sta_code = %s
                  AND e.pay_method IN ('CARD','VACCT','R')
                  AND {PDTI} >= %s AND {PDTI} <= %s
                UNION ALL
                SELECT 'M쇼핑몰',
                       CASE WHEN EXTRACT(DAY FROM o.confirm_date) < 25
                            THEN EXTRACT(MONTH FROM o.confirm_date)::integer
                            ELSE EXTRACT(MONTH FROM o.confirm_date + INTERVAL '1 month')::integer
                       END,
                       o.total_order_price
                FROM {CID} sub4
                JOIN shop_order o ON sub4.child_id = o.child_id
                WHERE o.is_finish = 'T' AND o.is_confirm = 'T'
                  AND o.confirm_date >= %s AND o.confirm_date <= %s
            ) combined
            GROUP BY ROLLUP(bill_label) ORDER BY bill_label NULLS LAST
        """, [sta, syear, sta, start_dt, end_dt, sta, syear, start_dt, end_dt])
        ctx['tot3_rows'] = _parse_bill(cursor)

        # ── 8. 월별 출석 집계 ──
        attd_pv = ', '.join([
            f"SUM(CASE WHEN s_month = '{m:02d}' THEN y_attd ELSE 0 END) AS yattd_{m:02d}, "
            f"SUM(CASE WHEN s_month = '{m:02d}' THEN n_attd ELSE 0 END) AS nattd_{m:02d}"
            for m in range(1, 13)])
        cursor.execute(f"""
            SELECT lec_title, {attd_pv}
            FROM (
                SELECT l.lecture_day::text || l.lecture_title AS lec_title,
                       SUBSTRING(a.attendance_dt, 6, 2) AS s_month,
                       CASE WHEN a.attendance_gbn IN ('Y','A') THEN 1 ELSE 0 END AS y_attd,
                       CASE WHEN a.attendance_gbn IN ('N','R','D') THEN 1 ELSE 0 END AS n_attd
                FROM enrollment_attendance a
                JOIN courses_lecture l ON a.lecture_code = l.lecture_code
                WHERE a.sta_code = %s AND SUBSTRING(a.attendance_dt, 1, 4) = %s
            ) sub
            GROUP BY lec_title ORDER BY lec_title
        """, [sta, syear])
        ctx['attd_rows'] = _parse_attd(cursor)

    return render(request, 'ba_office/lfreport/stadium_year.html', ctx)


@office_login_required
@office_permission_required('R')
def report_stadium_year_excel(request):
    """연도별 구장별 현황 Excel - ASP에 Excel 없음"""
    return HttpResponse('이 페이지는 Excel 다운로드를 지원하지 않습니다.')


# ── 16. 코치 실적 미반영 취소 ───────────────────────────

@office_login_required
@office_permission_required('R')
def report_coach_miban(request):
    """코치 실적 미반영 취소분 - ASP coach_month_miban.asp
    change_history + course_src(변경전) vs course(변경후) 비교하여
    수업료(1001) 금액이 다르고, 취소월 > 수강월인 건
    """
    search_date = request.GET.get('search_date', '')

    rows = []
    if search_date:
        sql = """
        SELECT a.ch_name, a.reg_dt, a.member_id, a.child_id,
               m.name AS member_name,
               mc.name AS child_name,
               s.sta_name,
               l.lecture_title AS lec_title,
               a.course_ym, a.course_ym_amt, a.course_ym_amt2,
               (a.course_ym_amt - a.course_ym_amt2) AS totsum
        FROM (
            SELECT x.ch_name, x.reg_dt, x.member_id, x.child_id,
                   x.src_seq, x.no_seq, x.pknum, x.lecture_code,
                   x.course_ym, x.course_ym_amt,
                   COALESCE(y.course_ym_amt, 0) AS course_ym_amt2
            FROM (
                SELECT b.ch_name, TO_CHAR(a.reg_dt, 'YYYYMM') AS reg_dt,
                       a.member_id, a.child_id, a.src_seq, a.no_seq,
                       b.pknum, b.lecture_code,
                       TO_CHAR(b.course_ym, 'YYYYMM') AS course_ym,
                       b.course_ym_amt
                FROM enrollment_changehistory a
                JOIN enrollment_enrollmentcoursesrc b ON a.src_seq = b.src_seq
                WHERE TO_CHAR(a.reg_dt, 'YYYYMM') = %s
                  AND b.bill_code = '1001'
            ) x LEFT OUTER JOIN (
                SELECT a.no_seq, b.pknum, b.course_ym_amt
                FROM enrollment_changehistory a
                JOIN enrollment_enrollmentcourse b ON a.no_seq = b.no_seq
                WHERE TO_CHAR(a.reg_dt, 'YYYYMM') = %s
                  AND b.bill_code = '1001'
            ) y ON x.no_seq = y.no_seq AND x.pknum = y.pknum
        ) a
        LEFT JOIN accounts_member m ON a.member_id = m.username
        LEFT JOIN accounts_memberchild mc ON a.child_id = mc.child_id
        LEFT JOIN courses_lecture l ON a.lecture_code = l.lecture_code
        LEFT JOIN courses_stadium s ON l.stadium_id = s.id
        WHERE a.course_ym_amt <> a.course_ym_amt2
          AND a.reg_dt > a.course_ym
        ORDER BY a.ch_name, a.reg_dt, a.member_id, a.child_id, a.course_ym
        """
        with connection.cursor() as cursor:
            cursor.execute(sql, [search_date, search_date])
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return render(request, 'ba_office/lfreport/coach_miban.html', {
        'search_date': search_date, 'rows': rows,
        'month_list': _month_choices_ym(),
    })


# ── 17~21. Performance (코치별 현황) ────────────────────
# [B방안] DailyCoachData/New/Month 집계 모델 대신 Enrollment 원본 직접 조회


def _query_coachdata_monthly(search_date):
    """월별 코치 데이터 조회 - Enrollment 원본 테이블에서 직접 집계.
    report_pay_master, report_month_coachdata 공통 사용."""
    sql = """
    SELECT e.id AS pay_seq, e.pay_method,
           CASE WHEN e.pay_method IN ('CARD','R','VACCT') THEN 'YES' ELSE 'NO' END AS kcp_yn,
           s.sta_code, s.sta_name, co.coach_code, co.coach_name,
           SUM(CASE WHEN ec.bill_code = '1001' THEN ec.course_ym_amt ELSE 0 END) AS m1001,
           SUM(CASE WHEN ec.bill_code = '1002' THEN ec.course_ym_amt ELSE 0 END) AS m1002,
           SUM(CASE WHEN ec.bill_code = '1003' THEN ec.course_ym_amt ELSE 0 END) AS m1003,
           0 AS m1003b,
           SUM(CASE WHEN ec.bill_code = '1006' THEN ec.course_ym_amt ELSE 0 END) AS m1006,
           SUM(CASE WHEN ec.bill_code = '1007' THEN ec.course_ym_amt ELSE 0 END) AS m1007b,
           SUM(CASE WHEN ec.bill_code = '1009' THEN ec.course_ym_amt ELSE 0 END) AS m1009b,
           SUM(CASE WHEN ec.bill_code = '2001' THEN ec.course_ym_amt ELSE 0 END) AS m2001,
           SUM(CASE WHEN ec.bill_code = '2002' THEN ec.course_ym_amt ELSE 0 END) AS m2002
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    WHERE TO_CHAR(ec.course_ym, 'YYYYMM') = %s
      AND e.pay_stats = 'PY'
      AND e.del_chk = 'N'
      AND ec.course_stats = 'LY'
    GROUP BY e.id, e.pay_method,
             s.sta_code, s.sta_name, co.coach_code, co.coach_name
    ORDER BY co.coach_name, s.sta_code
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [search_date])
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # EnrollmentBill 월할: per enrollment (pay_seq)
    bill_map = _query_bill_prorated(
        "TO_CHAR(ec.course_ym, 'YYYYMM') = %s",
        [search_date],
        "em.id",
        "em.id",
    )

    data = []
    for r in rows:
        b = bill_map.get(r['pay_seq'], {})
        r['m1003b'] = b.get('m1003_b', 0)
        r['m1007b'] = (r.get('m1007b', 0) or 0) + b.get('m1007_b', 0)
        r['m1009b'] = (r.get('m1009b', 0) or 0) + b.get('m1009_b', 0)
        tot = r['m1001'] + r['m1002'] + r['m1003'] + r['m1003b'] + r['m1006'] + r['m1007b']
        etc = r['m1009b'] + r['m2001'] + r['m2002']
        r['pay_method'] = _get_pay_method_display(r['pay_method'])
        r['cl_cnt'] = 1
        r['tot_sum'] = tot
        r['etc_sum'] = etc
        data.append(r)
    return data


@office_login_required
@office_permission_required('R')
def report_each_coachdata(request):
    """년간 코치 개별 - ASP 원본: 코치 실적 + 코치 결제방법별 (WITH ROLLUP 대체)"""
    search_date = request.GET.get('search_date', '')
    search_coach = request.GET.get('search_coach', '')

    if not search_date:
        search_date = str(date.today().year)

    coach_list = Coach.objects.filter(use_gbn='Y').order_by('coach_name').values_list('coach_name', flat=True)

    empty_ctx = {
        'search_date': search_date, 'search_coach': search_coach,
        'coach_list': coach_list, 'year_list': _year_choices(),
        'rows1': [], 'rows2': [],
    }

    if not search_coach:
        return render(request, 'ba_office/lfreport/each_coachdata.html', empty_ctx)

    # Base: GROUP BY course_ym, pay_method, kcp_yn → Table1, Table2 모두 커버
    sql = """
    SELECT TO_CHAR(ec.course_ym, 'YYYYMM') AS course_ym,
           CASE WHEN COALESCE(e.pay_method, '') = '' THEN 'XXXX' ELSE e.pay_method END AS pay_method,
           CASE WHEN e.pay_method IN ('CARD','R','VACCT') THEN 'YES' ELSE 'NO' END AS kcp_yn,
           COUNT(DISTINCT e.id) AS cl_cnt,
           SUM(CASE WHEN ec.bill_code='1001' THEN ec.course_ym_amt ELSE 0 END) AS m1001,
           SUM(CASE WHEN ec.bill_code='1002' THEN ec.course_ym_amt ELSE 0 END) AS m1002,
           SUM(CASE WHEN ec.bill_code='1003' THEN ec.course_ym_amt ELSE 0 END) AS m1003,
           SUM(CASE WHEN ec.bill_code='1006' THEN ec.course_ym_amt ELSE 0 END) AS m1006,
           SUM(CASE WHEN ec.bill_code='1007' THEN ec.course_ym_amt ELSE 0 END) AS m1007,
           SUM(CASE WHEN ec.bill_code='1009' THEN ec.course_ym_amt ELSE 0 END) AS m1009,
           SUM(CASE WHEN ec.bill_code='2001' THEN ec.course_ym_amt ELSE 0 END) AS m2001,
           SUM(CASE WHEN ec.bill_code='2002' THEN ec.course_ym_amt ELSE 0 END) AS m2002
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_coach co ON l.coach_id = co.id
    WHERE co.coach_name = %s
      AND TO_CHAR(ec.course_ym, 'YYYY') = %s
      AND e.pay_stats = 'PY'
      AND e.del_chk = 'N'
      AND ec.course_stats = 'LY'
    GROUP BY 1, 2, 3
    ORDER BY 1, 2, 3
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [search_coach, search_date])
        columns = [col[0] for col in cursor.description]
        raw = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # EnrollmentBill 월할: bill_amt / lec_period (결제금액차감, 주2회이상할인, 차량이용료)
    bill_map = _query_bill_prorated(
        "co.coach_name = %s AND TO_CHAR(ec.course_ym, 'YYYY') = %s",
        [search_coach, search_date],
        "em.course_ym, em.pay_method, em.kcp_yn",
        "em.course_ym, em.pay_method, em.kcp_yn",
    )
    # bill_map 병합: key=(course_ym, pay_method, kcp_yn)
    for r in raw:
        key = (r['course_ym'], r['pay_method'], r['kcp_yn'])
        b = bill_map.get(key, {})
        r['m1003_b'] = b.get('m1003_b', 0)
        r['m1007'] = (r.get('m1007', 0) or 0) + b.get('m1007_b', 0)  # EC(항상0) + Bill월할
        r['m1009'] = (r.get('m1009', 0) or 0) + b.get('m1009_b', 0)  # EC(항상0) + Bill월할

    PF = ['m1001', 'm1002', 'm1003', 'm1003_b', 'm1006', 'm1007', 'm1009', 'm2001', 'm2002']

    def _z():
        return {f: 0 for f in PF + ['cl_cnt']}

    def _add(dst, src):
        dst['cl_cnt'] += src.get('cl_cnt', 0) or 0
        for f in PF:
            dst[f] += src.get(f, 0) or 0

    def _row(d, ym, col2, pm_nm, style):
        # tot_sum = 수업료 관련: m1001+m1002+m1003+m1003_b+m1006+m1007
        tot = d['m1001'] + d['m1002'] + d['m1003'] + d['m1003_b'] + d['m1006'] + d['m1007']
        # etc_sum = 기타: m1009+m2001+m2002
        etc = d['m1009'] + d['m2001'] + d['m2002']
        return {
            'course_ym': ym, 'col2': col2, 'pay_method_nm': pm_nm,
            'cl_cnt': d['cl_cnt'], 'm1001': d['m1001'], 'm1002': d['m1002'],
            'm1003': d['m1003'], 'm1003_b': d['m1003_b'], 'm1007': d['m1007'],
            'm1009': d['m1009'], 'm2001': d['m2001'], 'm2002': d['m2002'],
            'tot_sum': tot, 'etc_sum': etc, 'style': style,
        }

    # ── Table 1: 코치 실적 (GROUP BY course_ym, pay_method) ──
    from collections import OrderedDict
    t1 = OrderedDict()
    for r in raw:
        key = (r['course_ym'], r['pay_method'])
        if key not in t1:
            t1[key] = _z()
        _add(t1[key], r)

    rows1, m_sub1, grand1, prev = [], OrderedDict(), _z(), None
    for (ym, pm), g in t1.items():
        if prev and prev != ym:
            rows1.append(_row(m_sub1[prev], prev, '계', '소계', 'subtotal'))
        rows1.append(_row(g, ym, search_coach, _get_pay_method_display(pm), ''))
        if ym not in m_sub1:
            m_sub1[ym] = _z()
        _add(m_sub1[ym], g)
        _add(grand1, g)
        prev = ym
    if prev:
        rows1.append(_row(m_sub1[prev], prev, '계', '소계', 'subtotal'))
    if grand1['cl_cnt'] > 0:
        rows1.append(_row(grand1, '합계', '계', '소계', 'total'))

    # ── Table 2: 코치 결제방법별 (GROUP BY course_ym, kcp_yn, pay_method) ──
    rows2, m_sub2, grand2, prev = [], OrderedDict(), _z(), None
    for r in raw:
        ym = r['course_ym']
        if prev and prev != ym:
            rows2.append(_row(m_sub2[prev], prev, '계', '소계', 'subtotal'))
        d = {f: r.get(f, 0) or 0 for f in PF}
        d['cl_cnt'] = r.get('cl_cnt', 0) or 0
        rows2.append(_row(d, ym, r['kcp_yn'], _get_pay_method_display(r['pay_method']), ''))
        if ym not in m_sub2:
            m_sub2[ym] = _z()
        _add(m_sub2[ym], r)
        _add(grand2, r)
        prev = ym
    if prev:
        rows2.append(_row(m_sub2[prev], prev, '계', '소계', 'subtotal'))
    if grand2['cl_cnt'] > 0:
        rows2.append(_row(grand2, '합계', '계', '소계', 'total'))

    return render(request, 'ba_office/lfreport/each_coachdata.html', {
        'search_date': search_date, 'search_coach': search_coach,
        'coach_list': coach_list, 'year_list': _year_choices(),
        'rows1': rows1, 'rows2': rows2,
    })


@office_login_required
@office_permission_required('R')
def report_pay_master(request):
    """월별 결제 DATA - ASP 원본 동일: Enrollment 결제 건별 상세"""
    search_date = request.GET.get('search_date', '')
    pg_yn = request.GET.get('pg_yn', '')
    search_coach = request.GET.get('search_coach', '')

    # 코치 드롭다운 (활성 코치)
    coach_list = Coach.objects.filter(use_gbn='Y').order_by('coach_name').values_list('coach_name', flat=True)

    month_list = _month_choices_dash()

    # 초기 접근 시 빈 화면
    if not search_date:
        return render(request, 'ba_office/lfreport/pay_master.html', {
            'search_date': '', 'pg_yn': '', 'search_coach': '',
            'rows': [], 'total_count': 0, 'total_price': 0,
            'coach_list': coach_list, 'month_list': month_list,
        })

    # 기본 SQL: 결제완료 + pay_price > 0 + 해당 월
    sql = """
    SELECT s.sta_name,
           '(' || co.coach_code::text || ')' || co.coach_name AS coach_name,
           l.lecture_title,
           e.id AS no_seq,
           m.name AS member_name,
           mc.name AS child_name,
           CASE WHEN EXISTS (
               SELECT 1 FROM payments_paymentkcp p
               WHERE p.pay_seq = e.id AND p.ordr_idxx IS NOT NULL AND p.ordr_idxx != ''
           ) THEN 'YES' ELSE 'NO' END AS kcp_yn,
           e.pay_method, e.pay_price,
           COALESCE(e.pay_dt, e.insert_dt) AS pay_dt
    FROM enrollment_enrollment e
    INNER JOIN (
        SELECT DISTINCT ON (no_seq) no_seq, lecture_code
        FROM enrollment_enrollmentcourse
        WHERE bill_code = '1001' AND course_stats = 'LY'
    ) ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    LEFT JOIN accounts_member m ON e.member_id = m.username
    LEFT JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    WHERE e.pay_stats = 'PY'
      AND e.del_chk = 'N'
      AND e.pay_price > 0
      AND TO_CHAR(COALESCE(e.pay_dt, e.insert_dt), 'YYYY-MM') = %s
    ORDER BY s.sta_name, l.lecture_title, co.coach_name, m.name, mc.name
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [search_date])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # PG여부 / 코치명 필터
    data = []
    total_price = 0
    for i, r in enumerate(result, 1):
        if pg_yn and r['kcp_yn'] != pg_yn:
            continue
        if search_coach and search_coach not in r['coach_name']:
            continue
        pay_dt = r.get('pay_dt')
        data.append({
            'num': len(data) + 1,
            'sta_name': r['sta_name'] or '',
            'coach_name': r['coach_name'] or '',
            'lecture_title': r['lecture_title'] or '',
            'no_seq': r['no_seq'],
            'member_name': r['member_name'] or '',
            'child_name': r['child_name'] or '',
            'kcp_yn': r['kcp_yn'],
            'pay_method': _get_pay_method_display(r['pay_method']),
            'pay_price': r['pay_price'] or 0,
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
        })
        total_price += r['pay_price'] or 0

    return render(request, 'ba_office/lfreport/pay_master.html', {
        'search_date': search_date, 'pg_yn': pg_yn,
        'search_coach': search_coach,
        'rows': data, 'total_count': len(data), 'total_price': total_price,
        'coach_list': coach_list, 'month_list': month_list,
    })


@office_login_required
@office_permission_required('R')
def report_raw_data(request):
    """수강 RAW DATA"""
    search_date = request.GET.get('search_date', '')
    if not search_date:
        search_date = _default_ym()

    sql = """
    SELECT e.id AS no_seq, e.member_id, e.child_id,
           mc.name AS child_name, s.sta_name, l.lecture_title,
           co.coach_name, e.apply_gubun, e.lecture_stats,
           e.pay_price, e.pay_stats, e.pay_method,
           e.pay_dt, e.lec_period, e.lec_cycle,
           e.start_dt, e.end_dt, ec.course_ym, ec.course_ym_amt,
           ec.bill_code, e.insert_dt
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    WHERE TO_CHAR(ec.course_ym, 'YYYYMM') = %s
    ORDER BY s.sta_name, co.coach_name, e.child_id
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [search_date])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    data = []
    for i, r in enumerate(result, 1):
        pay_dt = r.get('pay_dt')
        insert_dt = r.get('insert_dt')
        course_ym = r.get('course_ym')
        data.append({
            **r, 'num': i,
            'apply_gubun_nm': _get_apply_gubun_display(r['apply_gubun']),
            'lecture_stats_nm': _get_lecture_stats_display(r['lecture_stats']),
            'pay_stats_nm': _get_pay_stats_display(r['pay_stats']),
            'pay_method_nm': _get_pay_method_display(r['pay_method']),
            'pay_dt_str': pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            'insert_dt_str': insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
            'course_ym_str': course_ym.strftime('%Y-%m') if course_ym else '',
        })

    return render(request, 'ba_office/lfreport/raw_data.html', {
        'search_date': search_date, 'rows': data, 'total_count': len(data),
        'month_list': _month_choices_ym(),
    })


@office_login_required
@office_permission_required('R')
def report_raw_data_excel(request):
    """수강 RAW DATA Excel"""
    search_date = request.GET.get('search_date', '')
    if not search_date:
        return HttpResponse('조회월을 지정해주세요.')

    sql = """
    SELECT e.id AS no_seq, e.member_id, e.child_id,
           mc.name AS child_name, s.sta_name, l.lecture_title,
           co.coach_name, e.apply_gubun, e.lecture_stats,
           e.pay_price, e.pay_stats, e.pay_method,
           e.pay_dt, e.lec_period, e.lec_cycle,
           e.start_dt, e.end_dt, ec.course_ym, ec.course_ym_amt,
           ec.bill_code, e.insert_dt
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    INNER JOIN accounts_memberchild mc ON e.child_id = mc.child_id
    WHERE TO_CHAR(ec.course_ym, 'YYYYMM') = %s
    ORDER BY s.sta_name, co.coach_name, e.child_id
    LIMIT 10000
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [search_date])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = '수강RAW_DATA'
    headers = ['순번', 'NO_SEQ', '부모아이디', '자녀아이디', '자녀명', '구장',
               '클래스', '코치', '입단구분', '수강상태', '결제금액',
               '결제상태', '결제방법', '결제일자', '수강기간', '수업주기',
               '시작월', '종료월', '수강월', '수강금액', 'bill코드', '등록일자']
    hfill = _header_fill()
    border = _thin_border()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.border = border
        c.font = Font(bold=True)
    for i, r in enumerate(result, 1):
        pay_dt = r.get('pay_dt')
        insert_dt = r.get('insert_dt')
        course_ym = r.get('course_ym')
        vals = [
            i, r['no_seq'], r['member_id'], r['child_id'],
            r['child_name'], r['sta_name'], r['lecture_title'],
            r['coach_name'],
            _get_apply_gubun_display(r['apply_gubun']),
            _get_lecture_stats_display(r['lecture_stats']),
            r['pay_price'] or 0,
            _get_pay_stats_display(r['pay_stats']),
            _get_pay_method_display(r['pay_method']),
            pay_dt.strftime('%Y-%m-%d') if pay_dt else '',
            f"{r['lec_period']}개월", f"주{r['lec_cycle']}회",
            r['start_dt'] or '', r['end_dt'] or '',
            course_ym.strftime('%Y-%m') if course_ym else '',
            r['course_ym_amt'] or 0, r['bill_code'] or '',
            insert_dt.strftime('%Y-%m-%d') if insert_dt else '',
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=i + 1, column=ci, value=v).border = border
    return _excel_response(wb, f'raw_data_{search_date}.xlsx')


@office_login_required
@office_permission_required('R')
def report_month_coachdata(request):
    """월별 코치 전체 - ASP 원본: 실결제 일자별 + 코치 실적 WITH ROLLUP"""
    search_date = request.GET.get('search_date', '')
    if not search_date:
        search_date = _default_ym()

    # ── 기초 데이터: 건별 (enrollment × coach) ──
    base_sql = """
    SELECT e.id,
           s.sta_name, co.coach_name,
           CASE WHEN e.pay_method IN ('CARD','R','VACCT') THEN 'YES' ELSE 'NO' END AS kcp_yn,
           CASE WHEN COALESCE(e.pay_method, '') = '' THEN '미선택' ELSE e.pay_method END AS pay_method,
           TO_CHAR(COALESCE(e.pay_dt, e.insert_dt), 'YYYY-MM') AS real_dt,
           SUM(CASE WHEN ec.bill_code='1001' THEN ec.course_ym_amt ELSE 0 END) AS m1001,
           SUM(CASE WHEN ec.bill_code='1002' THEN ec.course_ym_amt ELSE 0 END) AS m1002,
           SUM(CASE WHEN ec.bill_code='1003' THEN ec.course_ym_amt ELSE 0 END) AS m1003,
           SUM(CASE WHEN ec.bill_code='1006' THEN ec.course_ym_amt ELSE 0 END) AS m1006,
           SUM(CASE WHEN ec.bill_code='2001' THEN ec.course_ym_amt ELSE 0 END) AS m2001,
           SUM(CASE WHEN ec.bill_code='2002' THEN ec.course_ym_amt ELSE 0 END) AS m2002
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    WHERE TO_CHAR(ec.course_ym, 'YYYYMM') = %s
      AND e.pay_stats = 'PY' AND e.del_chk = 'N' AND ec.course_stats = 'LY'
    GROUP BY e.id, s.sta_name, co.coach_name, e.pay_method, e.pay_dt, e.insert_dt
    ORDER BY s.sta_name, co.coach_name, 4, 5
    """
    with connection.cursor() as cursor:
        cursor.execute(base_sql, [search_date])
        columns = [col[0] for col in cursor.description]
        raw_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Bill 월할 (per enrollment)
    bill_map = _query_bill_prorated(
        "TO_CHAR(ec.course_ym, 'YYYYMM') = %s",
        [search_date], "em.id", "em.id",
    )
    for r in raw_rows:
        b = bill_map.get(r['id'], {})
        r['m1003_b'] = b.get('m1003_b', 0)
        r['m1007_b'] = b.get('m1007_b', 0)
        r['m1009_b'] = b.get('m1009_b', 0)
        for f in ('m1001', 'm1002', 'm1003', 'm1006', 'm2001', 'm2002'):
            r[f] = int(r[f] or 0)

    PF = ['m1001', 'm1002', 'm1003', 'm1003_b', 'm1006', 'm1007_b', 'm1009_b', 'm2001', 'm2002']

    def _tot(d):
        return d['m1001'] + d['m1002'] + d['m1003'] + d['m1003_b'] + d['m1006'] + d['m1007_b']

    def _etc(d):
        return d['m1009_b'] + d['m2001'] + d['m2002']

    def _z():
        return {'cl_cnt': 0, **{f: 0 for f in PF}}

    def _add(dst, src):
        dst['cl_cnt'] += 1
        for f in PF:
            dst[f] += src[f]

    # ── Table 1: 실결제 일자별 ──
    from collections import OrderedDict
    rdt_agg = OrderedDict()
    for r in raw_rows:
        rdt = r['real_dt']
        if rdt not in rdt_agg:
            rdt_agg[rdt] = _z()
        _add(rdt_agg[rdt], r)

    search_ym_dash = f"{search_date[:4]}-{search_date[4:]}" if len(search_date) == 6 else ''
    realdt_rows = []
    grand_rdt = _z()
    for rdt in sorted(rdt_agg):
        a = rdt_agg[rdt]
        style = 'current' if rdt == search_ym_dash else ''
        realdt_rows.append({
            'real_dt': rdt, 'style': style, **a,
            'tot_sum': _tot(a), 'etc_sum': _etc(a),
        })
        grand_rdt['cl_cnt'] += a['cl_cnt']
        for f in PF:
            grand_rdt[f] += a[f]
    realdt_rows.append({
        'real_dt': '합계', 'style': 'total', **grand_rdt,
        'tot_sum': _tot(grand_rdt), 'etc_sum': _etc(grand_rdt),
    })

    # ── Table 2: 코치 실적 WITH ROLLUP 시뮬레이션 ──
    coach_agg = OrderedDict()
    for r in raw_rows:
        key = (r['sta_name'], r['coach_name'], r['kcp_yn'], r['pay_method'])
        if key not in coach_agg:
            coach_agg[key] = _z()
        _add(coach_agg[key], r)

    # 소계 계산
    coach_sub = OrderedDict()   # (sta, coach) → 합산
    sta_sub = OrderedDict()     # sta → 합산
    grand = _z()
    for (sta, coach, kcp, pm), a in coach_agg.items():
        ck = (sta, coach)
        if ck not in coach_sub:
            coach_sub[ck] = _z()
        if sta not in sta_sub:
            sta_sub[sta] = _z()
        for f in PF:
            coach_sub[ck][f] += a[f]
            sta_sub[sta][f] += a[f]
            grand[f] += a[f]
        coach_sub[ck]['cl_cnt'] += a['cl_cnt']
        sta_sub[sta]['cl_cnt'] += a['cl_cnt']
        grand['cl_cnt'] += a['cl_cnt']

    def _row(sta, coach, kcp, pm, a, style):
        return {
            'sta_name': sta, 'coach_name': coach, 'kcp_yn': kcp,
            'pay_method': pm, 'style': style, **a,
            'tot_sum': _tot(a), 'etc_sum': _etc(a),
        }

    coach_rows = []
    prev_sta, prev_coach = None, None
    for (sta, coach, kcp, pm), a in coach_agg.items():
        # 코치 변경 → 이전 코치 소계
        if prev_coach is not None and (prev_sta != sta or prev_coach != coach):
            coach_rows.append(_row(prev_sta, prev_coach, '계', '소계',
                                   coach_sub[(prev_sta, prev_coach)], 'coach_sub'))
        # 구장 변경 → 이전 구장 소계
        if prev_sta is not None and prev_sta != sta:
            coach_rows.append(_row(prev_sta, '계', '계', '소계',
                                   sta_sub[prev_sta], 'sta_sub'))
        coach_rows.append(_row(sta, coach, kcp, pm, a, ''))
        prev_sta, prev_coach = sta, coach

    # 마지막 코치/구장 소계
    if prev_coach is not None:
        coach_rows.append(_row(prev_sta, prev_coach, '계', '소계',
                               coach_sub[(prev_sta, prev_coach)], 'coach_sub'))
    if prev_sta is not None:
        coach_rows.append(_row(prev_sta, '계', '계', '소계',
                               sta_sub[prev_sta], 'sta_sub'))
    if grand['cl_cnt'] > 0:
        coach_rows.append(_row('합계', '계', '계', '소계', grand, 'grand'))

    return render(request, 'ba_office/lfreport/month_coachdata.html', {
        'search_date': search_date,
        'realdt_rows': realdt_rows,
        'coach_rows': coach_rows,
        'total_count': len(raw_rows),
        'month_list': _month_choices_ym(),
    })


@office_login_required
@office_permission_required('R')
def report_year_coachdata(request):
    """년간 코치 전체 - Enrollment 원본 직접 조회 (12개월 피벗)"""
    search_date = request.GET.get('search_date', '')
    if not search_date:
        search_date = str(date.today().year)

    # 12개월 피벗 SQL 생성
    pivot_cols = []
    for m in range(1, 13):
        mm = f"{m:02d}"
        pivot_cols.append(
            f"COUNT(DISTINCT CASE WHEN EXTRACT(MONTH FROM ec.course_ym) = {m} THEN e.id END) AS cnt_{mm}"
        )
        pivot_cols.append(
            f"SUM(CASE WHEN EXTRACT(MONTH FROM ec.course_ym) = {m} "
            f"AND ec.bill_code IN ('1001','1002','1003','1006','1007') "
            f"THEN ec.course_ym_amt ELSE 0 END) AS tot_{mm}"
        )

    sql = f"""
    SELECT s.sta_code, co.coach_name,
           {', '.join(pivot_cols)}
    FROM enrollment_enrollment e
    INNER JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
    INNER JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
    INNER JOIN courses_stadium s ON l.stadium_id = s.id
    INNER JOIN courses_coach co ON l.coach_id = co.id
    WHERE TO_CHAR(ec.course_ym, 'YYYY') = %s
      AND e.pay_stats = 'PY'
      AND e.del_chk = 'N'
      AND ec.course_stats = 'LY'
    GROUP BY s.sta_code, co.coach_name
    ORDER BY s.sta_code, co.coach_name
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [search_date])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # EnrollmentBill 월할: 12개월 피벗으로 (sta_code, coach_name, month) 별 합산
    # EnrollmentBill 월할: 12개월 피벗으로 (sta_code, coach_name) 별 합산
    bill_pivot_cols = []
    for m in range(1, 13):
        mm = f"{m:02d}"
        bill_pivot_cols.append(
            f"SUM(CASE WHEN EXTRACT(MONTH FROM em.course_ym_dt) = {m} "
            f"THEN ROUND(COALESCE(bs.b_total, 0)::numeric / NULLIF(em.lec_period, 0), 0) "
            f"ELSE 0 END)::int AS bill_{mm}"
        )
    bill_sql = f"""
    SELECT em.sta_code, em.coach_name,
           {', '.join(bill_pivot_cols)}
    FROM (
        SELECT DISTINCT e.id, e.lec_period, s.sta_code, co.coach_name,
               ec.course_ym AS course_ym_dt
        FROM enrollment_enrollment e
        JOIN enrollment_enrollmentcourse ec ON e.id = ec.no_seq
        JOIN courses_lecture l ON ec.lecture_code = l.lecture_code
        JOIN courses_stadium s ON l.stadium_id = s.id
        JOIN courses_coach co ON l.coach_id = co.id
        WHERE TO_CHAR(ec.course_ym, 'YYYY') = %s
          AND e.pay_stats = 'PY' AND e.del_chk = 'N' AND ec.course_stats = 'LY'
    ) em
    LEFT JOIN (
        SELECT no_seq,
               SUM(CASE WHEN bill_code IN ('1003','1007') THEN bill_amt ELSE 0 END) AS b_total
        FROM enrollment_enrollmentbill
        WHERE bill_code IN ('1003','1007')
        GROUP BY no_seq
    ) bs ON em.id = bs.no_seq
    GROUP BY em.sta_code, em.coach_name
    """
    bill_map = {}
    with connection.cursor() as cursor:
        cursor.execute(bill_sql, [search_date])
        bcols = [col[0] for col in cursor.description]
        for row in cursor.fetchall():
            d = dict(zip(bcols, row))
            key = (d['sta_code'], d['coach_name'])
            bill_map[key] = d

    sta_map = dict(Stadium.objects.values_list('sta_code', 'sta_name'))

    data = []
    for r in result:
        r['sta_name'] = sta_map.get(r['sta_code'], '')
        # bill 월할 추가: tot_XX에 합산
        bkey = (r['sta_code'], r['coach_name'])
        bd = bill_map.get(bkey, {})
        for m in range(1, 13):
            mm = f"{m:02d}"
            bill_val = int(bd.get(f'bill_{mm}') or 0)
            r[f'tot_{mm}'] = (r.get(f'tot_{mm}') or 0) + bill_val
        data.append(r)

    return render(request, 'ba_office/lfreport/year_coachdata.html', {
        'search_date': search_date, 'rows': data, 'total_count': len(data),
        'year_list': _year_choices(),
    })


# ============================================================
# 마감 스냅샷/정산 (원본 sp_daily_* 데몬 적재본 조회)
#   기존 실시간 리포트와 별개로, 밤배치가 확정·동결한 값을 그대로 보여준다.
# ============================================================
@office_login_required
@office_permission_required('R')
def report_daily_coach(request):
    """일별 코치 정산 (확정본) — reports_dailycoachdatanew 적재본을 코치별 집계."""
    proc_dt = request.GET.get('proc_dt', '')
    course_ym = request.GET.get('course_ym', '')

    proc_list = list(DailyCoachDataNew.objects.values_list('proc_dt', flat=True)
                     .distinct().order_by('-proc_dt'))
    if not proc_dt and proc_list:
        proc_dt = proc_list[0]
    ym_list = list(DailyCoachDataNew.objects.filter(proc_dt=proc_dt)
                   .values_list('course_ym', flat=True).distinct().order_by('-course_ym')) if proc_dt else []

    rows, totals = [], None
    if proc_dt:
        qs = DailyCoachDataNew.objects.filter(proc_dt=proc_dt)
        if course_ym:
            qs = qs.filter(course_ym=course_ym)
        agg = qs.values('coach_code', 'coach_name').annotate(
            cl_cnt=Sum('cl_cnt'), m1001=Sum('m1001_price'), m1002=Sum('m1002_price'),
            m1003=Sum('m1003_price'), m1003_b=Sum('m1003_b_price'), m1006=Sum('m1006_price'),
            m1007_b=Sum('m1007_b_price'), m1009_b=Sum('m1009_b_price'),
            m2001=Sum('m2001_price'), m2002=Sum('m2002_price'),
        ).order_by('coach_name')
        rows = list(agg)
        for r in rows:
            r['tot_sum'] = r['m1001'] + r['m1002'] + r['m1003'] + r['m1003_b'] + r['m1006'] + r['m1007_b']
            r['etc_sum'] = r['m1009_b'] + r['m2001'] + r['m2002']
        t = qs.aggregate(
            cl_cnt=Sum('cl_cnt'), m1001=Sum('m1001_price'), m1002=Sum('m1002_price'),
            m1003=Sum('m1003_price'), m1003_b=Sum('m1003_b_price'), m1006=Sum('m1006_price'),
            m1007_b=Sum('m1007_b_price'), m1009_b=Sum('m1009_b_price'),
            m2001=Sum('m2001_price'), m2002=Sum('m2002_price'),
        )
        if t['cl_cnt']:
            t['tot_sum'] = (t['m1001'] or 0) + (t['m1002'] or 0) + (t['m1003'] or 0) + (t['m1003_b'] or 0) + (t['m1006'] or 0) + (t['m1007_b'] or 0)
            t['etc_sum'] = (t['m1009_b'] or 0) + (t['m2001'] or 0) + (t['m2002'] or 0)
            totals = t

    return render(request, 'ba_office/lfreport/daily_coach.html', {
        'proc_dt': proc_dt, 'course_ym': course_ym,
        'proc_list': proc_list, 'ym_list': ym_list,
        'rows': rows, 'totals': totals,
    })


@office_login_required
@office_permission_required('R')
def report_daily_snapshot(request):
    """일별 전체 수강 스냅샷 (확정본) — reports_dailytotaldata 적재본 조회."""
    proc_dt = request.GET.get('proc_dt', '')
    course_ym = request.GET.get('course_ym', '')
    page = request.GET.get('page', '1')

    proc_list = list(DailyTotalData.objects.values_list('proc_dt', flat=True)
                     .distinct().order_by('-proc_dt'))
    if not proc_dt and proc_list:
        proc_dt = proc_list[0]
    ym_list = list(DailyTotalData.objects.filter(proc_dt=proc_dt)
                   .values_list('course_ym', flat=True).distinct().order_by('-course_ym')) if proc_dt else []

    students, total_count = None, 0
    if proc_dt:
        qs = DailyTotalData.objects.filter(proc_dt=proc_dt)
        if course_ym:
            qs = qs.filter(course_ym=course_ym)
        qs = qs.order_by('rownum')
        total_count = qs.count()
        students = Paginator(qs, 50).get_page(page)

    return render(request, 'ba_office/lfreport/daily_snapshot.html', {
        'proc_dt': proc_dt, 'course_ym': course_ym,
        'proc_list': proc_list, 'ym_list': ym_list,
        'students': students, 'total_count': total_count,
    })
