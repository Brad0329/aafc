"""8-10 쇼핑몰관리 뷰 (ba_office/lfshop/)"""
import os
from datetime import date, datetime, timedelta

from django.conf import settings
from django.core.paginator import Paginator
from django.db import connection
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from apps.shop.models import (
    Category, Order, OrderDelivery, OrderItem, OrderItemOption,
    Product, ProductOption, ProductOptionItem, ProductOptionStock,
)
from .decorators import office_login_required, office_permission_required

STATE_NAMES = {
    100: '신규주문',
    200: '배송준비(입금확인)',
    301: '배송중',
    302: '배송완료',
    402: '주문취소',
}

PAYWAY_NAMES = {
    'CARD': '신용카드',
    'BANK': '계좌이체',
    'ZEROPAY': '제로페이',
}


# ============================================================
# 카테고리관리
# ============================================================

@office_login_required
@office_permission_required('S')
def category_list(request):
    parent = int(request.GET.get('parent', 0))
    cats = Category.objects.filter(cate_parent=0, del_ok='N').order_by('-insert_dt')
    for c in cats:
        c.child_cnt = Category.objects.filter(cate_parent=c.cate_code, del_ok='N').count()
    return render(request, 'ba_office/lfshop/category_list.html', {
        'cats': cats, 'parent': parent,
    })


@office_login_required
@office_permission_required('S')
def ajax_category2(request):
    parent = int(request.GET.get('parent', 0))
    cats = Category.objects.filter(cate_parent=parent, del_ok='N').order_by('-insert_dt')
    return render(request, 'ba_office/lfshop/category_depth2.html', {'cats': cats, 'parent': parent})


@office_login_required
@office_permission_required('S')
def category_write(request):
    cate_code = request.GET.get('cate_code') or request.POST.get('cate_code', '')
    parent = int(request.GET.get('parent', 0) or request.POST.get('parent', 0))
    mode = 'edit' if cate_code else 'add'

    cat = None
    parent_cat = None
    if cate_code:
        cat = get_object_or_404(Category, cate_code=int(cate_code), del_ok='N')
        parent = cat.cate_parent
    if parent:
        parent_cat = Category.objects.filter(cate_code=parent, del_ok='N').first()

    if request.method == 'POST':
        cate_name = request.POST.get('cate_name', '').strip()
        is_display = request.POST.get('is_display', 'Y')
        img_file = request.FILES.get('cate_img')
        office_user = request.session.get('office_user', {})
        admin_id = office_user.get('userid', '')

        if not cate_name:
            return render(request, 'ba_office/lfshop/category_write.html', {
                'cat': cat, 'parent': parent, 'parent_cat': parent_cat, 'mode': mode,
                'error': '분류명을 입력해주세요.'
            })

        # 이미지 저장
        img_name = cat.cate_img if cat else ''
        if img_file:
            ext = os.path.splitext(img_file.name)[1].lower()
            new_name = f"{admin_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_cate{ext}"
            save_dir = os.path.join(settings.MEDIA_ROOT, 'fcdata', 'shop_category')
            os.makedirs(save_dir, exist_ok=True)
            with open(os.path.join(save_dir, new_name), 'wb') as f:
                for chunk in img_file.chunks():
                    f.write(chunk)
            img_name = new_name

        if mode == 'add':
            # 새 cate_code 생성
            max_code = Category.objects.order_by('-cate_code').values_list('cate_code', flat=True).first() or 0
            new_code = max_code + 1
            depth = 2 if parent else 1
            Category.objects.create(
                cate_code=new_code,
                cate_name=cate_name,
                cate_depth=depth,
                cate_parent=parent,
                cate_img=img_name,
                cate_is_display=is_display,
                del_ok='N',
                insert_dt=timezone.now(),
                insert_id=admin_id,
            )
        else:
            cat.cate_name = cate_name
            cat.cate_is_display = is_display
            if img_file:
                cat.cate_img = img_name
            cat.save()

        return redirect(f"/ba_office/lfshop/category/list/?parent={parent}")

    return render(request, 'ba_office/lfshop/category_write.html', {
        'cat': cat, 'parent': parent, 'parent_cat': parent_cat, 'mode': mode,
    })


@office_login_required
@office_permission_required('S')
def category_delete(request):
    cate_code = request.POST.get('cate_code') or request.GET.get('cate_code', '')
    parent = request.POST.get('parent', 0) or request.GET.get('parent', 0)
    if cate_code:
        Category.objects.filter(cate_code=int(cate_code)).update(del_ok='Y')
    return redirect(f"/ba_office/lfshop/category/list/?parent={parent}")


# ============================================================
# 상품관리
# ============================================================

@office_login_required
@office_permission_required('S')
def goods_list(request):
    category1 = int(request.GET.get('category1', 0) or 0)
    category2 = int(request.GET.get('category2', 0) or 0)
    sstock = request.GET.get('sstock', '0')
    sdisplay = request.GET.get('sdisplay', '0')
    ssdate = request.GET.get('ssdate', '')
    sedate = request.GET.get('sedate', '')
    sword = request.GET.get('sword', '').strip()

    qs = Product.objects.filter(del_ok='N').select_related('category').order_by(
        '-order_number', '-gd_code'
    )

    # 카테고리 필터
    cate_id = category2 if category2 else category1
    if cate_id:
        qs = qs.filter(category__cate_code=cate_id)

    # 재고 필터
    if sstock == '1':
        qs = qs.exclude(gd_stock='0')
    elif sstock == '2':
        qs = qs.filter(gd_stock='0')
    elif sstock == '3':
        qs = qs.filter(gd_is_soldout='Y')

    # 진열 필터
    if sdisplay == '1':
        qs = qs.filter(gd_display='Y')
    elif sdisplay == '2':
        qs = qs.exclude(gd_display='Y')

    # 날짜 필터
    if ssdate:
        try:
            qs = qs.filter(insert_dt__date__gte=datetime.strptime(ssdate, '%Y-%m-%d').date())
        except ValueError:
            pass
    if sedate:
        try:
            qs = qs.filter(insert_dt__date__lte=datetime.strptime(sedate, '%Y-%m-%d').date())
        except ValueError:
            pass

    # 키워드 검색
    if sword:
        qs = qs.filter(gd_name__icontains=sword)

    # 1차 카테고리 목록
    cate1_list = Category.objects.filter(cate_parent=0, del_ok='N').order_by('-insert_dt')
    cate2_list = []
    if category1:
        cate2_list = list(Category.objects.filter(cate_parent=category1, del_ok='N').order_by('-insert_dt'))

    # 카테고리 경로 구성
    cate_map = {c.cate_code: c for c in Category.objects.filter(del_ok='N')}
    for p in qs:
        if p.category:
            cat = p.category
            path = cat.cate_name
            if cat.cate_parent:
                parent_cat = cate_map.get(cat.cate_parent)
                if parent_cat:
                    path = f"{parent_cat.cate_name} > {path}"
            p.cate_path = path
        else:
            p.cate_path = ''

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'ba_office/lfshop/goods_list.html', {
        'page_obj': page_obj,
        'cate1_list': cate1_list,
        'cate2_list': cate2_list,
        'category1': category1, 'category2': category2,
        'sstock': sstock, 'sdisplay': sdisplay,
        'ssdate': ssdate, 'sedate': sedate, 'sword': sword,
    })


@office_login_required
@office_permission_required('S')
def goods_modify(request, pk):
    """상품 기본정보 수정"""
    product = get_object_or_404(Product, pk=pk, del_ok='N')
    cate1_list = Category.objects.filter(cate_parent=0, del_ok='N').order_by('-insert_dt')
    cur_cate1 = 0
    cur_cate2 = 0
    cate2_list = []
    if product.category:
        if product.category.cate_parent:
            cur_cate1 = product.category.cate_parent
            cur_cate2 = product.category.cate_code
            cate2_list = list(Category.objects.filter(cate_parent=cur_cate1, del_ok='N').order_by('-insert_dt'))
        else:
            cur_cate1 = product.category.cate_code

    if request.method == 'POST':
        product.gd_name = request.POST.get('gd_name', '').strip()
        product.gd_market_price = int(request.POST.get('gd_market_price', 0) or 0)
        product.gd_price = int(request.POST.get('gd_price', 0) or 0)
        product.gd_pay_price = int(request.POST.get('gd_pay_price', 0) or 0)
        product.gd_point = int(request.POST.get('gd_point', 0) or 0)
        product.gd_stock = request.POST.get('gd_stock', '').strip()
        product.gd_display = request.POST.get('gd_display', 'Y')
        product.gd_is_soldout = request.POST.get('gd_is_soldout', 'N')
        product.order_number = request.POST.get('order_number', '').strip()

        # 카테고리 변경
        cate_code = int(request.POST.get('category2', 0) or request.POST.get('category1', 0) or 0)
        if cate_code:
            new_cat = Category.objects.filter(cate_code=cate_code, del_ok='N').first()
            product.category = new_cat

        product.save()
        return redirect(f"/ba_office/lfshop/goods/list/")

    return render(request, 'ba_office/lfshop/goods_modify.html', {
        'product': product,
        'cate1_list': cate1_list,
        'cate2_list': cate2_list,
        'cur_cate1': cur_cate1,
        'cur_cate2': cur_cate2,
    })


@office_login_required
@office_permission_required('S')
def goods_delete(request):
    gd_code = request.GET.get('gd_code', '')
    if gd_code:
        Product.objects.filter(gd_code=int(gd_code)).update(del_ok='Y')
    return redirect('/ba_office/lfshop/goods/list/')


@office_login_required
@office_permission_required('S')
def ajax_goods_cate2(request):
    """2차 카테고리 SELECT 옵션 (상품 등록용)"""
    parent = int(request.GET.get('parent', 0))
    category2 = int(request.GET.get('category2', 0))
    cats = Category.objects.filter(cate_parent=parent, del_ok='N').order_by('-insert_dt')
    return render(request, 'ba_office/lfshop/goods_cate2_select.html', {
        'cats': cats, 'category2': category2,
    })


# ============================================================
# 주문내역
# ============================================================

@office_login_required
@office_permission_required('S')
def order_list(request):
    sstate = request.GET.get('sstate', '')   # 콤마 구분: "101,102,..."
    spayway = request.GET.get('spayway', '')
    ssdate = request.GET.get('ssdate', '')
    sedate = request.GET.get('sedate', '')
    sword = request.GET.get('sword', '').strip()

    qs = Order.objects.filter(is_finish='T').order_by('-id')

    # 상태 필터 (ASP 그룹 코드 → Django state)
    STATE_MAP = {
        '101': [100],
        '102': [200, 301],
        '103': [302],
        '105': [402],
    }
    if sstate:
        state_vals = []
        for code in sstate.split(','):
            code = code.strip()
            if code in STATE_MAP:
                state_vals.extend(STATE_MAP[code])
        if state_vals:
            qs = qs.filter(state__in=state_vals)

    # 결제수단 필터
    if spayway == 'card':
        qs = qs.filter(payway='CARD')
    elif spayway == 'bank':
        qs = qs.filter(payway='BANK')
    elif spayway == 'zero':
        qs = qs.filter(payway='ZEROPAY')

    # 날짜 필터
    if ssdate:
        try:
            qs = qs.filter(reg_date__date__gte=datetime.strptime(ssdate, '%Y-%m-%d').date())
        except ValueError:
            pass
    if sedate:
        try:
            qs = qs.filter(reg_date__date__lte=datetime.strptime(sedate, '%Y-%m-%d').date())
        except ValueError:
            pass

    # 키워드 검색
    if sword:
        qs = qs.filter(
            Q(order_no__icontains=sword) |
            Q(user_id__icontains=sword) |
            Q(ord_name__icontains=sword) |
            Q(ord_mobile__icontains=sword) |
            Q(rcv_name__icontains=sword) |
            Q(rcv_mobile__icontains=sword) |
            Q(items__goods_title__icontains=sword)
        ).distinct()

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # 주문 상품 prefetch 및 상태명 부여
    orders = list(page_obj.object_list)
    order_ids = [o.id for o in orders]
    items_map = {}
    for item in OrderItem.objects.filter(order_id__in=order_ids).prefetch_related('options'):
        items_map.setdefault(item.order_id, []).append(item)
    delivery_map = {}
    for d in OrderDelivery.objects.filter(order_id__in=order_ids):
        delivery_map[d.order_id] = d

    total_count = paginator.count
    for i, o in enumerate(orders):
        o.state_name = STATE_NAMES.get(o.state, str(o.state))
        o.payway_name = PAYWAY_NAMES.get(o.payway, o.payway)
        o.order_items = items_map.get(o.id, [])
        o.delivery = delivery_map.get(o.id)
        o.row_num = total_count - (page_obj.number - 1) * 20 - i

    # 현재 검색 상태 체크박스용 dict
    sstate_list = sstate.split(',') if sstate else []

    return render(request, 'ba_office/lfshop/order_list.html', {
        'page_obj': page_obj,
        'orders': orders,
        'sstate': sstate,
        'sstate_list': sstate_list,
        'spayway': spayway,
        'ssdate': ssdate, 'sedate': sedate, 'sword': sword,
    })


@office_login_required
@office_permission_required('S')
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    items = OrderItem.objects.filter(order=order).prefetch_related('options')
    delivery = OrderDelivery.objects.filter(order=order).first()
    kcp = order.kcp_payments.first() if hasattr(order, 'kcp_payments') else None

    order.state_name = STATE_NAMES.get(order.state, str(order.state))
    order.payway_name = PAYWAY_NAMES.get(order.payway, order.payway)

    return render(request, 'ba_office/lfshop/order_detail.html', {
        'order': order,
        'items': items,
        'delivery': delivery,
        'kcp': kcp,
    })


@office_login_required
@office_permission_required('S')
def order_confirm(request):
    """입금확인 → 배송준비(200)"""
    if request.method == 'POST':
        pk = request.POST.get('pk', '')
        if pk:
            Order.objects.filter(pk=int(pk), is_finish='T').update(
                state=200, is_confirm='T', confirm_date=timezone.now()
            )
    return redirect(request.POST.get('next', '/ba_office/lfshop/order/list/'))


@office_login_required
@office_permission_required('S')
def order_delivery(request):
    """배송완료(302) + 배송사/송장 저장"""
    if request.method == 'POST':
        pk = request.POST.get('pk', '')
        delivery_name = request.POST.get('delivery_name', '').strip()
        delivery_no = request.POST.get('delivery_no', '').strip()
        if pk:
            Order.objects.filter(pk=int(pk)).update(
                state=302, is_delivery_finish='T'
            )
            d, _ = OrderDelivery.objects.get_or_create(order_id=int(pk))
            d.delivery_name = delivery_name
            d.delivery_no = delivery_no
            d.is_delivery = 'Y'
            d.delivery_date = timezone.now()
            d.save()
    return redirect(request.POST.get('next', '/ba_office/lfshop/order/list/'))


@office_login_required
@office_permission_required('S')
def order_cancel(request):
    """주문 취소(402)"""
    if request.method == 'POST':
        pk = request.POST.get('pk', '')
        if pk:
            Order.objects.filter(pk=int(pk)).update(
                state=402, is_cancel='T', cancel_date=timezone.now()
            )
    return redirect(request.POST.get('next', '/ba_office/lfshop/order/list/'))


@office_login_required
@office_permission_required('S')
def order_memo_save(request):
    """관리자 메모 저장"""
    if request.method == 'POST':
        pk = request.POST.get('pk', '')
        memo = request.POST.get('admin_memo', '').strip()
        if pk:
            Order.objects.filter(pk=int(pk)).update(admin_memo=memo)
    return redirect(request.POST.get('next', '/ba_office/lfshop/order/list/'))


# ============================================================
# 배송용리포트
# ============================================================

def _get_deliver_data(ssdate, sedate):
    """배송용 리포트 데이터 조회"""
    try:
        sd = datetime.strptime(ssdate, '%Y%m%d').date()
        ed = datetime.strptime(sedate, '%Y%m%d').date()
    except (ValueError, TypeError):
        return []

    # 현재 수강 구장명 dict {child_id: sta_name}
    current_ym = datetime.now().strftime('%Y%m')
    stadium_dict = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT e.child_id, s.sta_name
                FROM enrollment_enrollment e
                JOIN enrollment_enrollmentcourse ec ON ec.no_seq_id = e.id
                JOIN courses_lecture l ON l.lecture_code = ec.lecture_code
                JOIN courses_stadium s ON s.sta_code = l.sta_code
                WHERE ec.course_stats IN ('LY', 'PN', 'LP')
                AND TO_CHAR(ec.course_ym, 'YYYYMM') = %s
                AND e.cancel_code IS NULL
            """, [current_ym])
            stadium_dict = {row[0]: row[1] for row in cursor.fetchall()}
    except Exception:
        pass

    # 자녀명 dict
    child_name_dict = {}
    try:
        from apps.accounts.models import MemberChild
        with connection.cursor() as cursor:
            cursor.execute("SELECT child_id, name FROM accounts_memberchild WHERE child_id IS NOT NULL AND child_id != ''")
            child_name_dict = {row[0]: row[1] for row in cursor.fetchall()}
    except Exception:
        pass

    # 주문 + 주문상품 조회
    orders = Order.objects.filter(
        is_cancel='F', is_confirm='T',
        confirm_date__date__gte=sd,
        confirm_date__date__lte=ed,
    ).prefetch_related('items__options').order_by('-reg_date')

    rows = []
    for order in orders:
        child_name = child_name_dict.get(order.child_id, '') if order.child_id else ''
        sta_name = stadium_dict.get(order.child_id, '') if order.child_id else ''
        for item in order.items.all():
            # 옵션 텍스트 구성
            opts = item.options.order_by('sort').all()
            opt_str = ' / '.join([f"{o.title}:{o.item}" for o in opts]) if opts else item.option_txt or ''
            detail = f"{item.goods_title} / {opt_str}" if opt_str else item.goods_title
            rows.append({
                'reg_date': order.reg_date,
                'confirm_date': order.confirm_date,
                'ord_name': order.ord_name,
                'ord_mobile': order.ord_mobile,
                'child_name': child_name,
                'sta_name': sta_name,
                'goods_title': item.goods_title,
                'price': (item.price + item.option_price) * item.ea,
                'rcv_name': order.rcv_name,
                'rcv_mobile': order.rcv_mobile,
                'rcv_addr': f"{order.rcv_post} {order.rcv_addr} {order.rcv_addr_detail}".strip(),
                'detail': detail,
                'ea': item.ea,
                'order_memo': order.order_memo,
            })
    return rows


@office_login_required
@office_permission_required('S')
def deliver_list(request):
    ssdate = request.GET.get('ssdate', '')
    sedate = request.GET.get('sedate', '')
    rows = []
    searched = False
    if ssdate and sedate:
        searched = True
        rows = _get_deliver_data(ssdate, sedate)

    return render(request, 'ba_office/lfshop/deliver_list.html', {
        'rows': rows,
        'ssdate': ssdate,
        'sedate': sedate,
        'searched': searched,
    })


@office_login_required
@office_permission_required('S')
def deliver_excel(request):
    ssdate = request.GET.get('ssdate', '')
    sedate = request.GET.get('sedate', '')
    rows = _get_deliver_data(ssdate, sedate)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '배송용리포트'

    header_fill = PatternFill('solid', fgColor='D9E1F2')
    bold_font = Font(bold=True)
    headers = ['주문일자', '결제일자', '신청자', '신청자연락처', '자녀명', '구장명',
               '상품명', '가격', '수령인', '수령인연락처', '상세주소', '품목상세', '수량', '주문요청사항']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

    for row_i, r in enumerate(rows, 2):
        ws.cell(row=row_i, column=1, value=r['reg_date'].strftime('%Y-%m-%d %H:%M') if r['reg_date'] else '')
        ws.cell(row=row_i, column=2, value=r['confirm_date'].strftime('%Y-%m-%d %H:%M') if r['confirm_date'] else '')
        ws.cell(row=row_i, column=3, value=r['ord_name'])
        ws.cell(row=row_i, column=4, value=r['ord_mobile'])
        ws.cell(row=row_i, column=5, value=r['child_name'])
        ws.cell(row=row_i, column=6, value=r['sta_name'])
        ws.cell(row=row_i, column=7, value=r['goods_title'])
        ws.cell(row=row_i, column=8, value=r['price'])
        ws.cell(row=row_i, column=9, value=r['rcv_name'])
        ws.cell(row=row_i, column=10, value=r['rcv_mobile'])
        ws.cell(row=row_i, column=11, value=r['rcv_addr'])
        ws.cell(row=row_i, column=12, value=r['detail'])
        ws.cell(row=row_i, column=13, value=r['ea'])
        ws.cell(row=row_i, column=14, value=r['order_memo'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'deliver_{ssdate}_{sedate}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# ============================================================
# 통합재고관리
# ============================================================

@office_login_required
@office_permission_required('S')
def stock_list(request):
    products = (
        Product.objects.filter(del_ok='N', gd_option_kind__in=['100', '200'])
        .prefetch_related('option_stocks', 'options__items')
        .order_by('-gd_code')
    )

    # 각 상품의 옵션 재고 구성
    item_map = {}  # {ProductOptionItem.id: ProductOptionItem}
    for p in products:
        for opt in p.options.all():
            for itm in opt.items.all():
                item_map[itm.id] = itm

    for p in products:
        stocks = list(p.option_stocks.all())
        for s in stocks:
            s.item1 = item_map.get(s.opt_item_idx1)
            s.item2 = item_map.get(s.opt_item_idx2)
        p.stock_list = stocks

    return render(request, 'ba_office/lfshop/stock_list.html', {
        'products': products,
    })


@office_login_required
@office_permission_required('S')
def stock_save(request):
    if request.method == 'POST':
        for key, val in request.POST.items():
            # key 형식: stock_{product_id}_{opt_item_idx1}_{opt_item_idx2}
            if key.startswith('stock_'):
                parts = key.split('_')
                if len(parts) == 4:
                    try:
                        product_id = int(parts[1])
                        idx1 = int(parts[2])
                        idx2 = int(parts[3])
                        qty = int(val or 0)
                        ProductOptionStock.objects.filter(
                            product_id=product_id,
                            opt_item_idx1=idx1,
                            opt_item_idx2=idx2,
                        ).update(opt_stock=qty)
                    except (ValueError, TypeError):
                        pass
    return redirect('/ba_office/lfshop/stock/list/')
