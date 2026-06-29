import hashlib
import os
import re
import calendar
from datetime import datetime, date
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.conf import settings
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, Count, Max, Value, IntegerField, F
from django.db.models.functions import Coalesce
from .models import OfficeUser, OfficeLoginHistory
from .decorators import office_login_required, office_permission_required
from apps.notifications.models import OfficeNotification, Notification, SMSLog
from apps.notifications import infobank
from apps.common.models import CodeGroup, CodeValue, Setting
from apps.points.models import PointConfig, PointHistory
from apps.courses.models import (
    Coach, Stadium, StadiumCoach, Lecture, LectureSelDay,
    StadiumGoal, Promotion, PromotionMember, LectureTraining,
)
from apps.accounts.models import Member, MemberChild, OutMember
from apps.enrollment.models import (
    Enrollment, EnrollmentCourse, EnrollmentBill,
    Attendance, ChangeHistory, WaitStudent,
    EnrollmentSrc, EnrollmentBillSrc, EnrollmentCourseSrc,
)
from apps.consult.models import Consult, ConsultAnswer, ConsultFree, ConsultRegion
from apps.payments.models import PaymentToss
from apps.payments import toss


def _refund_enrollment_toss(enrollment):
    """입단 결제가 Toss면 실제 취소(환불). 반환 (성공여부, 실패메시지)."""
    log = PaymentToss.objects.filter(pay_seq=enrollment.id).order_by('-id').first()
    if not log or not log.payment_key:
        return True, ''  # Toss 결제건 아님(과거/KCP 등) → 환불 스킵
    http_code, body = toss.cancel(log.payment_key, '입단 취소')
    # 200=정상 취소, ALREADY_CANCELED=이미 취소됨(둘 다 성공으로 간주)
    if http_code == 200 or body.get('code') == 'ALREADY_CANCELED_PAYMENT':
        return True, ''
    return False, str(body.get('message', '환불 실패')).replace('"', "'")


def login_view(request):
    """관리자 로그인"""
    if request.session.get('office_user'):
        return redirect('office_main')

    error = ''
    if request.method == 'POST':
        login_id = request.POST.get('login_id', '').strip()
        login_pwd = request.POST.get('login_pwd', '').strip()

        if not login_id or not login_pwd:
            error = '아이디와 비밀번호를 입력해주세요.'
        else:
            try:
                user = OfficeUser.objects.get(office_id=login_id, del_chk='N')
                # SHA256 해시 비교
                hashed = hashlib.sha256(login_pwd.encode('utf-8')).hexdigest()
                if user.office_pwd != hashed:
                    error = '비밀번호가 일치하지 않습니다.'
                elif user.use_auth != 'W':
                    error = '허가된 사용자가 아닙니다.'
                else:
                    # 세션에 관리자 정보 저장
                    request.session['office_user'] = {
                        'office_code': user.office_code,
                        'office_id': user.office_id,
                        'office_name': user.office_name,
                        'power_level': user.power_level or '',
                        'coach_code': user.coach_code or '',
                    }
                    # 로그인 이력
                    ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
                    if not ip:
                        ip = request.META.get('REMOTE_ADDR', '')
                    OfficeLoginHistory.objects.create(
                        office_id=user.office_id,
                        action='로그인',
                        memo='로그인 하였습니다.',
                        login_ip=ip,
                    )
                    return redirect('office_main')
            except OfficeUser.DoesNotExist:
                error = '존재하지 않는 관리자 아이디입니다.'

    return render(request, 'ba_office/login.html', {'error': error})


def logout_view(request):
    """관리자 로그아웃"""
    office_user = request.session.get('office_user')
    if office_user:
        ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
        if not ip:
            ip = request.META.get('REMOTE_ADDR', '')
        OfficeLoginHistory.objects.create(
            office_id=office_user['office_id'],
            action='로그아웃',
            memo='로그아웃 하였습니다.',
            login_ip=ip,
        )
        del request.session['office_user']
    return redirect('office_login')


@office_login_required
def main_view(request):
    """관리자 메인 대시보드"""
    notices = OfficeNotification.objects.filter(del_chk='N').order_by('-no_seq')
    return render(request, 'ba_office/main.html', {
        'notices': notices,
    })


# ============================================================
# 시스템관리 > 관리자 관리
# ============================================================

@office_login_required
@office_permission_required('A')
def ofuser_list(request):
    """관리자 목록"""
    users = OfficeUser.objects.filter(del_chk='N').order_by('office_name')
    return render(request, 'ba_office/manage/ofuser_list.html', {
        'users': users,
    })


@office_login_required
@office_permission_required('A')
def ofuser_write(request):
    """관리자 등록"""
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')

    if request.method == 'POST':
        office_name = request.POST.get('office_name', '').strip()
        office_realname = request.POST.get('office_realname', '').strip()
        office_id = request.POST.get('office_id', '').strip()
        office_pwd = request.POST.get('office_pwd', '').strip()
        office_pwd_con = request.POST.get('office_pwd_con', '').strip()
        office_part = request.POST.get('office_part', '').strip()
        office_mail = request.POST.get('office_mail', '').strip()
        office_hp = request.POST.get('office_hp', '').strip()
        office_auth_list = request.POST.getlist('office_auth')
        use_auth = request.POST.get('use_auth', 'W')
        coach_code = request.POST.get('coach_code', '').strip()

        error = ''
        if not office_name:
            error = '표시명을 입력해주세요.'
        elif not office_id or len(office_id) < 6:
            error = '아이디는 6자 이상 입력해주세요.'
        elif not office_pwd or len(office_pwd) < 10:
            error = '비밀번호는 10자 이상 입력해주세요.'
        elif office_pwd != office_pwd_con:
            error = '비밀번호가 일치하지 않습니다.'
        elif office_id == office_pwd:
            error = '아이디와 비밀번호는 같을 수 없습니다.'
        elif not office_hp:
            error = '연락처를 입력해주세요.'
        elif not office_auth_list:
            error = '메뉴 권한을 1개 이상 선택해주세요.'
        elif OfficeUser.objects.filter(office_id=office_id).exists():
            error = '이미 사용 중인 아이디입니다.'
        elif coach_code and OfficeUser.objects.filter(coach_code=coach_code, del_chk='N').exists():
            error = '이미 등록된 코치코드입니다.'

        if error:
            return render(request, 'ba_office/manage/ofuser_write.html', {
                'error': error,
                'coaches': coaches,
                'form_data': request.POST,
            })

        power_level = ','.join(office_auth_list)
        hashed_pwd = hashlib.sha256(office_pwd.encode('utf-8')).hexdigest()

        OfficeUser.objects.create(
            office_name=office_name,
            office_realname=office_realname,
            office_id=office_id,
            office_pwd=hashed_pwd,
            office_part=office_part,
            office_mail=office_mail,
            office_hp=office_hp,
            power_level=power_level,
            use_auth=use_auth,
            coach_code=coach_code,
            del_chk='N',
            insert_dt=timezone.now(),
        )
        return redirect('office_ofuser_list')

    return render(request, 'ba_office/manage/ofuser_write.html', {
        'coaches': coaches,
    })


@office_login_required
@office_permission_required('A')
def ofuser_modify(request, pk):
    """관리자 수정"""
    user = get_object_or_404(OfficeUser, office_code=pk, del_chk='N')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')

    if request.method == 'POST':
        office_name = request.POST.get('office_name', '').strip()
        office_realname = request.POST.get('office_realname', '').strip()
        office_pwd = request.POST.get('office_pwd', '').strip()
        office_pwd_con = request.POST.get('office_pwd_con', '').strip()
        pwd_chk = request.POST.get('pwd_chk', '')
        office_part = request.POST.get('office_part', '').strip()
        office_mail = request.POST.get('office_mail', '').strip()
        office_hp = request.POST.get('office_hp', '').strip()
        office_auth_list = request.POST.getlist('office_auth')
        use_auth = request.POST.get('use_auth', 'W')
        coach_code = request.POST.get('coach_code', '').strip()

        error = ''
        if not office_name:
            error = '표시명을 입력해주세요.'
        elif not office_hp:
            error = '연락처를 입력해주세요.'
        elif not office_auth_list:
            error = '메뉴 권한을 1개 이상 선택해주세요.'
        elif pwd_chk == 'on':
            if not office_pwd or len(office_pwd) < 10:
                error = '비밀번호는 10자 이상 입력해주세요.'
            elif office_pwd != office_pwd_con:
                error = '비밀번호가 일치하지 않습니다.'
            elif user.office_id == office_pwd:
                error = '아이디와 비밀번호는 같을 수 없습니다.'
        if not error and coach_code:
            dup = OfficeUser.objects.filter(coach_code=coach_code, del_chk='N').exclude(office_code=pk)
            if dup.exists():
                error = '이미 등록된 코치코드입니다.'

        if error:
            return render(request, 'ba_office/manage/ofuser_modify.html', {
                'error': error,
                'user': user,
                'coaches': coaches,
                'form_data': request.POST,
            })

        user.office_name = office_name
        user.office_realname = office_realname
        user.office_part = office_part
        user.office_mail = office_mail
        user.office_hp = office_hp
        user.power_level = ','.join(office_auth_list)
        user.use_auth = use_auth
        user.coach_code = coach_code

        if pwd_chk == 'on' and office_pwd:
            user.office_pwd = hashlib.sha256(office_pwd.encode('utf-8')).hexdigest()

        user.save()
        return redirect('office_ofuser_list')

    return render(request, 'ba_office/manage/ofuser_modify.html', {
        'user': user,
        'coaches': coaches,
    })


@office_login_required
@office_permission_required('A')
def ofuser_del(request, pk):
    """관리자 삭제 (소프트 삭제)"""
    if request.method == 'POST':
        user = get_object_or_404(OfficeUser, office_code=pk)
        user.del_chk = 'Y'
        user.save()
    return redirect('office_ofuser_list')


@office_login_required
@office_permission_required('A')
def ofuser_idcheck(request):
    """관리자 아이디 중복 체크 (AJAX)"""
    office_id = request.GET.get('office_id', '').strip()
    exists = OfficeUser.objects.filter(office_id=office_id).exists()
    return JsonResponse({'exists': exists})


# ============================================================
# 시스템관리 > 코드 관리
# ============================================================

@office_login_required
@office_permission_required('A')
def code_list(request):
    """코드 그룹 목록"""
    groups = CodeGroup.objects.filter(del_chk='N').order_by('grpcode')
    selected = request.GET.get('grpcode', '')
    return render(request, 'ba_office/manage/code_list.html', {
        'groups': groups,
        'selected': selected,
    })


@office_login_required
@office_permission_required('A')
def code_sub_list(request):
    """서브 코드 목록 (iframe용)"""
    grpcode = request.GET.get('grpcode', '')
    grpcode_name = request.GET.get('grpcode_name', '')
    codes = []
    if grpcode:
        codes = CodeValue.objects.filter(group_id=grpcode, del_chk='N').order_by('code_order')
    return render(request, 'ba_office/manage/code_sub_list.html', {
        'grpcode': grpcode,
        'grpcode_name': grpcode_name,
        'codes': codes,
    })


@office_login_required
@office_permission_required('A')
def codegroup_write(request):
    """코드 그룹 등록"""
    if request.method == 'POST':
        grpcode = request.POST.get('grpcode', '').strip().upper()
        grpcode_name = request.POST.get('grpcode_name', '').strip()
        office_user = request.session.get('office_user', {})

        error = ''
        if not grpcode or len(grpcode) != 4:
            error = '그룹코드는 영문 대문자 4자리로 입력해주세요.'
        elif not grpcode.isalpha():
            error = '그룹코드는 영문 대문자만 가능합니다.'
        elif not grpcode_name:
            error = '그룹명을 입력해주세요.'
        elif CodeGroup.objects.filter(grpcode=grpcode).exists():
            error = '이미 존재하는 그룹코드입니다.'

        if error:
            return render(request, 'ba_office/manage/codegroup_write.html', {
                'error': error,
                'form_data': request.POST,
            })

        CodeGroup.objects.create(
            grpcode=grpcode,
            grpcode_name=grpcode_name,
            del_chk='N',
            insert_dt=timezone.now(),
            insert_id=office_user.get('office_id', ''),
        )
        return redirect('office_code_list')

    return render(request, 'ba_office/manage/codegroup_write.html')


@office_login_required
@office_permission_required('A')
def codegroup_modify(request, pk):
    """코드 그룹 수정"""
    group = get_object_or_404(CodeGroup, grpcode=pk, del_chk='N')

    if request.method == 'POST':
        grpcode_name = request.POST.get('grpcode_name', '').strip()

        if not grpcode_name:
            return render(request, 'ba_office/manage/codegroup_modify.html', {
                'error': '그룹명을 입력해주세요.',
                'group': group,
            })

        group.grpcode_name = grpcode_name
        group.save()
        return redirect('office_code_list')

    return render(request, 'ba_office/manage/codegroup_modify.html', {
        'group': group,
    })


@office_login_required
@office_permission_required('A')
def codegroup_del(request, pk):
    """코드 그룹 삭제 (소프트 삭제)"""
    if request.method == 'POST':
        group = get_object_or_404(CodeGroup, grpcode=pk)
        group.del_chk = 'Y'
        group.save()
        # 하위 코드도 삭제
        CodeValue.objects.filter(group_id=pk).update(del_chk='Y')
    return redirect('office_code_list')


@office_login_required
@office_permission_required('A')
def codesub_write(request):
    """서브 코드 등록"""
    grpcode = request.GET.get('grpcode', '') or request.POST.get('grpcode', '')
    grpcode_name = request.GET.get('grpcode_name', '') or request.POST.get('grpcode_name', '')

    if request.method == 'POST':
        code_name = request.POST.get('code_name', '').strip()
        code_desc = request.POST.get('code_desc', '').strip()
        code_order = request.POST.get('code_order', '0').strip()
        office_user = request.session.get('office_user', {})

        error = ''
        if not code_name:
            error = '코드명을 입력해주세요.'
        elif not code_order.isdigit():
            error = '정렬순서는 숫자만 입력해주세요.'

        if error:
            return render(request, 'ba_office/manage/codesub_write.html', {
                'error': error,
                'grpcode': grpcode,
                'grpcode_name': grpcode_name,
                'form_data': request.POST,
            })

        # 다음 subcode 계산
        max_sub = CodeValue.objects.filter(group_id=grpcode).order_by('-subcode').first()
        next_subcode = (max_sub.subcode + 1) if max_sub else 1

        CodeValue.objects.create(
            subcode=next_subcode,
            group_id=grpcode,
            code_name=code_name,
            code_desc=code_desc,
            code_order=int(code_order),
            del_chk='N',
            insert_dt=timezone.now(),
            insert_id=office_user.get('office_id', ''),
        )
        return redirect(f'/ba_office/manage/code/sub/?grpcode={grpcode}&grpcode_name={grpcode_name}')

    return render(request, 'ba_office/manage/codesub_write.html', {
        'grpcode': grpcode,
        'grpcode_name': grpcode_name,
    })


@office_login_required
@office_permission_required('A')
def codesub_modify(request, grpcode, subcode):
    """서브 코드 수정"""
    code = get_object_or_404(CodeValue, group_id=grpcode, subcode=subcode, del_chk='N')
    grpcode_name = request.GET.get('grpcode_name', '') or request.POST.get('grpcode_name', '')
    if not grpcode_name:
        grpcode_name = code.group.grpcode_name if code.group else ''

    if request.method == 'POST':
        code_name = request.POST.get('code_name', '').strip()
        code_desc = request.POST.get('code_desc', '').strip()
        code_order = request.POST.get('code_order', '0').strip()

        error = ''
        if not code_name:
            error = '코드명을 입력해주세요.'
        elif not code_order.isdigit():
            error = '정렬순서는 숫자만 입력해주세요.'

        if error:
            return render(request, 'ba_office/manage/codesub_modify.html', {
                'error': error,
                'code': code,
                'grpcode': grpcode,
                'grpcode_name': grpcode_name,
            })

        code.code_name = code_name
        code.code_desc = code_desc
        code.code_order = int(code_order)
        code.save()
        return redirect(f'/ba_office/manage/code/sub/?grpcode={grpcode}&grpcode_name={grpcode_name}')

    return render(request, 'ba_office/manage/codesub_modify.html', {
        'code': code,
        'grpcode': grpcode,
        'grpcode_name': grpcode_name,
    })


@office_login_required
@office_permission_required('A')
def codesub_del(request):
    """서브 코드 삭제 (소프트 삭제)"""
    if request.method == 'POST':
        grpcode = request.POST.get('grpcode', '')
        subcode = request.POST.get('subcode', '')
        grpcode_name = request.POST.get('grpcode_name', '')
        if grpcode and subcode:
            CodeValue.objects.filter(group_id=grpcode, subcode=int(subcode)).update(del_chk='Y')
        return redirect(f'/ba_office/manage/code/sub/?grpcode={grpcode}&grpcode_name={grpcode_name}')
    return redirect('office_code_list')


# ============================================================
# 시스템관리 > 포인트 설정
# ============================================================

@office_login_required
@office_permission_required('A')
def point_setup(request):
    """포인트 설정"""
    if request.method == 'POST':
        configs = PointConfig.objects.all().order_by('point_seq')
        for config in configs:
            use_yn = request.POST.get(f'use_yn_{config.point_seq}', config.use_yn)
            save_point = request.POST.get(f'save_point_{config.point_seq}', '')
            limit_point = request.POST.get(f'limit_point_{config.point_seq}', '')

            config.use_yn = use_yn
            if save_point != '':
                config.save_point = int(save_point)
            if limit_point != '':
                config.limit_point = int(limit_point)
            config.save()

        return redirect('office_point_setup')

    configs = PointConfig.objects.all().order_by('point_seq')
    return render(request, 'ba_office/manage/point_setup.html', {
        'configs': configs,
    })


# ============================================================
# 시스템관리 > 관리자 알림
# ============================================================

@office_login_required
@office_permission_required('A')
def office_alim_list(request):
    """관리자 알림 목록"""
    notices = OfficeNotification.objects.filter(del_chk='N').order_by('-no_seq')
    return render(request, 'ba_office/manage/office_alim_list.html', {
        'notices': notices,
    })


@office_login_required
@office_permission_required('A')
def office_alim_write(request):
    """관리자 알림 등록"""
    if request.method == 'POST':
        atitle = request.POST.get('atitle', '').strip()
        acontent = request.POST.get('acontent', '').strip()
        office_user = request.session.get('office_user', {})

        error = ''
        if not atitle:
            error = '제목을 입력해주세요.'
        elif not acontent:
            error = '내용을 입력해주세요.'

        if error:
            return render(request, 'ba_office/manage/office_alim_write.html', {
                'error': error,
                'form_data': request.POST,
            })

        # 다음 no_seq
        max_seq = OfficeNotification.objects.order_by('-no_seq').first()
        next_seq = (max_seq.no_seq + 1) if max_seq else 1

        OfficeNotification.objects.create(
            no_seq=next_seq,
            atitle=atitle,
            acontent=acontent,
            del_chk='N',
            reg_dt=timezone.now(),
            reg_id=office_user.get('office_id', ''),
        )
        return redirect('office_alim_list')

    return render(request, 'ba_office/manage/office_alim_write.html')


@office_login_required
@office_permission_required('A')
def office_alim_modify(request, pk):
    """관리자 알림 수정"""
    notice = get_object_or_404(OfficeNotification, no_seq=pk, del_chk='N')

    if request.method == 'POST':
        atitle = request.POST.get('atitle', '').strip()
        acontent = request.POST.get('acontent', '').strip()

        error = ''
        if not atitle:
            error = '제목을 입력해주세요.'
        elif not acontent:
            error = '내용을 입력해주세요.'

        if error:
            return render(request, 'ba_office/manage/office_alim_modify.html', {
                'error': error,
                'notice': notice,
            })

        notice.atitle = atitle
        notice.acontent = acontent
        notice.save()
        return redirect('office_alim_list')

    return render(request, 'ba_office/manage/office_alim_modify.html', {
        'notice': notice,
    })


@office_login_required
@office_permission_required('A')
def office_alim_del(request, pk):
    """관리자 알림 삭제 (소프트 삭제)"""
    if request.method == 'POST':
        notice = get_object_or_404(OfficeNotification, no_seq=pk)
        notice.del_chk = 'Y'
        notice.save()
    return redirect('office_alim_list')


# ============================================================
# 회원관리 > 회원정보
# ============================================================

@office_login_required
@office_permission_required('M')
def member_list(request):
    """회원 목록"""
    sch_type = request.GET.get('sch_type', 'name')
    sch_keyword = request.GET.get('sch_keyword', '').strip()
    sch_status = request.GET.get('sch_member_status', '')
    page = request.GET.get('page', '1')

    # 검색 조건이 하나라도 있을 때만 조회
    is_searched = bool(sch_keyword or sch_status)

    members = None
    total_count = 0

    if is_searched:
        qs = Member.objects.filter(is_superuser=False)
        if sch_keyword:
            if sch_type == 'phone':
                qs = qs.filter(phone__icontains=sch_keyword)
            else:
                qs = qs.filter(name__icontains=sch_keyword)
        if sch_status:
            qs = qs.filter(status=sch_status)

        # 포인트, 자녀수 어노테이션
        qs = qs.annotate(
            add_point=Coalesce(
                Sum('point_histories__app_point', filter=Q(point_histories__app_gbn='S')),
                Value(0), output_field=IntegerField()
            ),
            min_point=Coalesce(
                Sum('point_histories__app_point', filter=Q(point_histories__app_gbn='U')),
                Value(0), output_field=IntegerField()
            ),
            child_count=Count('children', distinct=True),
        ).order_by('-insert_dt', '-id')

        paginator = Paginator(qs, 20)
        members = paginator.get_page(page)
        total_count = paginator.count

    return render(request, 'ba_office/lfmember/member_list.html', {
        'members': members,
        'is_searched': is_searched,
        'sch_type': sch_type,
        'sch_keyword': sch_keyword,
        'sch_member_status': sch_status,
        'total_count': total_count,
    })


@office_login_required
@office_permission_required('M')
def member_list_excel(request):
    """회원 목록 엑셀 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sch_name = request.GET.get('sch_member_name', '').strip()
    sch_id = request.GET.get('sch_member_id', '').strip()
    sch_status = request.GET.get('sch_member_status', '')

    qs = Member.objects.filter(is_superuser=False)
    if sch_name:
        qs = qs.filter(name__icontains=sch_name)
    if sch_id:
        qs = qs.filter(username__icontains=sch_id)
    if sch_status:
        qs = qs.filter(status=sch_status)

    qs = qs.annotate(
        add_point=Coalesce(
            Sum('point_histories__app_point', filter=Q(point_histories__app_gbn='S')),
            Value(0), output_field=IntegerField()
        ),
        min_point=Coalesce(
            Sum('point_histories__app_point', filter=Q(point_histories__app_gbn='U')),
            Value(0), output_field=IntegerField()
        ),
        child_count=Count('children'),
    ).order_by('-insert_dt', '-id')[:10000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '회원목록'

    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['순번', '이름', '아이디', 'Point', '연락처', '등록자녀', '메일수신', 'SMS수신', '가입일', '상태']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, m in enumerate(qs, 1):
        point = m.add_point - m.min_point
        status_map = {'N': '정상', 'S': '중지', 'O': '탈퇴'}
        row = [i, m.name, m.username, point, m.phone, m.child_count,
               m.mail_consent, m.sms_consent,
               m.insert_dt.strftime('%Y-%m-%d') if m.insert_dt else '',
               status_map.get(m.status, m.status)]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=member_list.xlsx'
    wb.save(response)
    return response


@office_login_required
@office_permission_required('M')
def member_write(request):
    """회원 등록"""
    if request.method == 'POST':
        member_name = request.POST.get('member_name', '').strip()
        member_id = request.POST.get('member_id', '').strip()
        member_pwd = request.POST.get('member_pwd', '').strip()
        member_pwd_con = request.POST.get('member_pwd_con', '').strip()
        mhtel1 = request.POST.get('mhtel1', '').strip()
        mhtel2 = request.POST.get('mhtel2', '').strip()
        mhtel3 = request.POST.get('mhtel3', '').strip()
        mtel1 = request.POST.get('mtel1', '').strip()
        mtel2 = request.POST.get('mtel2', '').strip()
        mtel3 = request.POST.get('mtel3', '').strip()
        member_mail = request.POST.get('member_mail', '').strip()
        zipcode = request.POST.get('zipcode', '').strip()
        address1 = request.POST.get('address1', '').strip()
        address2 = request.POST.get('address2', '').strip()
        smsyn = request.POST.get('smsyn', 'Y')
        mailyn = request.POST.get('mailyn', 'Y')
        member_status = request.POST.get('member_status', 'N')

        error = ''
        if not member_name:
            error = '이름을 입력해주세요.'
        elif not member_id or len(member_id) < 6:
            error = '아이디는 6자 이상 입력해주세요.'
        elif not member_pwd or len(member_pwd) < 8:
            error = '비밀번호는 8자 이상 입력해주세요.'
        elif member_pwd != member_pwd_con:
            error = '비밀번호가 일치하지 않습니다.'
        elif member_id == member_pwd:
            error = '아이디와 비밀번호는 같을 수 없습니다.'
        elif not mhtel1 or not mhtel2 or not mhtel3:
            error = '휴대전화를 입력해주세요.'
        elif Member.objects.filter(username=member_id).exists():
            error = '이미 사용 중인 아이디입니다.'

        if error:
            return render(request, 'ba_office/lfmember/member_write.html', {
                'error': error, 'form_data': request.POST,
            })

        phone = f'{mhtel1}-{mhtel2}-{mhtel3}' if mhtel1 else ''
        tel = f'{mtel1}-{mtel2}-{mtel3}' if mtel1 and mtel2 else ''

        member = Member.objects.create_user(
            username=member_id,
            password=member_pwd,
            name=member_name,
            phone=phone,
            tel=tel,
            email=member_mail,
            zipcode=zipcode,
            address1=address1,
            address2=address2,
            sms_consent=smsyn,
            mail_consent=mailyn,
            status=member_status,
            insert_dt=timezone.now(),
        )

        # 자녀 등록
        child_cnt = int(request.POST.get('child_cnt', '1'))
        for i in range(1, child_cnt + 1):
            c_name = request.POST.get(f'child_name_{i}', '').strip()
            c_id = request.POST.get(f'child_id_{i}', '').strip()
            c_pwd = request.POST.get(f'child_pwd_{i}', '').strip()
            c_birth = request.POST.get(f'child_birth_{i}', '').strip()
            c_school = request.POST.get(f'sch_name_{i}', '').strip()
            c_grade = request.POST.get(f'sch_grade_{i}', '').strip()
            c_height = request.POST.get(f'child_height_{i}', '').strip()
            c_weight = request.POST.get(f'child_weight_{i}', '').strip()
            c_size = request.POST.get(f'child_size_{i}', '').strip()
            c_tel1 = request.POST.get(f'chtel1_{i}', '').strip()
            c_tel2 = request.POST.get(f'chtel2_{i}', '').strip()
            c_tel3 = request.POST.get(f'chtel3_{i}', '').strip()
            c_sex = request.POST.get(f'child_sexgbn_{i}', '').strip()

            if c_name and c_id:
                c_phone = f'{c_tel1}-{c_tel2}-{c_tel3}' if c_tel1 and c_tel2 else ''
                MemberChild.objects.create(
                    parent=member,
                    name=c_name,
                    child_id=c_id,
                    child_pwd=hashlib.sha256(c_pwd.encode('utf-8')).hexdigest() if c_pwd else '',
                    birth=c_birth,
                    school=c_school,
                    grade=c_grade,
                    height=c_height,
                    weight=c_weight,
                    size=c_size,
                    phone=c_phone,
                    gender=c_sex,
                    status='N',
                    course_state='CAN',
                    insert_dt=timezone.now(),
                )

        return redirect('office_member_list')

    return render(request, 'ba_office/lfmember/member_write.html')


@office_login_required
@office_permission_required('M')
def member_modify(request, member_id):
    """회원 수정"""
    member = get_object_or_404(Member, username=member_id)

    if request.method == 'POST':
        member_pwd = request.POST.get('member_pwd', '').strip()
        member_pwd_con = request.POST.get('member_pwd_con', '').strip()
        pwd_chk = request.POST.get('pwd_chk', '')
        mhtel1 = request.POST.get('mhtel1', '').strip()
        mhtel2 = request.POST.get('mhtel2', '').strip()
        mhtel3 = request.POST.get('mhtel3', '').strip()
        mtel1 = request.POST.get('mtel1', '').strip()
        mtel2 = request.POST.get('mtel2', '').strip()
        mtel3 = request.POST.get('mtel3', '').strip()
        member_mail = request.POST.get('member_mail', '').strip()
        zipcode = request.POST.get('zipcode', '').strip()
        address1 = request.POST.get('address1', '').strip()
        address2 = request.POST.get('address2', '').strip()
        smsyn = request.POST.get('smsyn', 'Y')
        mailyn = request.POST.get('mailyn', 'Y')
        member_status = request.POST.get('member_status', 'N')

        error = ''
        if not mhtel1 or not mhtel2 or not mhtel3:
            error = '휴대전화를 입력해주세요.'
        elif pwd_chk == 'on':
            if not member_pwd or len(member_pwd) < 8:
                error = '비밀번호는 8자 이상 입력해주세요.'
            elif member_pwd != member_pwd_con:
                error = '비밀번호가 일치하지 않습니다.'

        if error:
            return render(request, 'ba_office/lfmember/member_modify.html', {
                'error': error, 'member': member, 'form_data': request.POST,
            })

        member.phone = f'{mhtel1}-{mhtel2}-{mhtel3}' if mhtel1 else ''
        member.tel = f'{mtel1}-{mtel2}-{mtel3}' if mtel1 and mtel2 else ''
        member.email = member_mail
        member.zipcode = zipcode
        member.address1 = address1
        member.address2 = address2
        member.sms_consent = smsyn
        member.mail_consent = mailyn
        member.status = member_status

        if pwd_chk == 'on' and member_pwd:
            member.set_password(member_pwd)

        member.save()
        return redirect('office_member_list')

    # 전화번호 분리
    phone_parts = member.phone.split('-') if member.phone else ['', '', '']
    tel_parts = member.tel.split('-') if member.tel else ['', '', '']
    while len(phone_parts) < 3:
        phone_parts.append('')
    while len(tel_parts) < 3:
        tel_parts.append('')

    return render(request, 'ba_office/lfmember/member_modify.html', {
        'member': member,
        'mhtel1': phone_parts[0], 'mhtel2': phone_parts[1], 'mhtel3': phone_parts[2],
        'mtel1': tel_parts[0], 'mtel2': tel_parts[1], 'mtel3': tel_parts[2],
    })


@office_login_required
@office_permission_required('M')
def member_idcheck(request):
    """회원 아이디 중복 체크 (AJAX)"""
    member_id = request.GET.get('member_id', '').strip()
    exists = Member.objects.filter(username=member_id).exists()
    return JsonResponse({'exists': exists})


@office_login_required
@office_permission_required('M')
def member_childadd(request, member_id):
    """자녀 추가"""
    member = get_object_or_404(Member, username=member_id)

    if request.method == 'POST':
        c_name = request.POST.get('child_name', '').strip()
        c_id = request.POST.get('child_id', '').strip()
        c_pwd = request.POST.get('child_pwd', '').strip()
        c_pwd_con = request.POST.get('child_pwd_con', '').strip()
        c_birth = request.POST.get('child_birth', '').strip()
        c_school = request.POST.get('sch_name', '').strip()
        c_grade = request.POST.get('sch_grade', '').strip()
        c_height = request.POST.get('child_height', '').strip()
        c_weight = request.POST.get('child_weight', '').strip()
        c_size = request.POST.get('child_size', '').strip()
        c_tel1 = request.POST.get('chtel1', '').strip()
        c_tel2 = request.POST.get('chtel2', '').strip()
        c_tel3 = request.POST.get('chtel3', '').strip()
        c_sex = request.POST.get('child_sexgbn', '').strip()

        error = ''
        if not c_name:
            error = '자녀 이름을 입력해주세요.'
        elif not c_id or len(c_id) < 6:
            error = '자녀 아이디는 6자 이상 입력해주세요.'
        elif not c_pwd or len(c_pwd) < 8:
            error = '비밀번호는 8자 이상 입력해주세요.'
        elif c_pwd != c_pwd_con:
            error = '비밀번호가 일치하지 않습니다.'
        elif MemberChild.objects.filter(child_id=c_id).exists():
            error = '이미 사용 중인 자녀 아이디입니다.'

        if error:
            return render(request, 'ba_office/lfmember/member_childadd.html', {
                'error': error, 'member': member, 'form_data': request.POST,
            })

        c_phone = f'{c_tel1}-{c_tel2}-{c_tel3}' if c_tel1 and c_tel2 else ''
        MemberChild.objects.create(
            parent=member,
            name=c_name,
            child_id=c_id,
            child_pwd=hashlib.sha256(c_pwd.encode('utf-8')).hexdigest(),
            birth=c_birth,
            school=c_school,
            grade=c_grade,
            height=c_height,
            weight=c_weight,
            size=c_size,
            phone=c_phone,
            gender=c_sex,
            status='N',
            course_state='CAN',
            insert_dt=timezone.now(),
        )
        return redirect('office_member_list')

    return render(request, 'ba_office/lfmember/member_childadd.html', {
        'member': member,
    })


# ============================================================
# 회원관리 > 자녀정보
# ============================================================

@office_login_required
@office_permission_required('M')
def child_list(request):
    """자녀 목록"""
    sch_type = request.GET.get('sch_type', 'child_name')
    sch_keyword = request.GET.get('sch_keyword', '').strip()
    page = request.GET.get('page', '1')

    # 검색 조건이 있을 때만 조회
    is_searched = bool(sch_keyword)

    children = None
    total_count = 0

    if is_searched:
        qs = MemberChild.objects.select_related('parent').all()
        if sch_type == 'parent_name':
            qs = qs.filter(parent__name__icontains=sch_keyword)
        else:
            qs = qs.filter(name__icontains=sch_keyword)

        qs = qs.order_by('-insert_dt', '-id')

        paginator = Paginator(qs, 15)
        children = paginator.get_page(page)
        total_count = paginator.count

    return render(request, 'ba_office/lfmember/child_list.html', {
        'children': children,
        'is_searched': is_searched,
        'sch_type': sch_type,
        'sch_keyword': sch_keyword,
        'total_count': total_count,
    })


@office_login_required
@office_permission_required('M')
def child_modify(request, child_id):
    """자녀 수정"""
    child = get_object_or_404(MemberChild, child_id=child_id)

    if request.method == 'POST':
        child_name = request.POST.get('child_name', '').strip()
        child_birth = request.POST.get('child_birth', '').strip()
        sch_name = request.POST.get('sch_name', '').strip()
        sch_grade = request.POST.get('sch_grade', '').strip()
        child_pwd = request.POST.get('child_pwd', '').strip()
        child_pwd_con = request.POST.get('child_pwd_con', '').strip()
        pwd_chk = request.POST.get('pwd_chk', '')
        child_height = request.POST.get('child_height', '').strip()
        child_weight = request.POST.get('child_weight', '').strip()
        child_size = request.POST.get('child_size', '').strip()
        chtel1 = request.POST.get('chtel1', '').strip()
        chtel2 = request.POST.get('chtel2', '').strip()
        chtel3 = request.POST.get('chtel3', '').strip()
        child_sexgbn = request.POST.get('child_sexgbn', '').strip()
        child_status = request.POST.get('child_status', 'N')
        course_state = request.POST.get('course_state', 'CAN')
        card_num = request.POST.get('card_num', '').strip()

        error = ''
        if not child_name:
            error = '이름을 입력해주세요.'
        elif pwd_chk == 'on':
            if not child_pwd or len(child_pwd) < 8:
                error = '비밀번호는 8자 이상 입력해주세요.'
            elif child_pwd != child_pwd_con:
                error = '비밀번호가 일치하지 않습니다.'

        if error:
            return render(request, 'ba_office/lfmember/child_modify.html', {
                'error': error, 'child': child, 'form_data': request.POST,
            })

        child.name = child_name
        child.birth = child_birth
        child.school = sch_name
        child.grade = sch_grade
        child.height = child_height
        child.weight = child_weight
        child.size = child_size
        child.phone = f'{chtel1}-{chtel2}-{chtel3}' if chtel1 and chtel2 else ''
        child.gender = child_sexgbn
        child.status = child_status
        child.course_state = course_state
        child.card_num = card_num

        if pwd_chk == 'on' and child_pwd:
            child.child_pwd = hashlib.sha256(child_pwd.encode('utf-8')).hexdigest()

        child.save()
        return redirect('office_child_list')

    phone_parts = child.phone.split('-') if child.phone else ['', '', '']
    while len(phone_parts) < 3:
        phone_parts.append('')

    return render(request, 'ba_office/lfmember/child_modify.html', {
        'child': child,
        'chtel1': phone_parts[0], 'chtel2': phone_parts[1], 'chtel3': phone_parts[2],
    })


@office_login_required
@office_permission_required('M')
def child_detail(request, child_id):
    """자녀 상세 - 수강 이력"""
    child = get_object_or_404(MemberChild.objects.select_related('parent'), child_id=child_id)

    from apps.courses.models import Lecture, Stadium
    enrollments = Enrollment.objects.filter(child_id=child_id, del_chk='N').order_by('-id')

    # 수강과정 정보 매핑
    enrollment_data = []
    for enroll in enrollments:
        courses = EnrollmentCourse.objects.filter(enrollment=enroll).order_by('course_ym')
        for course in courses:
            lecture = Lecture.objects.filter(lecture_code=course.lecture_code).first()
            sta = Stadium.objects.filter(sta_code=lecture.stadium_id).first() if lecture else None
            enrollment_data.append({
                'enroll': enroll,
                'course': course,
                'lecture': lecture,
                'stadium': sta,
            })

    return render(request, 'ba_office/lfmember/child_detail.html', {
        'child': child,
        'enrollment_data': enrollment_data,
    })


# ============================================================
# 회원관리 > 회원통계
# ============================================================

@office_login_required
@office_permission_required('M')
def member_stat(request):
    """회원 가입 통계 (일별)"""
    import datetime
    syear = request.GET.get('syear', str(timezone.now().year))
    try:
        year = int(syear)
    except ValueError:
        year = timezone.now().year

    # 월/일별 가입수 집계
    from django.db.models.functions import ExtractMonth, ExtractDay
    parent_stats = (
        Member.objects.filter(
            is_superuser=False, status='N',
            insert_dt__year=year
        )
        .annotate(month=ExtractMonth('insert_dt'), day=ExtractDay('insert_dt'))
        .values('month', 'day')
        .annotate(cnt=Count('id'))
    )

    child_stats = (
        MemberChild.objects.filter(
            status='N',
            insert_dt__year=year
        )
        .annotate(month=ExtractMonth('insert_dt'), day=ExtractDay('insert_dt'))
        .values('month', 'day')
        .annotate(cnt=Count('id'))
    )

    # 데이터 구조화: stat_data[day][month] = {'p': 0, 'c': 0}
    stat_data = {}
    for d in range(1, 32):
        stat_data[d] = {}
        for m in range(1, 13):
            stat_data[d][m] = {'p': 0, 'c': 0}

    for s in parent_stats:
        stat_data[s['day']][s['month']]['p'] = s['cnt']
    for s in child_stats:
        stat_data[s['day']][s['month']]['c'] = s['cnt']

    # 월별 합계
    month_totals = []
    for m in range(1, 13):
        p_total = sum(stat_data[d][m]['p'] for d in range(1, 32))
        c_total = sum(stat_data[d][m]['c'] for d in range(1, 32))
        month_totals.append({'p': p_total, 'c': c_total})

    # 행 데이터를 리스트로 변환 (템플릿에서 쉽게 접근)
    rows = []
    for d in range(1, 32):
        row = {'day': d, 'months': []}
        day_total_p = 0
        day_total_c = 0
        for m in range(1, 13):
            row['months'].append(stat_data[d][m])
            day_total_p += stat_data[d][m]['p']
            day_total_c += stat_data[d][m]['c']
        row['total'] = {'p': day_total_p, 'c': day_total_c}
        rows.append(row)

    # 연도 목록 (2019 ~ 현재)
    current_year = timezone.now().year
    years = list(range(2019, current_year + 1))

    # 전체 합계
    grand_total_p = sum(mt['p'] for mt in month_totals)
    grand_total_c = sum(mt['c'] for mt in month_totals)

    return render(request, 'ba_office/lfmember/member_stat.html', {
        'rows': rows,
        'month_totals': month_totals,
        'grand_total': {'p': grand_total_p, 'c': grand_total_c},
        'syear': year,
        'years': years,
    })


# ============================================================
# 회원관리 > SMS/LMS
# ============================================================

_PHONE_RE = re.compile(r'^0\d{1,2}-?\d{3,4}-?\d{4}$')


def _normalize_phone(raw):
    """숫자만 추출 + 앞 0 유실(엑셀 숫자셀) 보정. 유효시 digits, 아니면 None."""
    d = re.sub(r'\D', '', str(raw if raw is not None else ''))
    if not d:
        return None
    if not d.startswith('0'):
        d = '0' + d
    return d if 10 <= len(d) <= 11 else None


def _parse_excel_phones(f):
    """업로드 엑셀(.xls/.xlsx)에서 전화번호 추출. 시트 'sms'·'번호' 컬럼 우선, 없으면 전체 스캔."""
    name = (f.name or '').lower()
    rows = []
    if name.endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=f.read())
        names = wb.sheet_names()
        sh = wb.sheet_by_name('sms') if 'sms' in names else wb.sheet_by_index(0)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb['sms'] if 'sms' in wb.sheetnames else wb.worksheets[0]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]

    if not rows:
        return []

    # 헤더에서 '번호'(전화/휴대/연락처) 컬럼 찾기
    header = [str(c or '').strip() for c in rows[0]]
    col = next((i for i, h in enumerate(header)
                if any(k in h for k in ('번호', '전화', '휴대', '연락처', 'phone', 'mobile', 'hp'))), None)

    phones, seen = [], set()
    if col is not None:
        for row in rows[1:]:
            p = _normalize_phone(row[col]) if col < len(row) else None
            if p and p not in seen:
                seen.add(p); phones.append(p)
    else:
        # 헤더 매칭 실패 → 전 셀에서 전화패턴(대시 포함)만 추출
        for row in rows:
            for cell in row:
                s = str(cell if cell is not None else '').strip()
                if _PHONE_RE.match(s):
                    p = re.sub(r'\D', '', s)
                    if p not in seen:
                        seen.add(p); phones.append(p)
    return phones


@office_login_required
@office_permission_required('M')
def sms_excel_upload(request):
    """SMS 엑셀 일괄발송 — 업로드 엑셀에서 전화번호 추출(AJAX). JSON 반환."""
    f = request.FILES.get('excel_file')
    if request.method != 'POST' or not f:
        return JsonResponse({'ok': False, 'error': '파일이 없습니다.'})
    if not (f.name or '').lower().endswith(('.xls', '.xlsx')):
        return JsonResponse({'ok': False, 'error': '엑셀(.xls/.xlsx) 파일만 업로드 가능합니다.'})
    try:
        phones = _parse_excel_phones(f)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'엑셀을 읽지 못했습니다: {e}'})
    if not phones:
        return JsonResponse({'ok': False, 'error': "전화번호를 찾지 못했습니다. (시트 'sms'의 '번호' 컬럼 확인)"})
    return JsonResponse({'ok': True, 'count': len(phones), 'numbers': phones})


@office_login_required
@office_permission_required('M')
def sms_send(request):
    """SMS/LMS 발송 폼 + 발송 처리.

    수신대상: individual(직접입력, 콤마/줄바꿈) / excel(엑셀 일괄) / coach(코치 다중선택).
    메시지 90byte 초과 시 자동 LMS. 인포뱅크 자격증명 미설정 시 테스트모드(실발송 X, 이력만).
    """
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')
    result = None

    if request.method == 'POST':
        callback = request.POST.get('callback', '').strip()
        message = request.POST.get('smsMsg', '').strip()
        rcv_type = request.POST.get('rcv_type', 'individual')

        # 수신번호 목록 구성
        if rcv_type == 'coach':
            codes = set(request.POST.getlist('coach_code'))
            recipients = [c.phone for c in coaches if str(c.coach_code) in codes and c.phone]
        else:
            # individual=rcv_num(직접입력) / excel=excel_numbers(업로드 파싱결과)
            raw = request.POST.get('excel_numbers' if rcv_type == 'excel' else 'rcv_num', '')
            recipients = [n.strip() for n in re.split(r'[,\n\r]+', raw) if n.strip()]

        errors = []
        if not callback:
            errors.append('발신번호를 선택하세요.')
        if not message:
            errors.append('메시지를 입력하세요.')
        if not recipients:
            errors.append('수신번호가 없습니다.')

        if errors:
            result = {'error': ' '.join(errors)}
        else:
            broadcast = len(recipients) > 1
            sent = failed = 0
            for to in recipients:
                res = infobank.send_and_log(to, message, callback, broadcast=broadcast)
                if res.get('ok'):
                    sent += 1
                else:
                    failed += 1
            result = {
                'sent': sent, 'failed': failed, 'total': len(recipients),
                'test': not infobank.is_configured(),
                'service': 'LMS' if infobank.sms_byte_len(message) > 90 else 'SMS',
            }

    return render(request, 'ba_office/lfmember/sms_send.html', {
        'coaches': coaches,
        'configured': infobank.is_configured(),
        'result': result,
    })


# ============================================================
# 회원관리 > SMS/LMS 내역
# ============================================================

@office_login_required
@office_permission_required('M')
def sms_send_list(request):
    """SMS 발송 내역"""
    page = request.GET.get('page', '1')
    sch_sdate = request.GET.get('sch_sdate', '')
    sch_edate = request.GET.get('sch_edate', '')

    qs = SMSLog.objects.all().order_by('-date_client_req')
    if sch_sdate:
        qs = qs.filter(date_client_req__date__gte=sch_sdate)
    if sch_edate:
        qs = qs.filter(date_client_req__date__lte=sch_edate)

    paginator = Paginator(qs, 40)
    logs = paginator.get_page(page)

    return render(request, 'ba_office/lfmember/sms_send_list.html', {
        'logs': logs,
        'sch_sdate': sch_sdate,
        'sch_edate': sch_edate,
        'total_count': paginator.count,
    })


# ============================================================
# 회원관리 > 회원포인트내역
# ============================================================

@office_login_required
@office_permission_required('M')
def memberpoint_list(request):
    """회원 포인트 내역"""
    page = request.GET.get('page', '1')
    sch_startdt = request.GET.get('sch_startdt', '')
    sch_enddt = request.GET.get('sch_enddt', '')
    sch_app_gbn = request.GET.get('sch_app_gbn', '')
    sch_val = request.GET.get('sch_val', '').strip()
    # 검색 필드 선택자(ASP memberpoint_list.asp): member_id(아이디) | member_name(이름).
    # 회원목록 '포인트' 링크는 member_id 로, 직접검색 폼은 회원명(기본)으로 진입.
    sch_txt = request.GET.get('sch_txt', '')
    sch_desc_gbn = request.GET.get('sch_desc_gbn', '')

    # 검색 조건이 하나라도 있을 때만 조회
    is_searched = bool(sch_startdt or sch_enddt or sch_app_gbn or sch_val or sch_desc_gbn)

    histories = None
    total_count = 0

    if is_searched:
        qs = PointHistory.objects.all().order_by('-insert_dt', '-id')
        if sch_startdt:
            qs = qs.filter(point_dt__gte=sch_startdt)
        if sch_enddt:
            qs = qs.filter(point_dt__lte=sch_enddt)
        if sch_app_gbn:
            qs = qs.filter(app_gbn=sch_app_gbn)
        if sch_val:
            # member_id 는 FK(to_field=username)라 icontains 불가 → 정확 일치
            # (회원목록 링크가 정확한 로그인ID를 넘기므로 정밀하게 매칭). 그 외엔 회원명 부분일치.
            if sch_txt == 'member_id':
                qs = qs.filter(member_id=sch_val)
            else:
                qs = qs.filter(member_name__icontains=sch_val)
        if sch_desc_gbn:
            qs = qs.filter(point_desc__icontains=sch_desc_gbn)

        paginator = Paginator(qs, 20)
        histories = paginator.get_page(page)
        total_count = paginator.count

    return render(request, 'ba_office/lfmember/memberpoint_list.html', {
        'histories': histories,
        'is_searched': is_searched,
        'sch_startdt': sch_startdt,
        'sch_enddt': sch_enddt,
        'sch_app_gbn': sch_app_gbn,
        'sch_val': sch_val,
        'sch_txt': sch_txt,
        'sch_desc_gbn': sch_desc_gbn,
        'total_count': total_count,
    })


@office_login_required
@office_permission_required('M')
def memberpoint_write(request):
    """포인트 수동 등록"""
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '').strip()
        app_gbn = request.POST.get('app_gbn', 'S')
        app_point = request.POST.get('app_point', '0').strip()
        point_desc = request.POST.get('point_desc', '').strip()
        office_user = request.session.get('office_user', {})

        error = ''
        if not member_id:
            error = '회원 아이디를 입력해주세요.'
        elif not Member.objects.filter(username=member_id).exists():
            error = '존재하지 않는 회원 아이디입니다.'
        elif not app_point or not app_point.isdigit() or int(app_point) <= 0:
            error = '포인트를 올바르게 입력해주세요.'
        elif not point_desc:
            error = '내용을 입력해주세요.'

        if error:
            return render(request, 'ba_office/lfmember/memberpoint_write.html', {
                'error': error, 'form_data': request.POST,
            })

        member = Member.objects.get(username=member_id)
        now = timezone.now()
        PointHistory.objects.create(
            point_dt=now.strftime('%Y%m%d'),
            member=member,
            member_name=member.name,
            app_gbn=app_gbn,
            app_point=int(app_point),
            point_desc=point_desc,
            desc_detail='관리자 직권',
            insert_dt=now,
            insert_id=office_user.get('office_id', ''),
        )
        return redirect('office_memberpoint_list')

    return render(request, 'ba_office/lfmember/memberpoint_write.html')


@office_login_required
@office_permission_required('M')
def memberpoint_del(request, pk):
    """포인트 삭제"""
    if request.method == 'POST':
        history = get_object_or_404(PointHistory, pk=pk)
        history.delete()
    return redirect('office_memberpoint_list')


# ============================================================
# 회원관리 > 탈퇴회원리스트
# ============================================================

@office_login_required
@office_permission_required('M')
def secession_list(request):
    """탈퇴회원 목록"""
    sch_name = request.GET.get('sch_member_name', '').strip()
    sch_id = request.GET.get('sch_member_id', '').strip()
    page = request.GET.get('page', '1')

    qs = OutMember.objects.all().order_by('-out_dt')
    if sch_name:
        qs = qs.filter(member_name__icontains=sch_name)
    if sch_id:
        qs = qs.filter(member_id__icontains=sch_id)

    paginator = Paginator(qs, 20)
    members = paginator.get_page(page)

    return render(request, 'ba_office/lfmember/secession_list.html', {
        'members': members,
        'sch_member_name': sch_name,
        'sch_member_id': sch_id,
        'total_count': paginator.count,
    })


# ============================================================
# 상담관리 > 상담 리스트
# ============================================================

def _get_consult_codes():
    """상담 관련 공통코드 조회 헬퍼"""
    return {
        'locd_codes': CodeValue.objects.filter(group_id='LOCD', del_chk='N').order_by('code_order'),
        'path_codes': CodeValue.objects.filter(group_id='PATH', del_chk='N').order_by('code_order'),
        'line_codes': CodeValue.objects.filter(group_id='LINE', del_chk='N').order_by('code_order'),
        'stat_codes': CodeValue.objects.filter(group_id='STAT', del_chk='N').order_by('code_order'),
        'cont_codes': CodeValue.objects.filter(group_id='CONT', del_chk='N').order_by('code_order'),
        'cust_codes': CodeValue.objects.filter(group_id='CUST', del_chk='N').order_by('code_order'),
    }


@office_login_required
@office_permission_required('C')
def consult_list(request):
    """상담 리스트 - 최근 1년 자동 로딩, 구장/담당코치 필터"""
    from datetime import date, timedelta
    from django.db.models import Prefetch

    page = request.GET.get('page', '1')
    sch_stadium = request.GET.get('sch_stadium', '')
    sch_coach = request.GET.get('sch_coach', '')

    # 최근 1년 기준
    date_to = date.today()
    date_from = date_to - timedelta(days=365)

    qs = Consult.objects.filter(
        del_chk='N',
        consult_dt__date__gte=date_from,
        consult_dt__date__lte=date_to,
    ).order_by('-id')

    if sch_stadium:
        qs = qs.filter(sta_code=sch_stadium)

    # 담당코치 필터: 코치 선택 시 해당 코치 OR 담당코치 없는 것
    if sch_coach:
        qs = qs.filter(
            Q(answers__coach_code=int(sch_coach)) | Q(answers__isnull=True) | Q(answers__coach_code__isnull=True)
        )

    qs = qs.distinct()
    qs = qs.prefetch_related(
        Prefetch('answers', queryset=ConsultAnswer.objects.order_by('-id'))
    )

    paginator = Paginator(qs, 20)
    consults = paginator.get_page(page)

    # 코드명 매핑
    codes = _get_consult_codes()
    stat_dict = {c.subcode: c.code_name for c in codes['stat_codes']}
    cont_dict = {c.subcode: c.code_name for c in codes['cont_codes']}
    cust_dict = {c.subcode: c.code_name for c in codes['cust_codes']}
    line_dict = {c.subcode: c.code_name for c in codes['line_codes']}

    sta_dict = {s.sta_code: s.sta_name for s in Stadium.objects.filter(use_gbn='Y')}
    coach_dict = {c.coach_code: c.coach_name for c in Coach.objects.filter(use_gbn='Y')}
    manage_dict = {u.office_id: u.office_name for u in OfficeUser.objects.filter(del_chk='N')}

    for con in consults:
        latest = con.answers.first()
        con.stat_name = stat_dict.get(latest.stat_code, '') if latest else ''
        con.cont_name = cont_dict.get(latest.consult_category, '') if latest else ''
        con.cust_name = cust_dict.get(latest.cus_stat_code, '') if latest else ''
        con.coach_name_display = coach_dict.get(latest.coach_code, '') if latest else ''
        con.line_name = line_dict.get(con.line_code, '')
        con.sta_name = sta_dict.get(int(con.sta_code), '') if con.sta_code and con.sta_code.isdigit() else ''
        con.manage_name = manage_dict.get(con.manage_id, '')
        con.member_gbn_name = '기존회원' if con.consult_gbn == 'old' else ('신규회원' if con.consult_gbn in ('new', 'guest') else '미가입자')

    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')

    return render(request, 'ba_office/lfconsult/consult_list.html', {
        'consults': consults,
        'total_count': paginator.count,
        'stadiums': stadiums,
        'coaches': coaches,
        'sch_stadium': sch_stadium,
        'sch_coach': sch_coach,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
    })


# ============================================================
# 상담관리 > 상담 상세/수정
# ============================================================

@office_login_required
@office_permission_required('C')
def consult_detail(request, pk):
    """상담 상세 보기/수정"""
    consult = get_object_or_404(Consult, pk=pk, del_chk='N')

    if request.method == 'POST':
        consult.consult_gbn = request.POST.get('consult_gbn', '')
        consult.member_id = request.POST.get('member_id', '')
        consult.child_id = request.POST.get('child_id', '')
        # [UX변경] 필드(권역) 직접입력 제거 → 선택 구장의 local_code에서 파생
        _scode = request.POST.get('sta_code', '')
        _lcode = request.POST.get('local_code', '')
        if _scode and not _lcode:
            _sta = Stadium.objects.filter(sta_code=int(_scode)).first()
            _lcode = str(_sta.local_code) if _sta and _sta.local_code else ''
        consult.local_code = _lcode
        consult.sta_code = _scode
        consult.consult_name = request.POST.get('consult_name', '')
        consult.consult_tel = request.POST.get('consult_tel', '')
        consult.stu_name = request.POST.get('stu_name', '')
        consult.stu_sex = request.POST.get('stu_sex', '')
        try:
            consult.stu_age = int(request.POST.get('stu_age', '0'))
        except ValueError:
            consult.stu_age = 0
        consult.path_code = int(request.POST.get('path_code', '44') or '44')
        consult.line_code = int(request.POST.get('line_code', '33') or '33')
        consult.consult_title = request.POST.get('consult_title', '')
        consult.consult_content = request.POST.get('consult_content', '')
        consult.save()
        return redirect('office_consult_detail', pk=pk)

    answers = consult.answers.all().order_by('id')
    codes = _get_consult_codes()
    # [UX변경] 권역 필터 제거 → 구장 직접선택(전체)
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')

    # 코치명 매핑
    coach_dict = {c.coach_code: c.coach_name for c in coaches}
    stat_dict = {c.subcode: c.code_name for c in codes['stat_codes']}
    cont_dict = {c.subcode: c.code_name for c in codes['cont_codes']}
    cust_dict = {c.subcode: c.code_name for c in codes['cust_codes']}

    for ans in answers:
        ans.coach_name_display = coach_dict.get(ans.coach_code, '')
        ans.stat_name = stat_dict.get(ans.stat_code, '')
        ans.cont_name = cont_dict.get(ans.consult_category, '')
        ans.cust_name = cust_dict.get(ans.cus_stat_code, '')

    ctx = {
        'consult': consult,
        'answers': answers,
        'stadiums': stadiums,
        'coaches': coaches,
    }
    ctx.update(codes)
    return render(request, 'ba_office/lfconsult/consult_detail.html', ctx)


# ============================================================
# 상담관리 > 답변 추가/수정/삭제
# ============================================================

@office_login_required
@office_permission_required('C')
def consult_answer_add(request):
    """답변 추가"""
    if request.method == 'POST':
        con_id = request.POST.get('con_id', '')
        consult = get_object_or_404(Consult, pk=con_id)
        office_user = request.session.get('office_user', {})

        ConsultAnswer.objects.create(
            consult=consult,
            consult_category=int(request.POST.get('consult_category', '0') or '0'),
            stat_code=int(request.POST.get('stat_code', '76') or '76'),
            cus_stat_code=int(request.POST.get('cus_stat_code', '0') or '0') or None,
            coach_code=int(request.POST.get('coach_code', '0') or '0') or None,
            consult_answer=request.POST.get('consult_answer', ''),
            receive_code=office_user.get('coach_code', None) or None,
            con_answer_dt=timezone.now(),
        )
        return redirect('office_consult_detail', pk=con_id)
    return redirect('office_consult_list')


@office_login_required
@office_permission_required('C')
def consult_answer_edit(request, pk):
    """답변 수정"""
    answer = get_object_or_404(ConsultAnswer, pk=pk)
    if request.method == 'POST':
        answer.consult_category = int(request.POST.get('consult_category', '0') or '0')
        answer.stat_code = int(request.POST.get('stat_code', '76') or '76')
        answer.cus_stat_code = int(request.POST.get('cus_stat_code', '0') or '0') or None
        answer.coach_code = int(request.POST.get('coach_code', '0') or '0') or None
        answer.consult_answer = request.POST.get('consult_answer', '')
        answer.con_answer_dt = timezone.now()
        answer.save()
    return redirect('office_consult_detail', pk=answer.consult_id)


@office_login_required
@office_permission_required('C')
def consult_answer_del(request, pk):
    """답변 삭제"""
    answer = get_object_or_404(ConsultAnswer, pk=pk)
    con_id = answer.consult_id
    if request.method == 'POST':
        answer.delete()
    return redirect('office_consult_detail', pk=con_id)


# ============================================================
# 상담관리 > 상담 등록
# ============================================================

@office_login_required
@office_permission_required('C')
def consult_input(request):
    """상담 등록"""
    office_user = request.session.get('office_user', {})

    if request.method == 'POST':
        # [UX변경] 필드(권역) 직접입력 제거 → 선택 구장의 local_code에서 파생
        sta_code = request.POST.get('sta_code', '')
        local_code = request.POST.get('local_code', '')
        if sta_code and not local_code:
            _sta = Stadium.objects.filter(sta_code=int(sta_code)).first()
            local_code = str(_sta.local_code) if _sta and _sta.local_code else ''
        consult = Consult.objects.create(
            consult_gbn=request.POST.get('consult_gbn', 'new'),
            member_id=request.POST.get('member_id', ''),
            member_name=request.POST.get('member_name', ''),
            child_id=request.POST.get('child_id', ''),
            child_name=request.POST.get('child_name', ''),
            local_code=local_code,
            sta_code=sta_code,
            consult_name=request.POST.get('consult_name', ''),
            consult_tel=request.POST.get('consult_tel', ''),
            stu_name=request.POST.get('stu_name', ''),
            stu_sex=request.POST.get('stu_sex', ''),
            stu_age=int(request.POST.get('stu_age', '0') or '0'),
            path_code=int(request.POST.get('path_code', '44') or '44'),
            line_code=int(request.POST.get('line_code', '33') or '33'),
            consult_title=request.POST.get('consult_title', ''),
            consult_content=request.POST.get('consult_content', ''),
            manage_id=office_user.get('office_id', ''),
            consult_dt=timezone.now(),
        )

        # 답변 동시 등록
        answer_content = request.POST.get('consult_answer', '').strip()
        if answer_content:
            ConsultAnswer.objects.create(
                consult=consult,
                consult_category=int(request.POST.get('answer_category', '0') or '0'),
                stat_code=int(request.POST.get('answer_stat', '76') or '76'),
                cus_stat_code=int(request.POST.get('answer_cust', '0') or '0') or None,
                coach_code=int(request.POST.get('answer_coach', '0') or '0') or None,
                consult_answer=answer_content,
                receive_code=office_user.get('coach_code', None) or None,
                con_answer_dt=timezone.now(),
            )

        return redirect('office_consult_list')

    codes = _get_consult_codes()
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')
    # [UX변경] 권역→구장 cascade 제거 → 구장 직접선택
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    ctx = {'coaches': coaches, 'stadiums': stadiums}
    ctx.update(codes)
    return render(request, 'ba_office/lfconsult/consult_input.html', ctx)


# ============================================================
# 상담관리 > AJAX 엔드포인트
# ============================================================

@office_login_required
def ajax_consult_stadium(request):
    """AJAX: 권역별 구장 목록"""
    local_code = request.GET.get('local_code', '')
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    if local_code:
        stadiums = stadiums.filter(local_code=int(local_code))
    data = [{'sta_code': s.sta_code, 'sta_name': s.sta_name} for s in stadiums]
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_consult_coach(request):
    """AJAX: 권역별 코치 목록"""
    local_code = request.GET.get('local_code', '')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('coach_name')
    if local_code:
        # 권역코드의 code_desc로 코치 dpart 매칭
        code_val = CodeValue.objects.filter(group_id='LOCD', subcode=int(local_code), del_chk='N').first()
        if code_val and code_val.code_desc:
            coaches = coaches.filter(dpart=code_val.code_desc)
    data = [{'coach_code': c.coach_code, 'coach_name': f'{c.coach_name}({c.dpart})'} for c in coaches]
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_consult_member_search(request):
    """AJAX: 회원 검색"""
    skey = request.GET.get('skey', 'member_id')
    sword = request.GET.get('sword', '').strip()
    if not sword:
        return JsonResponse([], safe=False)

    qs = Member.objects.filter(status='N', is_superuser=False)
    if skey == 'member_id':
        qs = qs.filter(username__icontains=sword)
    else:
        qs = qs.filter(name__icontains=sword)

    members = []
    for m in qs[:20]:
        child_cnt = MemberChild.objects.filter(parent=m, status='N').count()
        members.append({
            'member_id': m.username,
            'member_name': m.name,
            'phone': m.phone,
            'email': m.email,
            'address': m.address1,
            'child_cnt': child_cnt,
            'insert_dt': m.insert_dt.strftime('%Y-%m-%d') if m.insert_dt else '',
        })
    return JsonResponse(members, safe=False)


@office_login_required
def ajax_consult_child_list(request):
    """AJAX: 회원 자녀 목록"""
    member_id = request.GET.get('member_id', '')
    if not member_id:
        return JsonResponse([], safe=False)

    children = MemberChild.objects.filter(
        parent__username=member_id, status='N'
    ).values('child_id', 'name')
    return JsonResponse(list(children), safe=False)


# ============================================================
# 상담관리 > 상담권역설정
# ============================================================

@office_login_required
@office_permission_required('C')
def consult_local(request):
    """상담 권역설정"""
    regions = ConsultRegion.objects.filter(del_chk='N').order_by('id')
    return render(request, 'ba_office/lfconsult/consult_local.html', {
        'regions': regions,
    })


@office_login_required
@office_permission_required('C')
def consult_local_write(request):
    """권역 등록"""
    if request.method == 'POST':
        ConsultRegion.objects.create(
            reg_gbn=request.POST.get('reg_gbn', 'L'),
            reg_name=request.POST.get('reg_name', ''),
            mphone=request.POST.get('mphone', ''),
        )
    return redirect('office_consult_local')


@office_login_required
@office_permission_required('C')
def consult_local_modify(request, pk):
    """권역 수정"""
    region = get_object_or_404(ConsultRegion, pk=pk)
    if request.method == 'POST':
        region.reg_gbn = request.POST.get('reg_gbn', 'L')
        region.reg_name = request.POST.get('reg_name', '')
        region.mphone = request.POST.get('mphone', '')
        region.save()
    return redirect('office_consult_local')


@office_login_required
@office_permission_required('C')
def consult_local_del(request, pk):
    """권역 삭제"""
    if request.method == 'POST':
        region = get_object_or_404(ConsultRegion, pk=pk)
        region.del_chk = 'Y'
        region.save()
    return redirect('office_consult_local')


# ============================================================
# 상담관리 > 메인상담신청 (무료체험)
# ============================================================

@office_login_required
@office_permission_required('C')
def consult_free_list(request):
    """무료체험 신청 목록"""
    page = request.GET.get('page', '1')
    sch_confirm = request.GET.get('sch_confirm', '')
    sch_local = request.GET.get('sch_local', '')
    sch_name = request.GET.get('sch_name', '').strip()
    sch_phone = request.GET.get('sch_phone', '').strip()

    qs = ConsultFree.objects.filter(del_chk='N').order_by('-id')
    if sch_confirm:
        qs = qs.filter(confirm_yn=sch_confirm)
    if sch_local:
        qs = qs.filter(jlocal__icontains=sch_local)
    if sch_name:
        qs = qs.filter(jname__icontains=sch_name)
    if sch_phone:
        qs = qs.filter(Q(jphone2__icontains=sch_phone) | Q(jphone3__icontains=sch_phone))

    paginator = Paginator(qs, 20)
    frees = paginator.get_page(page)
    regions = ConsultRegion.objects.filter(del_chk='N', reg_gbn='L').order_by('reg_name')

    return render(request, 'ba_office/lfconsult/consult_free.html', {
        'frees': frees,
        'regions': regions,
        'total_count': paginator.count,
        'sch_confirm': sch_confirm,
        'sch_local': sch_local,
        'sch_name': sch_name,
        'sch_phone': sch_phone,
    })


@office_login_required
@office_permission_required('C')
def consult_free_confirm(request, pk):
    """무료체험 확인 처리"""
    if request.method == 'POST':
        free = get_object_or_404(ConsultFree, pk=pk)
        office_user = request.session.get('office_user', {})
        free.confirm_memo = request.POST.get('confirm_memo', '')
        free.confirm_yn = 'Y'
        free.confirm_id = office_user.get('office_id', '')
        free.confirm_name = office_user.get('office_name', '')
        free.confirm_date = timezone.now()
        free.save()
    return redirect('office_consult_free')


# ============================================================
# 수강생관리 > AJAX 엔드포인트
# ============================================================

@office_login_required
def ajax_student_local(request):
    """AJAX: 필드명(code_desc)별 지역(LOCD) 목록"""
    code_desc = request.GET.get('code_desc', '')
    qs = CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N')
    if code_desc:
        qs = qs.filter(code_desc=code_desc)
    qs = qs.order_by('code_desc', 'code_order')
    data = [{'subcode': c.subcode, 'code_name': c.code_name, 'code_desc': c.code_desc} for c in qs]
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_student_stadium(request):
    """AJAX: 필드명(grplocal_code=code_desc)별 구장 목록
    ASP sel_grpstadium.asp 패턴: code_desc → LOCD subcodes → stadiums
    """
    grplocal_code = request.GET.get('grplocal_code', '')
    qs = Stadium.objects.filter(use_gbn='Y')
    if grplocal_code:
        subcodes = CodeValue.objects.filter(
            group__grpcode='LOCD', del_chk='N', code_desc=grplocal_code
        ).values_list('subcode', flat=True)
        qs = qs.filter(local_code__in=subcodes)
    qs = qs.order_by('sta_name')
    data = [{'sta_code': s.sta_code, 'sta_name': s.sta_name} for s in qs]
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_student_course(request):
    """AJAX: 구장(sta_code)별 강좌 목록"""
    sta_code = request.GET.get('sta_code', '')
    qs = Lecture.objects.filter(stadium__sta_code=int(sta_code)) if sta_code else Lecture.objects.none()
    qs = qs.order_by('-use_gbn', 'lecture_day', 'class_gbn', 'lecture_time')
    data = []
    for l in qs:
        title = l.lecture_title
        if l.use_gbn == 'N':
            title = '(종료)' + title
        data.append({'lecture_code': l.lecture_code, 'lecture_title': title})
    return JsonResponse(data, safe=False)


# ============================================================
# 수강생관리 > 기본금액관리
# ============================================================

@office_login_required
@office_permission_required('H')
def under_development(request, page_title=''):
    """개발준비중 페이지 (미구현 메뉴 공통)"""
    return render(request, 'ba_office/under_development.html', {'page_title': page_title})


@office_login_required
@office_permission_required('H')
def dues_setting(request):
    """기본금액관리 - 교육용품비/패키지금액 설정"""
    setting = Setting.objects.first()
    if not setting:
        setting = Setting.objects.create(join_price=0, pk_price=0, insert_dt=timezone.now())

    if request.method == 'POST':
        setting.join_price = int(request.POST.get('join_price', '0') or '0')
        setting.pk_price = int(request.POST.get('pk_price', '0') or '0')
        setting.insert_id = request.session.get('office_user', {}).get('office_id', '')
        setting.insert_dt = timezone.now()
        setting.save()
        return redirect('office_dues_setting')

    return render(request, 'ba_office/lfstudent/dues_setting.html', {
        'setting': setting,
    })


# ============================================================
# 수강생관리 > 수강생정보(NEW) 목록
# ============================================================

def _get_default_ym():
    """기본 기준월 계산 (22일 이후면 다음달)"""
    now = datetime.now()
    y, m = now.year, now.month
    if now.day > 22:
        m += 1
        if m > 12:
            m = 1
            y += 1
    return f'{y}{m:02d}'


def _build_ym_choices():
    """기준월 드롭다운 옵션 (현재부터 과거 24개월)"""
    now = datetime.now()
    choices = []
    for i in range(-2, 25):
        y = now.year
        m = now.month - i + 2
        while m > 12:
            m -= 12
            y += 1
        while m < 1:
            m += 12
            y -= 1
        ym = f'{y}{m:02d}'
        choices.append(ym)
    # 중복 제거 후 역순 정렬
    choices = sorted(set(choices), reverse=True)
    return choices


LECTURE_STATS_TEXT = {
    'LY': '수강확정', 'LP': '수강예정', 'LN': '퇴단',
    'PN': '일시중지', 'LS': '중도취소',
}
PAY_STATS_TEXT = {
    'PP': '결제대기', 'PQ': '입금확인대기', 'PY': '결제완료',
    'PZ': '취소', 'UN': '미결제', 'PN': '결제대기',
}
APPLY_GUBUN_TEXT = {
    'NEW': '신규입단', 'RENEW': '재입단', 'AGAIN': '재수강', 'RE': '재등록',
}
PAY_METHOD_TEXT = {
    'CARD': '카드', 'R': '계좌이체', 'VACCT': '가상계좌',
    'ACCT': '무통장 입금', 'BENEFIT': '복지', 'ZEROPAY': '제로페이',
    'MUCU': '문화상품권', 'LOTTE': '롯데', 'EMART': '이마트',
    'ESNC': 'ESNC', 'GRPY': '그룹페이', 'SPBA': 'SPBA',
}
COURSE_STATE_TEXT = {
    'ING': '수강중', 'END': '퇴단', 'CAN': '미수강',
}


@office_login_required
@office_permission_required('H')
def student_list(request):
    """수강생정보 목록 (ASP lfstudent_list.asp)"""
    # 검색 파라미터
    sch_ym = request.GET.get('sch_ym', _get_default_ym())
    # [UX변경] 필드+구장 2단계→구장 직접선택 1단계. 원본: sch_code_desc = request.GET.get('sch_code_desc', '')
    sch_code_desc = ''
    sch_sta_code = request.GET.get('sch_sta_code', '')
    sch_lecture_code = request.GET.get('sch_lecture_code', '')
    sch_lecture_stats = request.GET.get('sch_lecture_stats', '')
    sch_pay_stats = request.GET.get('sch_pay_stats', '')
    sch_lecture_title = request.GET.get('sch_lecture_title', '')
    page = request.GET.get('page', '1')

    # [UX변경] 구장 선택이 조회 트리거. 원본: searched = bool(sch_code_desc)
    searched = bool(sch_sta_code)
    students = None
    total_count = 0
    wait_students = []
    wait_count = 0

    if searched:
        # 기준월 → date 변환
        try:
            ym_year = int(sch_ym[:4])
            ym_month = int(sch_ym[4:6])
            course_ym_date = date(ym_year, ym_month, 1)
        except (ValueError, IndexError):
            course_ym_date = date(datetime.now().year, datetime.now().month, 1)

        # 강좌 필터링을 위한 lecture_code 집합
        lecture_filter_codes = None
        if sch_lecture_code:
            lecture_filter_codes = [int(sch_lecture_code)]
        elif sch_sta_code:
            lecture_filter_codes = list(Lecture.objects.filter(
                stadium__sta_code=int(sch_sta_code)
            ).values_list('lecture_code', flat=True))
        # [UX변경] 원본: elif sch_code_desc: subcodes→lecture_filter_codes (필드→구장 cascade 제거)

        # 강좌명 텍스트 검색
        if sch_lecture_title:
            title_lec_codes = list(Lecture.objects.filter(
                lecture_title__icontains=sch_lecture_title
            ).values_list('lecture_code', flat=True))
            if lecture_filter_codes is not None:
                lecture_filter_codes = [c for c in lecture_filter_codes if c in title_lec_codes]
            else:
                lecture_filter_codes = title_lec_codes

        # EnrollmentCourse에서 해당 월의 no_seq 목록 추출
        courses_qs = EnrollmentCourse.objects.filter(
            bill_code='1001',
            course_ym=course_ym_date,
            enrollment__del_chk='N',
        )
        if lecture_filter_codes is not None:
            courses_qs = courses_qs.filter(lecture_code__in=lecture_filter_codes)

        enrollment_ids = courses_qs.values_list('enrollment_id', flat=True).distinct()

        # Enrollment 목록
        qs = Enrollment.objects.filter(id__in=enrollment_ids)
        if sch_lecture_stats:
            qs = qs.filter(lecture_stats=sch_lecture_stats)
        if sch_pay_stats:
            qs = qs.filter(pay_stats=sch_pay_stats)

        qs = qs.select_related('member', 'child').order_by('child__name', '-id')

        paginator = Paginator(qs, 30)
        students = paginator.get_page(page)
        total_count = paginator.count

        # 현재 페이지 enrollment들의 강좌정보 가져오기
        page_enrollment_ids = [e.id for e in students]
        course_entries = list(EnrollmentCourse.objects.filter(
            enrollment_id__in=page_enrollment_ids,
            bill_code='1001',
            course_ym=course_ym_date,
        ).order_by('lecture_code'))

        lec_codes = set(c.lecture_code for c in course_entries)
        lectures_map = {}
        if lec_codes:
            for l in Lecture.objects.filter(lecture_code__in=lec_codes).select_related('stadium'):
                lectures_map[l.lecture_code] = l

        # enrollment_id → [lecture_title, ...] (ASP는 다중 강좌를 <br/>로 연결)
        enrollment_lectures = {}
        for c in course_entries:
            lec = lectures_map.get(c.lecture_code)
            if lec:
                enrollment_lectures.setdefault(c.enrollment_id, []).append(lec.lecture_title)

        # isLP 배치 체크: 재수강 대상자 = end_dt == sch_ym AND 다음달 enrollment 없음
        check_pairs = []
        for e in students:
            if e.end_dt == sch_ym:
                try:
                    end_y, end_m = int(e.end_dt[:4]), int(e.end_dt[4:6])
                    end_m += 1
                    if end_m > 12:
                        end_m, end_y = 1, end_y + 1
                    next_ym = f'{end_y}{end_m:02d}'
                    check_pairs.append((e.member_id, e.child_id, next_ym))
                except (ValueError, IndexError):
                    pass

        has_next_set = set()
        if check_pairs:
            q = Q()
            for mid, cid, nym in check_pairs:
                q |= Q(member_id=mid, child_id=cid, start_dt=nym)
            for mid, cid, sdt in Enrollment.objects.filter(
                q, lecture_stats__in=['LY', 'LP', 'PN', 'LN']
            ).values_list('member_id', 'child_id', 'start_dt'):
                has_next_set.add((mid, cid, sdt))

        # 각 enrollment에 표시 텍스트 부착
        for e in students:
            e.lecture_title_display = '<br/>'.join(enrollment_lectures.get(e.id, []))
            e.apply_gubun_text = APPLY_GUBUN_TEXT.get(e.apply_gubun, e.apply_gubun)
            e.lecture_stats_text = LECTURE_STATS_TEXT.get(e.lecture_stats, e.lecture_stats)
            e.pay_stats_text = PAY_STATS_TEXT.get(e.pay_stats, e.pay_stats)
            e.pay_method_text = PAY_METHOD_TEXT.get(e.pay_method, e.pay_method)
            # 재수강 대상자 여부
            e.is_pink = False
            if e.end_dt == sch_ym:
                try:
                    end_y, end_m = int(e.end_dt[:4]), int(e.end_dt[4:6])
                    end_m += 1
                    if end_m > 12:
                        end_m, end_y = 1, end_y + 1
                    next_ym = f'{end_y}{end_m:02d}'
                    if (e.member_id, e.child_id, next_ym) not in has_next_set:
                        e.is_pink = True
                except (ValueError, IndexError):
                    pass

        # 대기등록자 (ASP와 동일 필터 적용)
        wait_qs = WaitStudent.objects.filter(trans_gbn='N', del_chk='N')
        # [UX변경] 원본: if sch_code_desc: wait_qs.filter(local_code__in=sub_codes) (필드 필터 제거)
        if sch_sta_code:
            wait_qs = wait_qs.filter(sta_code=int(sch_sta_code))
        if sch_lecture_code:
            wait_qs = wait_qs.filter(lecture_code=int(sch_lecture_code))
        wait_students = list(wait_qs.order_by('id'))
        wait_count = len(wait_students)

        # 대기등록자 추가 정보 (전화번호, 회원상태)
        if wait_students:
            w_member_ids = set(w.member_id for w in wait_students)
            w_child_ids = set(w.child_id for w in wait_students)
            member_phones = dict(
                Member.objects.filter(username__in=w_member_ids).values_list('username', 'phone'))
            child_states = dict(
                MemberChild.objects.filter(child_id__in=w_child_ids).values_list('child_id', 'course_state'))
            for w in wait_students:
                w.phone = member_phones.get(w.member_id, '')
                w.course_state_display = COURSE_STATE_TEXT.get(
                    child_states.get(w.child_id, ''), '')

    # [UX변경] 원본: code_descs = CodeValue.objects.filter(group__grpcode='LOCD'...) (필드 목록 제거)

    # [UX변경] 구장 전체 직접 표시. 원본: stadiums=[], if sch_code_desc: filter by local_code
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    courses = []
    if sch_sta_code:
        courses = Lecture.objects.filter(stadium__sta_code=int(sch_sta_code)).order_by('lecture_title')

    return render(request, 'ba_office/lfstudent/student_list.html', {
        'searched': searched,
        'students': students,
        'total_count': total_count,
        'wait_students': wait_students,
        'wait_count': wait_count,
        'stadiums': stadiums,
        'courses': courses,
        'sch_ym': sch_ym,
        'sch_sta_code': sch_sta_code,
        'sch_lecture_code': sch_lecture_code,
        'sch_lecture_stats': sch_lecture_stats,
        'sch_pay_stats': sch_pay_stats,
        'sch_lecture_title': sch_lecture_title,
    })


@office_login_required
@office_permission_required('H')
def lec_attendance_excel(request):
    """수강생 리스트 > 출석부출력 (ASP report/lec_attendance.asp)

    특정 강좌(lecture_code)·기준월(sch_ym)의 출석부를 Excel로 다운로드.
    학생별 행 + 수업일자별 출결(Y:출석/R:우천취소/N:결석/D:수업연기) 동적 컬럼.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        lecture_code = int(request.GET.get('lecture_code', '0') or '0')
    except ValueError:
        lecture_code = 0
    sch_ym = request.GET.get('sch_ym', '').strip()

    # 필수값 검증 (원본: lecture_code=0 또는 month=0 이면 AlertBack)
    if not lecture_code or not (len(sch_ym) == 6 and sch_ym.isdigit()):
        return HttpResponse('<script>alert("필수정보가 부족합니다.");history.back();</script>')

    year, month = int(sch_ym[:4]), int(sch_ym[4:6])
    try:
        course_ym_date = date(year, month, 1)
    except ValueError:
        return HttpResponse('<script>alert("기준월이 올바르지 않습니다.");history.back();</script>')

    # 강좌 정보 (구장명, 강좌명)
    lecture = Lecture.objects.filter(lecture_code=lecture_code).select_related('stadium').first()
    if not lecture:
        return HttpResponse('<script>alert("수업정보가 없습니다.");history.back();</script>')
    sta_name = lecture.stadium.sta_name if lecture.stadium_id else ''
    lec_name = lecture.lecture_title

    # 수업일 목록 (LectureSelDay) - 일자 오름차순 (원본은 정렬 없음 → 출석부 가독성 위해 일자순)
    sel_days = list(LectureSelDay.objects.filter(
        lecture_code=lecture_code, syear=year, smonth=month,
    ).values_list('sday', flat=True).distinct().order_by('sday'))

    # 학생 목록: bill_code='1001', 해당 월·강좌, 수강상태 IN(LY,LP,PN)
    course_qs = EnrollmentCourse.objects.filter(
        bill_code='1001',
        course_ym=course_ym_date,
        lecture_code=lecture_code,
        enrollment__lecture_stats__in=['LY', 'LP', 'PN'],
    ).select_related(
        'enrollment', 'enrollment__member', 'enrollment__child',
    ).order_by('enrollment__child__name')

    # child_id별 1행 (원본 temp_id 중복 제거)
    seen, students = set(), []
    for c in course_qs:
        cid = c.enrollment.child_id
        if cid in seen:
            continue
        seen.add(cid)
        students.append(c)

    # 출결 매핑: (child_id, day) -> attendance_gbn
    # 원본 sql_gbn은 child_id+날짜만 조회하나, 출석부=해당 강좌이므로 lecture_code 포함(정확도↑)
    att_map = {}
    ym_prefix = f'{year:04d}-{month:02d}'
    for cid, adt, gbn in Attendance.objects.filter(
        lecture_code=lecture_code, attendance_dt__startswith=ym_prefix,
    ).values_list('child_id', 'attendance_dt', 'attendance_gbn'):
        try:
            att_map[(cid, int(adt[8:10]))] = gbn
        except (ValueError, IndexError):
            continue

    # ── 엑셀 생성 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '출석부'

    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center')

    # 상단 정보 (원본 헤더 테이블 재현)
    ws.cell(row=1, column=1, value='AAFC 출석부').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'구장명: {sta_name}')
    ws.cell(row=2, column=4, value=f'기준월: {sch_ym}')
    ws.cell(row=3, column=1, value=f'CLASS: {lec_name}')
    ws.cell(row=3, column=4, value='Y:출석 / R:우천취소 / N:결석 / D:수업연기')

    base_headers = ['부모명', '자녀명', '자녀아이디', '카드번호', '전화번호',
                    '입단구분', '수강상태', '결제금액', '수강기간', '시작월',
                    '종료월', '수업시작일자', '결제상태', '결제방법', '결제일']
    headers = base_headers + [f'{d}일' for d in sel_days]
    HROW = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=HROW, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = thin

    r = HROW + 1
    for c in students:
        e = c.enrollment
        member, child = e.member, e.child
        row_vals = [
            member.name if member else '',
            child.name if child else '',
            e.child_id,
            child.card_num if child else '',
            member.phone if member else '',
            APPLY_GUBUN_TEXT.get(e.apply_gubun, e.apply_gubun),
            LECTURE_STATS_TEXT.get(e.lecture_stats, e.lecture_stats),
            e.pay_price,
            e.lec_period,
            e.start_dt,
            e.end_dt,
            c.start_ymd.strftime('%Y%m%d') if c.start_ymd else '',
            PAY_STATS_TEXT.get(e.pay_stats, e.pay_stats),
            PAY_METHOD_TEXT.get(e.pay_method, e.pay_method),
            e.pay_dt.strftime('%Y-%m-%d') if e.pay_dt else '',
        ]
        for d in sel_days:
            row_vals.append(att_map.get((e.child_id, d), ''))
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin
            if col > len(base_headers):
                cell.alignment = center
        r += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={lecture_code}_{sch_ym}.xlsx'
    wb.save(response)
    return response


# ============================================================
# 수강생관리 > 수강생조회(이름검색)
# ============================================================

@office_login_required
@office_permission_required('H')
def student_search(request):
    """수강생조회 - 이름검색"""
    sch_ym = request.GET.get('sch_ym', _get_default_ym())
    sch_child_name = request.GET.get('sch_child_name', '')
    results = []

    if sch_child_name and len(sch_child_name) >= 2:
        try:
            ym_year = int(sch_ym[:4])
            ym_month = int(sch_ym[4:6])
            course_ym_date = date(ym_year, ym_month, 1)
        except (ValueError, IndexError):
            course_ym_date = date(datetime.now().year, datetime.now().month, 1)

        enrollment_ids = EnrollmentCourse.objects.filter(
            bill_code='1001',
            course_ym=course_ym_date,
        ).values_list('enrollment_id', flat=True).distinct()

        qs = Enrollment.objects.filter(
            id__in=enrollment_ids,
            del_chk='N',
            child__name__icontains=sch_child_name,
        ).select_related('member', 'child').order_by('child__name')

        # 각 enrollment에 강좌명, 카드번호, 상태 텍스트 부여
        # 재수강 대상 판정: 종료월 == 검색월 AND 다음달 수강 없음
        next_ym = date(ym_year + (1 if ym_month == 12 else 0),
                       1 if ym_month == 12 else ym_month + 1, 1)
        next_enrolled_ids = set(EnrollmentCourse.objects.filter(
            bill_code='1001', course_ym=next_ym, enrollment__del_chk='N',
            enrollment__lecture_stats='LY',
        ).values_list('enrollment__child_id', flat=True))

        for e in qs:
            # 강좌명 조합
            lec_codes = EnrollmentCourse.objects.filter(
                enrollment=e, bill_code='1001', course_ym=course_ym_date,
            ).values_list('lecture_code', flat=True)
            lec_titles = list(Lecture.objects.filter(
                lecture_code__in=lec_codes
            ).values_list('lecture_title', flat=True))
            e.lecture_title_display = '<br/>'.join(lec_titles) if lec_titles else ''
            e.card_num = e.child.card_num if e.child else ''
            e.lecture_stats_text = LECTURE_STATS_TEXT.get(e.lecture_stats, e.lecture_stats)
            e.pay_stats_text = PAY_STATS_TEXT.get(e.pay_stats, e.pay_stats)
            e.pay_method_text = PAY_METHOD_TEXT.get(e.pay_method, e.pay_method)
            # 핑크: 종료월이 검색월이고 다음달 수강 없음
            e.is_pink = (e.end_dt == sch_ym and e.child_id not in next_enrolled_ids)

        results = qs

    return render(request, 'ba_office/lfstudent/student_search.html', {
        'results': results,
        'ym_choices': _build_ym_choices(),
        'sch_ym': sch_ym,
        'sch_child_name': sch_child_name,
    })


# ============================================================
# 수강생관리 > 수강생상세
# ============================================================

@office_login_required
@office_permission_required('H')
def student_detail(request, no_seq):
    """수강생상세 보기/수정"""
    enrollment = get_object_or_404(Enrollment, id=no_seq)
    member = enrollment.member
    child = enrollment.child
    office_user = request.session.get('office_user', {})

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'update':
            # 결제완료(PY) 건을 취소(중도취소 LN / 결제취소 PZ)로 바꾸면 Toss 실제 환불 먼저 (실패 시 중단)
            if enrollment.pay_stats == 'PY' and (request.POST.get('lecture_stats') == 'LN' or request.POST.get('pay_stats') == 'PZ'):
                _ok, _msg = _refund_enrollment_toss(enrollment)
                if not _ok:
                    return HttpResponse('<script>alert("Toss 환불 실패: ' + _msg + '\\n취소를 중단합니다. 확인 후 다시 시도하세요.");history.back();</script>')
            with transaction.atomic():
                new_lecture_stats = request.POST.get('lecture_stats', enrollment.lecture_stats)
                new_pay_stats = request.POST.get('pay_stats', enrollment.pay_stats)
                new_pay_method = request.POST.get('pay_method', enrollment.pay_method)
                new_apply_gubun = request.POST.get('apply_gubun', enrollment.apply_gubun)
                new_bigo = request.POST.get('bigo_content', enrollment.bigo_content)
                new_account_no = request.POST.get('account_no', enrollment.account_no)
                cancel_code = request.POST.get('cancel_code', '')
                cancel_desc = request.POST.get('cancel_desc', '')

                # 수강상태 변경 처리
                if new_lecture_stats != enrollment.lecture_stats:
                    if new_lecture_stats == 'LY':
                        EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='LY')
                        MemberChild.objects.filter(child_id=child.child_id).update(course_state='ING')
                    elif new_lecture_stats == 'PN':
                        EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='PN')
                        MemberChild.objects.filter(child_id=child.child_id).update(course_state='PAU')
                        enrollment.cancel_code = cancel_code
                        enrollment.cancel_desc = cancel_desc
                    elif new_lecture_stats == 'LN':
                        EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='LN')
                        EnrollmentBill.objects.filter(enrollment=enrollment).update(pay_stats='PZ')
                        MemberChild.objects.filter(child_id=child.child_id).update(course_state='END')
                        enrollment.cancel_code = cancel_code
                        enrollment.cancel_desc = cancel_desc
                        if not enrollment.cancel_date:
                            enrollment.cancel_date = timezone.now()
                        new_pay_stats = 'PZ'
                    elif new_lecture_stats == 'LP':
                        EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='LP')

                # 결제상태 변경 처리
                if new_pay_stats != enrollment.pay_stats:
                    EnrollmentBill.objects.filter(enrollment=enrollment).update(pay_stats=new_pay_stats)
                    if new_pay_stats == 'PY' and not enrollment.pay_dt:
                        enrollment.pay_dt = timezone.now()

                enrollment.lecture_stats = new_lecture_stats
                enrollment.pay_stats = new_pay_stats
                enrollment.pay_method = new_pay_method
                enrollment.apply_gubun = new_apply_gubun
                enrollment.bigo_content = new_bigo
                enrollment.account_no = new_account_no
                enrollment.save()

            return redirect('office_student_detail', no_seq=no_seq)

        elif action == 'delete':
            with transaction.atomic():
                EnrollmentCourse.objects.filter(enrollment=enrollment).delete()
                EnrollmentBill.objects.filter(enrollment=enrollment).delete()
                enrollment.delete()
            return redirect('office_student_list')

    # 수강과정 목록
    course_entries = list(EnrollmentCourse.objects.filter(
        enrollment=enrollment, bill_code='1001'
    ).order_by('-course_ym'))

    # 각 과정행에 수강상태 라벨 + 수업횟수 부여 (원본 lfstudent_detail.asp 과정정보 테이블)
    #   수업횟수 cnt = 시작월 중 시작일 이후 수업일 수 (ASP: COUNT(lf_lecture_selday.sday))
    for c in course_entries:
        c.course_stats_text = LECTURE_STATS_TEXT.get(c.course_stats, c.course_stats)
        if c.start_ymd:
            c.cnt = LectureSelDay.objects.filter(
                lecture_code=c.lecture_code,
                syear=c.start_ymd.year,
                smonth=c.start_ymd.month,
                sday__gte=c.start_ymd.day,
            ).count()
        else:
            c.cnt = 0

    # 강좌 정보 매핑
    lec_codes = set(c.lecture_code for c in course_entries)
    lectures_map = {}
    if lec_codes:
        for l in Lecture.objects.filter(lecture_code__in=lec_codes).select_related('stadium'):
            lectures_map[l.lecture_code] = l

    # 청구내역
    bills = EnrollmentBill.objects.filter(enrollment=enrollment).order_by('bill_code')

    # 셔틀비 정보
    shuttle_bill = EnrollmentBill.objects.filter(enrollment=enrollment, bill_code='1009').first()

    # 청구 합계
    bill_total = sum(b.bill_amt for b in bills if b.bill_amt)

    # 알림글 이력
    notifications = Notification.objects.filter(
        child_id=child.child_id, del_chk='N'
    ).order_by('-insert_dt')[:20]

    # 취소사유 코드 (RESN 그룹)
    cancel_reasons = CodeValue.objects.filter(
        group__grpcode='RESN', del_chk='N'
    ).order_by('code_order')

    # 출결이력 캘린더 (ASP: 월별 31일 그리드, O/X)
    att_records = Attendance.objects.filter(
        child_id=child.child_id
    ).order_by('attendance_dt')

    # 월별 피벗: {YYYY-MM: {day: 'O'|'X'}}
    att_calendar = {}
    for a in att_records:
        if len(a.attendance_dt) >= 10:
            att_mon = a.attendance_dt[:7]  # YYYY-MM
            try:
                att_day = int(a.attendance_dt[8:10])
            except (ValueError, IndexError):
                continue
            if att_mon not in att_calendar:
                att_calendar[att_mon] = {}
            # Y(출석), A(보강) → O, 나머지 → X
            att_calendar[att_mon][att_day] = 'O' if a.attendance_gbn in ('Y', 'A') else 'X'

    # 정렬된 월 목록
    att_months = sorted(att_calendar.keys())
    att_rows = []
    for mon in att_months:
        days = att_calendar[mon]
        row = [days.get(d, '') for d in range(1, 32)]
        att_rows.append({'month': mon, 'days': row})

    # 성별 표시
    gender_text = {'M': '남자', 'F': '여자'}.get(child.gender, '')

    # 등록처구분 표시
    source_gubun_text = {
        '01': '온라인', '02': '오프라인', 'M': '모바일',
    }.get(enrollment.source_gubun, enrollment.source_gubun)

    # 텍스트 표시용 (기본 모드에서 select 대신 텍스트 표시)
    apply_gubun_text = APPLY_GUBUN_TEXT.get(enrollment.apply_gubun, enrollment.apply_gubun)
    lecture_stats_text = LECTURE_STATS_TEXT.get(enrollment.lecture_stats, enrollment.lecture_stats)
    pay_stats_text = PAY_STATS_TEXT.get(enrollment.pay_stats, enrollment.pay_stats)
    pay_method_text = PAY_METHOD_TEXT.get(enrollment.pay_method, enrollment.pay_method)

    return render(request, 'ba_office/lfstudent/student_detail.html', {
        'enrollment': enrollment,
        'member': member,
        'child': child,
        'course_entries': course_entries,
        'lectures_map': lectures_map,
        'bills': bills,
        'bill_total': bill_total,
        'shuttle_bill': shuttle_bill,
        'notifications': notifications,
        'cancel_reasons': cancel_reasons,
        'att_rows': att_rows,
        'gender_text': gender_text,
        'source_gubun_text': source_gubun_text,
        'apply_gubun_text': apply_gubun_text,
        'lecture_stats_text': lecture_stats_text,
        'pay_stats_text': pay_stats_text,
        'pay_method_text': pay_method_text,
    })


@office_login_required
@office_permission_required('H')
def student_shuttle_proc(request):
    """셔틀비 관리 처리 (AJAX POST)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST만 허용'}, status=405)

    no_seq = int(request.POST.get('no_seq', '0'))
    gubun = request.POST.get('gubun', '')  # insert, update, delete
    shuttle_amt = int(request.POST.get('shuttle_amt', '0') or '0')

    enrollment = get_object_or_404(Enrollment, id=no_seq)

    with transaction.atomic():
        existing = EnrollmentBill.objects.filter(enrollment=enrollment, bill_code='1009').first()

        if gubun == 'insert' and not existing:
            EnrollmentBill.objects.create(
                enrollment=enrollment,
                bill_code='1009',
                bill_desc='차량이용료',
                bill_amt=shuttle_amt,
                pay_stats='PP',
            )
            enrollment.shuttle_yn = 'Y'
            enrollment.pay_price = enrollment.pay_price + shuttle_amt
            enrollment.save()

        elif gubun == 'update' and existing:
            old_amt = existing.bill_amt
            existing.bill_amt = shuttle_amt
            existing.save()
            enrollment.pay_price = enrollment.pay_price - old_amt + shuttle_amt
            enrollment.save()

        elif gubun == 'delete' and existing:
            old_amt = existing.bill_amt
            existing.delete()
            enrollment.shuttle_yn = ''
            enrollment.pay_price = enrollment.pay_price - old_amt
            enrollment.save()

    return JsonResponse({'result': 'ok'})


@office_login_required
@office_permission_required('H')
def student_alim_proc(request):
    """알림글 등록/수정 (AJAX POST) - ASP lfstudent_alim_proc.asp"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST만 허용'}, status=405)

    mode = request.POST.get('mode', '')  # insert, mod
    no_seq = request.POST.get('no_seq', '')
    child_id = request.POST.get('child_id', '')
    member_id = request.POST.get('member_id', '')
    member_name = request.POST.get('member_name', '')
    alim_title = request.POST.get('alim_title', '').strip()
    alim_content = request.POST.get('alim_content', '').strip()
    office_user = request.session.get('office_user', {})

    if not alim_title or not alim_content:
        return JsonResponse({'error': '제목과 내용을 입력하세요.'}, status=400)

    if mode == 'insert':
        # 새 no_seq 생성
        max_seq = Notification.objects.order_by('-no_seq').values_list('no_seq', flat=True).first()
        new_seq = (max_seq or 0) + 1
        Notification.objects.create(
            no_seq=new_seq,
            alim_gbn='P',
            member_id=member_id,
            member_name=member_name,
            child_id=child_id,
            alim_title=alim_title,
            alim_content=alim_content,
            insert_id=office_user.get('office_id', ''),
            insert_name=office_user.get('office_name', ''),
            insert_dt=timezone.now(),
        )
    elif mode == 'mod' and no_seq:
        try:
            noti = Notification.objects.get(no_seq=int(no_seq))
            noti.alim_title = alim_title
            noti.alim_content = alim_content
            noti.save()
        except Notification.DoesNotExist:
            return JsonResponse({'error': '알림글을 찾을 수 없습니다.'}, status=404)

    return JsonResponse({'result': 'ok'})


# ============================================================
# 수강생관리 > 입단신청내역
# ============================================================

@office_login_required
@office_permission_required('H')
def master_list(request):
    """입단신청내역 목록"""
    from datetime import date as _date
    today_str = _date.today().strftime('%Y-%m-%d')

    # [UX변경] 원본: sch_code_desc = request.GET.get('sch_code_desc', '') (필드 제거)
    sch_sta_code = request.GET.get('sch_sta_code', '')
    sch_apply_gubun = request.GET.get('sch_apply_gubun', '')
    sch_lecture_stats = request.GET.get('sch_lecture_stats', '')
    sch_pay_method = request.GET.get('sch_pay_method', '')
    sch_pay_stats = request.GET.get('sch_pay_stats', '')
    sch_sdate = request.GET.get('sch_sdate', '') or today_str
    sch_edate = request.GET.get('sch_edate', '') or today_str
    sch_skey = request.GET.get('sch_skey', 'child_name')
    sch_sword = request.GET.get('sch_sword', '')
    page = request.GET.get('page', '1')

    # 날짜는 항상 기본값이 있으므로 searched는 날짜 포함 항상 True
    searched = True

    qs = Enrollment.objects.filter(del_chk='N').select_related('member', 'child')

    # 구장 필터
    if sch_sta_code:
        lec_codes = Lecture.objects.filter(
            stadium__sta_code=int(sch_sta_code)
        ).values_list('lecture_code', flat=True)
        enrollment_ids = EnrollmentCourse.objects.filter(
            lecture_code__in=lec_codes
        ).values_list('enrollment_id', flat=True).distinct()
        qs = qs.filter(id__in=enrollment_ids)
    # [UX변경] 원본: elif sch_code_desc: local_code→lec_codes 필터 (필드→구장 cascade 제거)

    if sch_apply_gubun:
        qs = qs.filter(apply_gubun=sch_apply_gubun)
    if sch_lecture_stats:
        qs = qs.filter(lecture_stats=sch_lecture_stats)
    if sch_pay_method:
        qs = qs.filter(pay_method=sch_pay_method)
    if sch_pay_stats:
        qs = qs.filter(pay_stats=sch_pay_stats)

    # 기간 필터 (ASP 동일: pay_stats=PY이면 pay_dt, lecture_stats=LN이면 cancel_date, 나머지는 insert_dt)
    if sch_sdate or sch_edate:
        if sch_pay_stats == 'PY':
            date_field = 'pay_dt__date'
        elif sch_lecture_stats == 'LN':
            date_field = 'cancel_date__date'
        else:
            date_field = 'insert_dt__date'
        if sch_sdate:
            qs = qs.filter(**{f'{date_field}__gte': sch_sdate})
        if sch_edate:
            qs = qs.filter(**{f'{date_field}__lte': sch_edate})

    # 텍스트 검색
    if sch_sword:
        if sch_skey == 'member_id':
            qs = qs.filter(member__username__icontains=sch_sword)
        elif sch_skey == 'member_name':
            qs = qs.filter(member__name__icontains=sch_sword)
        elif sch_skey == 'child_id':
            qs = qs.filter(child__child_id__icontains=sch_sword)
        else:  # child_name
            qs = qs.filter(child__name__icontains=sch_sword)

    qs = qs.order_by('-insert_dt')

    paginator = Paginator(qs, 30)
    masters = paginator.get_page(page)
    total_count = paginator.count

    # 각 enrollment에 구장명 부여
    for e in masters:
        ec = EnrollmentCourse.objects.filter(enrollment=e, bill_code='1001').first()
        if ec:
            lec = Lecture.objects.filter(lecture_code=ec.lecture_code).select_related('stadium').first()
            e.sta_name = lec.stadium.sta_nickname if lec and lec.stadium else ''
        else:
            e.sta_name = ''

    # [UX변경] 구장 전체 직접 표시. 원본: stadiums=[], if sch_code_desc: filter by local_code
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    # [UX변경] 원본: code_descs = CodeValue.objects.filter(group__grpcode='LOCD'...) (필드 목록 제거)

    return render(request, 'ba_office/lfstudent/master_list.html', {
        'searched': searched,
        'masters': masters,
        'total_count': total_count,
        'stadiums': stadiums,
        'sch_sta_code': sch_sta_code,
        'sch_apply_gubun': sch_apply_gubun,
        'sch_lecture_stats': sch_lecture_stats,
        'sch_pay_method': sch_pay_method,
        'sch_pay_stats': sch_pay_stats,
        'sch_sdate': sch_sdate,
        'sch_edate': sch_edate,
        'sch_skey': sch_skey,
        'sch_sword': sch_sword,
    })


@office_login_required
@office_permission_required('H')
def master_detail(request, no_seq):
    """입단신청상세 보기/수정"""
    enrollment = get_object_or_404(Enrollment, id=no_seq)
    member = enrollment.member
    child = enrollment.child
    office_user = request.session.get('office_user', {})

    if request.method == 'POST':
        # 결제완료(PY) 건을 취소(중도취소 LN / 결제취소 PZ)로 바꾸면 Toss 실제 환불 먼저 (실패 시 중단)
        if enrollment.pay_stats == 'PY' and (request.POST.get('lecture_stats') == 'LN' or request.POST.get('pay_stats') == 'PZ'):
            _ok, _msg = _refund_enrollment_toss(enrollment)
            if not _ok:
                return HttpResponse('<script>alert("Toss 환불 실패: ' + _msg + '\\n취소를 중단합니다. 확인 후 다시 시도하세요.");history.back();</script>')
        with transaction.atomic():
            new_lecture_stats = request.POST.get('lecture_stats', enrollment.lecture_stats)
            new_pay_stats = request.POST.get('pay_stats', enrollment.pay_stats)
            new_pay_method = request.POST.get('pay_method', enrollment.pay_method)
            new_apply_gubun = request.POST.get('apply_gubun', enrollment.apply_gubun)
            new_account_no = request.POST.get('account_no', enrollment.account_no)
            cancel_code = request.POST.get('cancel_code', '')
            cancel_desc = request.POST.get('cancel_desc', '')

            if new_lecture_stats != enrollment.lecture_stats:
                if new_lecture_stats == 'LY':
                    EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='LY')
                    MemberChild.objects.filter(child_id=child.child_id).update(course_state='ING')
                elif new_lecture_stats == 'PN':
                    EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='PN')
                    MemberChild.objects.filter(child_id=child.child_id).update(course_state='PAU')
                    enrollment.cancel_code = cancel_code
                    enrollment.cancel_desc = cancel_desc
                elif new_lecture_stats == 'LN':
                    EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='LN')
                    EnrollmentBill.objects.filter(enrollment=enrollment).update(pay_stats='PZ')
                    MemberChild.objects.filter(child_id=child.child_id).update(course_state='END')
                    enrollment.cancel_code = cancel_code
                    enrollment.cancel_desc = cancel_desc
                    if not enrollment.cancel_date:
                        enrollment.cancel_date = timezone.now()
                    new_pay_stats = 'PZ'
                elif new_lecture_stats == 'LP':
                    EnrollmentCourse.objects.filter(enrollment=enrollment).update(course_stats='LP')

            if new_pay_stats != enrollment.pay_stats:
                EnrollmentBill.objects.filter(enrollment=enrollment).update(pay_stats=new_pay_stats)
                if new_pay_stats == 'PY' and not enrollment.pay_dt:
                    enrollment.pay_dt = timezone.now()

            enrollment.lecture_stats = new_lecture_stats
            enrollment.pay_stats = new_pay_stats
            enrollment.pay_method = new_pay_method
            enrollment.apply_gubun = new_apply_gubun
            enrollment.account_no = new_account_no
            enrollment.save()

        return redirect('office_master_detail', no_seq=no_seq)

    # 수강과정 목록
    course_entries = EnrollmentCourse.objects.filter(
        enrollment=enrollment, bill_code='1001'
    ).order_by('-course_ym')
    lec_codes = set(c.lecture_code for c in course_entries)
    lectures_map = {}
    if lec_codes:
        for l in Lecture.objects.filter(lecture_code__in=lec_codes).select_related('stadium'):
            lectures_map[l.lecture_code] = l

    # 청구내역
    bills = EnrollmentBill.objects.filter(enrollment=enrollment).order_by('bill_code')
    bill_total = sum(b.bill_amt for b in bills)

    # 취소사유 코드
    cancel_reasons = CodeValue.objects.filter(
        group__grpcode='RESN', del_chk='N'
    ).order_by('code_order')

    # 등록처구분
    SOURCE_GUBUN_TEXT = {
        '01': '홈페이지', '02': '관리자', '03': '모바일',
    }

    return render(request, 'ba_office/lfstudent/master_detail.html', {
        'enrollment': enrollment,
        'member': member,
        'child': child,
        'course_entries': course_entries,
        'lectures_map': lectures_map,
        'bills': bills,
        'bill_total': bill_total,
        'cancel_reasons': cancel_reasons,
        'source_gubun_text': SOURCE_GUBUN_TEXT.get(enrollment.source_gubun, enrollment.source_gubun),
    })


@office_login_required
@office_permission_required('H')
def master_list_excel(request):
    """입단신청내역 Excel 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sch_code_desc = request.GET.get('sch_code_desc', '')
    sch_sta_code = request.GET.get('sch_sta_code', '')
    sch_apply_gubun = request.GET.get('sch_apply_gubun', '')
    sch_lecture_stats = request.GET.get('sch_lecture_stats', '')
    sch_pay_method = request.GET.get('sch_pay_method', '')
    sch_pay_stats = request.GET.get('sch_pay_stats', '')
    sch_sdate = request.GET.get('sch_sdate', '')
    sch_edate = request.GET.get('sch_edate', '')
    sch_skey = request.GET.get('sch_skey', 'child_name')
    sch_sword = request.GET.get('sch_sword', '')

    qs = Enrollment.objects.filter(del_chk='N').select_related('member', 'child')

    if sch_sta_code:
        lec_codes = Lecture.objects.filter(stadium__sta_code=int(sch_sta_code)).values_list('lecture_code', flat=True)
        eids = EnrollmentCourse.objects.filter(lecture_code__in=lec_codes).values_list('enrollment_id', flat=True).distinct()
        qs = qs.filter(id__in=eids)
    elif sch_code_desc:
        subcodes = list(CodeValue.objects.filter(
            group__grpcode='LOCD', del_chk='N', code_desc=sch_code_desc
        ).values_list('subcode', flat=True))
        lec_codes = Lecture.objects.filter(local_code__in=subcodes).values_list('lecture_code', flat=True)
        eids = EnrollmentCourse.objects.filter(lecture_code__in=lec_codes).values_list('enrollment_id', flat=True).distinct()
        qs = qs.filter(id__in=eids)
    if sch_apply_gubun:
        qs = qs.filter(apply_gubun=sch_apply_gubun)
    if sch_lecture_stats:
        qs = qs.filter(lecture_stats=sch_lecture_stats)
    if sch_pay_method:
        qs = qs.filter(pay_method=sch_pay_method)
    if sch_pay_stats:
        qs = qs.filter(pay_stats=sch_pay_stats)
    if sch_sdate:
        qs = qs.filter(insert_dt__date__gte=sch_sdate)
    if sch_edate:
        qs = qs.filter(insert_dt__date__lte=sch_edate)
    if sch_sword:
        if sch_skey == 'member_id':
            qs = qs.filter(member__username__icontains=sch_sword)
        elif sch_skey == 'member_name':
            qs = qs.filter(member__name__icontains=sch_sword)
        elif sch_skey == 'child_id':
            qs = qs.filter(child__child_id__icontains=sch_sword)
        else:
            qs = qs.filter(child__name__icontains=sch_sword)

    qs = qs.order_by('-insert_dt')[:10000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '입단신청내역'

    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['No', '회원ID', '회원명', '자녀ID', '자녀명', '신청구분',
               '수강상태', '결제방법', '결제상태', '결제금액', '시작월', '종료월', '등록일']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    apply_map = dict(Enrollment.apply_gubun.field.choices)
    lstat_map = dict(Enrollment.lecture_stats.field.choices)
    pmethod_map = dict(Enrollment.pay_method.field.choices)
    pstat_map = dict(Enrollment.pay_stats.field.choices)

    for idx, e in enumerate(qs, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=e.member_id).border = thin_border
        ws.cell(row=row, column=3, value=e.member.name if e.member else '').border = thin_border
        ws.cell(row=row, column=4, value=e.child_id).border = thin_border
        ws.cell(row=row, column=5, value=e.child.name if e.child else '').border = thin_border
        ws.cell(row=row, column=6, value=apply_map.get(e.apply_gubun, e.apply_gubun)).border = thin_border
        ws.cell(row=row, column=7, value=lstat_map.get(e.lecture_stats, e.lecture_stats)).border = thin_border
        ws.cell(row=row, column=8, value=pmethod_map.get(e.pay_method, e.pay_method)).border = thin_border
        ws.cell(row=row, column=9, value=pstat_map.get(e.pay_stats, e.pay_stats)).border = thin_border
        ws.cell(row=row, column=10, value=e.pay_price).border = thin_border
        ws.cell(row=row, column=11, value=e.start_dt).border = thin_border
        ws.cell(row=row, column=12, value=e.end_dt).border = thin_border
        dt_str = e.insert_dt.strftime('%Y-%m-%d') if e.insert_dt else ''
        ws.cell(row=row, column=13, value=dt_str).border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="enrollment_list.xlsx"'
    wb.save(response)
    return response


# ============================================================
# 수강생관리 > 변경이력조회
# ============================================================

@office_login_required
@office_permission_required('H')
def chghis_list(request):
    """변경이력 목록"""
    sch_reg_month = request.GET.get('sch_reg_month', '')
    sch_child_name = request.GET.get('sch_child_name', '')
    page = request.GET.get('page', '1')

    qs = ChangeHistory.objects.all()

    if sch_reg_month:
        try:
            ym_year = int(sch_reg_month[:4])
            ym_month = int(sch_reg_month[4:6])
            qs = qs.filter(reg_dt__year=ym_year, reg_dt__month=ym_month)
        except (ValueError, IndexError):
            pass
    if sch_child_name:
        # child_id로 MemberChild 검색
        child_ids = MemberChild.objects.filter(
            name__icontains=sch_child_name
        ).values_list('child_id', flat=True)
        qs = qs.filter(child_id__in=child_ids)

    qs = qs.order_by('-reg_dt')

    paginator = Paginator(qs, 20)
    histories = paginator.get_page(page)

    # 이력에 연결된 멤버/자녀 이름 매핑
    member_ids = set(h.member_id for h in histories)
    child_ids = set(h.child_id for h in histories)
    member_names = {m.username: m.name for m in Member.objects.filter(username__in=member_ids)}
    child_names = {c.child_id: c.name for c in MemberChild.objects.filter(child_id__in=child_ids)}

    # 변경 전/후 비교 데이터 조회
    no_seqs = set(h.no_seq for h in histories if h.no_seq)
    src_seqs = set(h.src_seq for h in histories if h.src_seq)
    # 현재 enrollment (변경 후 = no_seq → Enrollment.id)
    cur_map = {e.id: e for e in Enrollment.objects.filter(id__in=no_seqs)}
    # 변경 전 = src_seq → EnrollmentSrc 스냅샷 (Enrollment 아님! src_seq는 master_src PK)
    src_map = {e.src_seq: e for e in EnrollmentSrc.objects.filter(src_seq__in=src_seqs)}

    # 각 history 객체에 이름/비교 속성 부여
    for h in histories:
        h.member_name = member_names.get(h.member_id, '')
        h.child_name = child_names.get(h.child_id, '')
        cur = cur_map.get(h.no_seq)
        src = src_map.get(h.src_seq)
        h.cur_pay_price = cur.pay_price if cur else 0
        h.cur_lec_period = cur.lec_period if cur else 0
        h.cur_start_dt = cur.start_dt if cur else ''
        h.cur_end_dt = cur.end_dt if cur else ''
        h.src_pay_price = src.pay_price if src else 0
        h.src_lec_period = src.lec_period if src else 0
        h.src_start_dt = src.start_dt if src else ''
        h.src_end_dt = src.end_dt if src else ''
        h.price_diff = h.cur_pay_price - h.src_pay_price

    return render(request, 'ba_office/lfstudent/chghis_list.html', {
        'histories': histories,
        'total_count': paginator.count,
        'ym_choices': _build_ym_choices(),
        'sch_reg_month': sch_reg_month,
        'sch_child_name': sch_child_name,
    })


@office_login_required
@office_permission_required('H')
def chghis_detail(request, pk):
    """변경이력 상세"""
    history = get_object_or_404(ChangeHistory, pk=pk)

    # 현재 Enrollment 정보 (변경 후)
    enrollment = None
    if history.no_seq:
        enrollment = Enrollment.objects.filter(id=history.no_seq).select_related('member', 'child').first()

    # 변경 전 = src_seq → EnrollmentSrc 스냅샷 (Enrollment 아님)
    src_enrollment = None
    if history.src_seq:
        src_enrollment = EnrollmentSrc.objects.filter(src_seq=history.src_seq).first()

    # 수강과정/청구 (현재=변경 후). ASP 동일: bill_code='1001', course_ym 오름차순
    course_entries = []
    lectures_map = {}
    bills = []
    if enrollment:
        course_entries = EnrollmentCourse.objects.filter(
            enrollment=enrollment, bill_code='1001'
        ).order_by('course_ym', 'lecture_code')
        lec_codes = set(c.lecture_code for c in course_entries)
        if lec_codes:
            for l in Lecture.objects.filter(lecture_code__in=lec_codes).select_related('stadium'):
                lectures_map[l.lecture_code] = l
        bills = EnrollmentBill.objects.filter(enrollment=enrollment).order_by('bill_code')

    # 수강과정/청구 (변경 전 = 스냅샷, src_seq 기준)
    src_course_entries = []
    src_bills = []
    if src_enrollment:
        src_course_entries = EnrollmentCourseSrc.objects.filter(
            src_seq=history.src_seq, bill_code='1001'
        ).order_by('course_ym', 'lecture_code')
        src_lec_codes = set(c.lecture_code for c in src_course_entries)
        if src_lec_codes:
            for l in Lecture.objects.filter(lecture_code__in=src_lec_codes).select_related('stadium'):
                if l.lecture_code not in lectures_map:
                    lectures_map[l.lecture_code] = l
        src_bills = EnrollmentBillSrc.objects.filter(src_seq=history.src_seq).order_by('bill_code')

    # 멤버/자녀 이름
    member_name = ''
    child_name = ''
    try:
        member_name = Member.objects.get(username=history.member_id).name
    except Member.DoesNotExist:
        pass
    try:
        child_name = MemberChild.objects.get(child_id=history.child_id).name
    except MemberChild.DoesNotExist:
        pass

    return render(request, 'ba_office/lfstudent/chghis_detail.html', {
        'history': history,
        'enrollment': enrollment,
        'src_enrollment': src_enrollment,
        'course_entries': course_entries,
        'src_course_entries': src_course_entries,
        'lectures_map': lectures_map,
        'bills': bills,
        'src_bills': src_bills,
        'member_name': member_name,
        'child_name': child_name,
    })


# ============================================================
# 수강생관리 > 출결관리
# ============================================================

@office_login_required
@office_permission_required('H')
def attendance_view(request):
    """출결관리 - 검색/조회"""
    # [UX변경] 원본: sch_code_desc = request.GET.get('sch_code_desc', '') (필드 제거)
    sch_sta_code = request.GET.get('sch_sta_code', '')
    sch_lecture_code = request.GET.get('sch_lecture_code', '')
    # ASP 원본 파라미터명으로 변경: sch_mode→sch_process_gbn, sch_date→sch_lecture_month, sch_sdate/sch_edate→sch_lecture_month/sch_lecture_end_month
    sch_process_gbn = request.GET.get('sch_process_gbn', '1')  # 1=출석체크, 2=조회용
    sch_lecture_month = request.GET.get('sch_lecture_month', '')   # 단일날짜(mode1) or 시작일(mode2)
    sch_lecture_end_month = request.GET.get('sch_lecture_end_month', '')  # 종료일(mode2만)
    att_gbn = request.GET.get('att_gbn', '')
    suc_gbn = request.GET.get('suc_gbn', '')

    students = []
    records = []
    searched = False

    # [UX변경] 원본: code_descs = CodeValue.objects.filter(group__grpcode='LOCD'...) (필드 목록 제거)
    # [UX변경] 구장 전체 직접 표시. 원본: stadiums=[], if sch_code_desc: filter by local_code
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    courses = []
    if sch_sta_code:
        courses = Lecture.objects.filter(stadium__sta_code=int(sch_sta_code), use_gbn='Y').order_by('lecture_title')

    if sch_lecture_code and sch_lecture_month:
        searched = True
        lecture_code = int(sch_lecture_code)

        if sch_process_gbn == '1':
            # 모드 1: 출석체크 - 해당 강좌의 수강생 목록
            sch_date = sch_lecture_month
            try:
                att_date_parts = sch_date.split('-')
                att_year = int(att_date_parts[0])
                att_month = int(att_date_parts[1])
                course_ym_date = date(att_year, att_month, 1)
            except (ValueError, IndexError):
                course_ym_date = date(datetime.now().year, datetime.now().month, 1)

            # 수강 중인 학생들
            enrolled = EnrollmentCourse.objects.filter(
                bill_code='1001',
                lecture_code=lecture_code,
                course_ym=course_ym_date,
                enrollment__del_chk='N',
            ).filter(
                Q(enrollment__apply_gubun__in=['NEW', 'RENEW'], enrollment__lecture_stats='LY') |
                Q(enrollment__apply_gubun='AGAIN', enrollment__lecture_stats__in=['LY', 'LP'])
            ).select_related('enrollment__child')

            # 기존 출결 데이터
            existing_att = {}
            for a in Attendance.objects.filter(lecture_code=lecture_code, attendance_dt=sch_date):
                existing_att[a.child_id] = a

            seen_children = set()
            for ec in enrolled:
                cid = ec.enrollment.child_id
                if cid in seen_children:
                    continue
                seen_children.add(cid)
                att = existing_att.get(cid)
                students.append({
                    'child_id': cid,
                    'child_name': ec.enrollment.child.name if ec.enrollment.child else '',
                    'no_seq': ec.enrollment_id,
                    'attendance_gbn': att.attendance_gbn if att else '',
                    'attendance_desc': att.attendance_desc if att else '',
                    'mata': att.mata if att else '',
                    'app_month': att.app_month if att else '',
                    'complete_yn': att.complete_yn if att else 'N',
                    'uid': att.id if att else 0,
                })

        elif sch_process_gbn == '2' and sch_lecture_end_month:
            # 모드 2: 조회 - 기간별 출결 데이터
            records = Attendance.objects.filter(
                lecture_code=lecture_code,
                attendance_dt__gte=sch_lecture_month,
                attendance_dt__lte=sch_lecture_end_month,
            )
            if att_gbn:
                records = records.filter(attendance_gbn=att_gbn)
            if suc_gbn:
                records = records.filter(complete_yn=suc_gbn)
            records = records.order_by('attendance_dt', 'child_id')

            # 자녀 이름 매핑
            child_ids = set(r.child_id for r in records)
            child_map = {c.child_id: c.name for c in MemberChild.objects.filter(child_id__in=child_ids)}
            for r in records:
                r.child_name = child_map.get(r.child_id, '')

    return render(request, 'ba_office/lfstudent/attendance.html', {
        'stadiums': stadiums,
        'courses': courses,
        'students': students,
        'records': records,
        'searched': searched,
        'sch_sta_code': sch_sta_code,
        'sch_lecture_code': sch_lecture_code,
        'sch_process_gbn': sch_process_gbn,
        'sch_lecture_month': sch_lecture_month,
        'sch_lecture_end_month': sch_lecture_end_month,
        'att_gbn': att_gbn,
        'suc_gbn': suc_gbn,
    })


@office_login_required
@office_permission_required('H')
def attendance_proc(request):
    """출결관리 처리 (저장)"""
    if request.method != 'POST':
        return redirect('office_attendance')

    lecture_code = int(request.POST.get('lecture_code', '0'))
    att_date = request.POST.get('att_date', '')
    att_mode = request.POST.get('att_mode', '2')  # 1=일괄, 2=개별
    child_ids = request.POST.getlist('child_ids')

    office_user = request.session.get('office_user', {})
    insert_id = office_user.get('office_id', '')

    # 강좌 정보
    lecture = Lecture.objects.filter(lecture_code=lecture_code).first()
    local_code = lecture.local_code if lecture else 0
    sta_code = lecture.stadium.sta_code if lecture and lecture.stadium else 0

    with transaction.atomic():
        if att_mode == '1':
            # 일괄처리: 모든 학생에게 동일 출결
            batch_gbn = request.POST.get('batch_gbn', 'Y')
            batch_desc = request.POST.get('batch_desc', '')

            # 기존 삭제
            Attendance.objects.filter(
                lecture_code=lecture_code, attendance_dt=att_date
            ).delete()

            # 일괄 입력
            for cid in child_ids:
                Attendance.objects.create(
                    local_code=local_code,
                    sta_code=sta_code,
                    lecture_code=lecture_code,
                    child_id=cid,
                    attendance_dt=att_date,
                    attendance_gbn=batch_gbn,
                    attendance_desc=batch_desc,
                    insert_id=insert_id,
                    insert_dt=timezone.now(),
                )
        else:
            # 개별처리
            for cid in child_ids:
                gbn = request.POST.get(f'gbn_{cid}', '')
                if not gbn:
                    continue
                desc = request.POST.get(f'desc_{cid}', '')
                mata = request.POST.get(f'mata_{cid}', '')
                app_month = request.POST.get(f'app_month_{cid}', '')
                complete_yn = request.POST.get(f'complete_yn_{cid}', 'N')

                Attendance.objects.filter(
                    child_id=cid, lecture_code=lecture_code, attendance_dt=att_date
                ).delete()

                Attendance.objects.create(
                    local_code=local_code,
                    sta_code=sta_code,
                    lecture_code=lecture_code,
                    child_id=cid,
                    attendance_dt=att_date,
                    attendance_gbn=gbn,
                    attendance_desc=desc,
                    mata=mata,
                    app_month=app_month,
                    complete_yn=complete_yn,
                    insert_id=insert_id,
                    insert_dt=timezone.now(),
                )

    # 원래 검색 조건으로 리다이렉트 (ASP 파라미터명 사용)
    # [UX변경] 원본: sch_code_desc = request.POST.get('sch_code_desc', '') (필드 제거)
    sch_sta_code = request.POST.get('sch_sta_code', '')
    return redirect(f'/ba_office/lfstudent/attendance/?sch_process_gbn=1&sch_lecture_code={lecture_code}&sch_lecture_month={att_date}&sch_sta_code={sch_sta_code}')


# ============================================================
# 수강생관리 > 대기정보관리
# ============================================================

@office_login_required
@office_permission_required('H')
def wait_list(request):
    """대기자정보 목록"""
    # 개별 삭제 처리 (GET 파라미터)
    del_id = request.GET.get('del_id', '')
    if del_id:
        WaitStudent.objects.filter(id=int(del_id)).update(del_chk='Y')
        return redirect('office_wait_list')

    sch_locd_code = request.GET.get('sch_locd_code', '')
    sch_sta_code = request.GET.get('sch_sta_code', '')
    trans_gbn = request.GET.get('trans_gbn', '')

    qs = WaitStudent.objects.filter(del_chk='N')
    if sch_locd_code:
        qs = qs.filter(local_code=int(sch_locd_code))
    if sch_sta_code:
        qs = qs.filter(sta_code=int(sch_sta_code))
    if trans_gbn:
        qs = qs.filter(trans_gbn=trans_gbn)
    qs = qs.order_by('sta_code', 'lecture_code', 'insert_dt')

    wait_data = list(qs)
    sta_codes = set(w.sta_code for w in wait_data)
    lec_codes = set(w.lecture_code for w in wait_data)
    sta_map = {}
    if sta_codes:
        for s in Stadium.objects.filter(sta_code__in=sta_codes):
            sta_map[s.sta_code] = s.sta_nickname or s.sta_name
    lec_map = {}
    if lec_codes:
        for lec in Lecture.objects.filter(lecture_code__in=lec_codes):
            lec_map[lec.lecture_code] = lec.lecture_title

    member_ids = set(w.member_id for w in wait_data)
    phone_map = dict(
        Member.objects.filter(username__in=member_ids).values_list('username', 'phone')
    ) if member_ids else {}

    for w in wait_data:
        w.sta_name = sta_map.get(w.sta_code, '')
        w.lec_title = lec_map.get(w.lecture_code, '')
        w.mhtel = phone_map.get(w.member_id, '')
        w.trans_gbn_text = '처리완료' if w.trans_gbn == 'Y' else '미처리'

    # [UX변경] 권역→구장 cascade 제거 → 구장 직접선택(전체)
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    return render(request, 'ba_office/lfstudent/wait_list.html', {
        'wait_list': wait_data,
        'stadiums': stadiums,
        'sch_sta_code': sch_sta_code,
        'trans_gbn': trans_gbn,
    })


@office_login_required
@office_permission_required('H')
def wait_write(request):
    """대기자 등록"""
    if request.method == 'POST':
        member_id = request.POST.get('member_id', '').strip()
        member_name = request.POST.get('member_name', '').strip()
        child_id = request.POST.get('child_id', '').strip()
        child_name = request.POST.get('child_name', '').strip()
        sta_code_str = request.POST.get('sta_code', '').strip()
        lecture_code_str = request.POST.get('lecture_code', '').strip()
        wait_seq_str = request.POST.get('wait_seq', '').strip()
        bigo = request.POST.get('bigo', '').strip()
        insert_id = request.session.get('office_user', {}).get('office_id', '')

        # 서버측 필수값 검증 ([UX변경] 권역 입력 제거)
        if not member_id or not child_id or not sta_code_str or not lecture_code_str or not wait_seq_str:
            return render(request, 'ba_office/lfstudent/wait_write.html', {
                'children': [],
                'stadiums': Stadium.objects.filter(use_gbn='Y').order_by('sta_name'),
                'sch_field': 'member_id', 'sch_value': '',
                'error_msg': '필수정보가 부족합니다.',
            })

        sta_code = int(sta_code_str)
        lecture_code = int(lecture_code_str)
        wait_seq = int(wait_seq_str)
        # [UX변경] local_code는 선택 구장에서 파생
        _sta = Stadium.objects.filter(sta_code=sta_code).first()
        local_code = _sta.local_code if _sta and _sta.local_code else 0

        # 중복체크: 미처리 대기자 중 동일 child_id
        if WaitStudent.objects.filter(child_id=child_id, del_chk='N', trans_gbn='N').exists():
            return render(request, 'ba_office/lfstudent/wait_write.html', {
                'children': [],
                'stadiums': Stadium.objects.filter(use_gbn='Y').order_by('sta_name'),
                'sch_field': 'member_id', 'sch_value': '',
                'error_msg': f'{child_id} 은(는) 대기자 명단에 존재합니다. 등록할 수 없습니다.',
            })

        phone = ''
        try:
            member = Member.objects.get(username=member_id)
            phone = member.phone or ''
        except Member.DoesNotExist:
            pass

        WaitStudent.objects.create(
            local_code=local_code,
            sta_code=sta_code,
            lecture_code=lecture_code,
            member_id=member_id,
            member_name=member_name,
            child_id=child_id,
            child_name=child_name,
            wait_seq=wait_seq,
            bigo=bigo,
            phone=phone,
            insert_id=insert_id,
        )
        return redirect('office_wait_list')

    # GET
    sch_field = request.GET.get('sch_field', 'member_id')
    sch_value = request.GET.get('sch_value', '')
    children = []
    if sch_value:
        qs = MemberChild.objects.filter(status='N')
        if sch_field == 'member_id':
            qs = qs.filter(parent__username__icontains=sch_value)
        elif sch_field == 'child_id':
            qs = qs.filter(child_id__icontains=sch_value)
        elif sch_field == 'child_name':
            qs = qs.filter(name__icontains=sch_value)
        children = list(qs.select_related('parent').order_by('name'))

    # [UX변경] 권역→구장 cascade 제거 → 구장 직접선택
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    return render(request, 'ba_office/lfstudent/wait_write.html', {
        'children': children,
        'stadiums': stadiums,
        'sch_field': sch_field,
        'sch_value': sch_value,
    })


@office_login_required
@office_permission_required('H')
def wait_modify(request, pk):
    """대기자 수정"""
    wait = get_object_or_404(WaitStudent, pk=pk)

    if request.method == 'POST':
        wait.sta_code = int(request.POST.get('sta_code', '0') or '0')
        # [UX변경] local_code는 선택 구장에서 파생
        _sta = Stadium.objects.filter(sta_code=wait.sta_code).first()
        wait.local_code = _sta.local_code if _sta and _sta.local_code else 0
        wait.lecture_code = int(request.POST.get('lecture_code', '0') or '0')
        wait.wait_seq = int(request.POST.get('wait_seq', '0') or '0')
        wait.trans_gbn = request.POST.get('trans_gbn', 'N')
        wait.bigo = request.POST.get('bigo', '')
        wait.save()
        return redirect('office_wait_list')

    # [UX변경] 권역→구장 cascade 제거 → 구장 직접선택(전체)
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    courses = Lecture.objects.filter(
        stadium__sta_code=wait.sta_code, use_gbn='Y'
    ).order_by('lecture_day', 'lecture_time') if wait.sta_code else []

    return render(request, 'ba_office/lfstudent/wait_modify.html', {
        'wait': wait,
        'stadiums': stadiums,
        'courses': courses,
    })


@office_login_required
@office_permission_required('H')
def wait_delete_proc(request):
    """대기자 삭제 (일괄)"""
    if request.method == 'POST':
        wait_ids = request.POST.getlist('wait_sel_chk')
        if wait_ids:
            WaitStudent.objects.filter(id__in=wait_ids).update(del_chk='Y')
    return redirect(request.META.get('HTTP_REFERER', '/ba_office/lfstudent/wait/'))


# ============================================================
# 수강생관리 > 일괄처리
# ============================================================

@office_login_required
@office_permission_required('H')
def batch_confirm_proc(request):
    """수강확정 일괄처리 (LP → LY)"""
    if request.method == 'POST':
        child_ids = request.POST.getlist('sel_chk')
        sch_ym = request.POST.get('sch_ym', '')

        if child_ids and sch_ym:
            try:
                course_ym_date = date(int(sch_ym[:4]), int(sch_ym[4:6]), 1)
            except (ValueError, IndexError):
                return redirect(request.META.get('HTTP_REFERER', '/ba_office/lfstudent/student/'))

            for child_id in child_ids:
                enrollments = Enrollment.objects.filter(
                    child_id=child_id, lecture_stats='LP', del_chk='N',
                )
                for enrollment in enrollments:
                    has_course = EnrollmentCourse.objects.filter(
                        enrollment=enrollment, course_ym=course_ym_date, bill_code='1001',
                    ).exists()
                    if has_course:
                        enrollment.lecture_stats = 'LY'
                        enrollment.save()
                        EnrollmentCourse.objects.filter(
                            enrollment=enrollment
                        ).update(course_stats='LY')
                        MemberChild.objects.filter(
                            child_id=child_id
                        ).update(course_state='ING')

    return redirect(request.META.get('HTTP_REFERER', '/ba_office/lfstudent/student/'))


@office_login_required
@office_permission_required('H')
def batch_confirm_pay_proc(request):
    """수강확정+결제완료 일괄처리 (LP→LY, PP→PY)"""
    if request.method == 'POST':
        child_ids = request.POST.getlist('sel_chk')
        sch_ym = request.POST.get('sch_ym', '')

        if child_ids and sch_ym:
            try:
                course_ym_date = date(int(sch_ym[:4]), int(sch_ym[4:6]), 1)
            except (ValueError, IndexError):
                return redirect(request.META.get('HTTP_REFERER', '/ba_office/lfstudent/student/'))

            for child_id in child_ids:
                enrollments = Enrollment.objects.filter(
                    child_id=child_id, lecture_stats='LP', del_chk='N',
                )
                for enrollment in enrollments:
                    has_course = EnrollmentCourse.objects.filter(
                        enrollment=enrollment, course_ym=course_ym_date, bill_code='1001',
                    ).exists()
                    if has_course:
                        enrollment.lecture_stats = 'LY'
                        enrollment.pay_stats = 'PY'
                        enrollment.pay_dt = timezone.now()
                        enrollment.save()
                        EnrollmentCourse.objects.filter(
                            enrollment=enrollment
                        ).update(course_stats='LY')
                        EnrollmentBill.objects.filter(
                            enrollment=enrollment
                        ).update(pay_stats='PY')
                        MemberChild.objects.filter(
                            child_id=child_id
                        ).update(course_state='ING')

    return redirect(request.META.get('HTTP_REFERER', '/ba_office/lfstudent/student/'))


@office_login_required
@office_permission_required('H')
@transaction.atomic
def batch_next_month_proc(request):
    """익월수강예정 일괄생성"""
    if request.method != 'POST':
        return redirect('/ba_office/lfstudent/student/')

    child_ids = request.POST.getlist('sel_chk')
    sch_ym = request.POST.get('sch_ym', '')
    insert_id = request.session.get('office_user', {}).get('office_id', '')
    # DN모드: 개월선택 (GET 파라미터)
    sel_lec_period = int(request.GET.get('sel_lec_period', '0') or '0')

    if not child_ids or not sch_ym:
        return redirect('/ba_office/lfstudent/student/')

    try:
        cur_year = int(sch_ym[:4])
        cur_month = int(sch_ym[4:6])
    except (ValueError, IndexError):
        return redirect('/ba_office/lfstudent/student/')

    next_month = cur_month + 1
    next_year = cur_year
    if next_month > 12:
        next_month = 1
        next_year += 1
    next_ym = f'{next_year}{next_month:02d}'
    next_ym_date = date(next_year, next_month, 1)

    for child_id in child_ids:
        # 현재월 수강정보 조회
        current_enrollments = Enrollment.objects.filter(
            child_id=child_id,
            end_dt=sch_ym,
            lecture_stats__in=['LY', 'LP', 'PN'],
            del_chk='N',
        ).order_by('-id')

        if not current_enrollments.exists():
            continue

        cur_enrollment = current_enrollments.first()

        # 익월 중복 체크
        if Enrollment.objects.filter(
            child_id=child_id, start_dt=next_ym,
            lecture_stats__in=['LY', 'LP', 'PN'], del_chk='N',
        ).exists():
            continue

        # 현재 강좌 코드
        cur_courses = EnrollmentCourse.objects.filter(
            enrollment=cur_enrollment, bill_code='1001',
        )
        lecture_codes = list(set(c.lecture_code for c in cur_courses))
        if not lecture_codes:
            continue

        lec_cycle = cur_enrollment.lec_cycle or 1
        # DN모드(개월선택)이면 사용자 선택값 사용, 아니면 현재 enrollment에서 복사
        lec_period = sel_lec_period if sel_lec_period > 0 else (cur_enrollment.lec_period or 1)

        # 자동이체 여부 (ASP 원본과 동일)
        auto_pay_methods = ['MUCU', 'MCDO', 'SDM', 'TA', 'CAU', 'YDP']
        is_auto = cur_enrollment.pay_method in auto_pay_methods

        # 시간표 유효성 검증: 모든 강좌에 대해 lec_period 개월분 스케줄 존재 확인
        schedule_valid = True
        for lec_code in lecture_codes:
            for offset in range(lec_period):
                m = next_month + offset
                y = next_year
                while m > 12:
                    m -= 12
                    y += 1
                if not LectureSelDay.objects.filter(
                    lecture_code=lec_code, syear=y, smonth=m,
                ).exists():
                    schedule_valid = False
                    break
            if not schedule_valid:
                break
        if not schedule_valid:
            continue

        # 종료년월 계산
        end_m = next_month + lec_period - 1
        end_y = next_year
        while end_m > 12:
            end_m -= 12
            end_y += 1
        end_ym = f'{end_y}{end_m:02d}'

        new_enrollment = Enrollment.objects.create(
            member_id=cur_enrollment.member_id,
            child_id=child_id,
            pay_stats='PY' if is_auto else 'PP',
            pay_method=cur_enrollment.pay_method,
            pay_price=0,
            lecture_stats='LY' if is_auto else 'LP',
            lec_cycle=lec_cycle,
            lec_period=lec_period,
            start_dt=next_ym,
            end_dt=end_ym,
            apply_gubun='AGAIN',
            source_gubun='02',
            shuttle_yn=cur_enrollment.shuttle_yn,
            del_chk='N',
            insert_id=insert_id,
        )

        # 강좌별 Course 생성 (lec_period 개월분)
        total_lecture_price = 0
        for lec_code in lecture_codes:
            try:
                lec = Lecture.objects.get(lecture_code=lec_code)
            except Lecture.DoesNotExist:
                continue

            for offset in range(lec_period):
                m = next_month + offset
                y = next_year
                while m > 12:
                    m -= 12
                    y += 1

                day_count = LectureSelDay.objects.filter(
                    lecture_code=lec_code, syear=y, smonth=m,
                ).count()

                course_amt = lec.lec_price * day_count
                total_lecture_price += course_amt

                EnrollmentCourse.objects.create(
                    enrollment=new_enrollment,
                    bill_code='1001',
                    course_ym=date(y, m, 1),
                    course_ym_amt=course_amt,
                    lecture_code=lec_code,
                    start_ymd=date(y, m, 1),
                    course_stats='LY' if is_auto else 'LP',
                )

        # 다회할인 계산 (dc_X × lec_period, ASP 원본 동일)
        multi_discount = 0
        if lec_cycle >= 2:
            for lec_code in lecture_codes:
                try:
                    lec = Lecture.objects.get(lecture_code=lec_code)
                    if lec_cycle == 2:
                        multi_discount += (lec.dc_2 or 0) * lec_period
                    elif lec_cycle == 3:
                        multi_discount += (lec.dc_3 or 0) * lec_period
                    elif lec_cycle >= 4:
                        multi_discount += (lec.dc_4 or 0) * lec_period
                except Lecture.DoesNotExist:
                    pass

        # 프로모션 할인 (bill_code 1003)
        promo_discount = 0
        active_promos = Promotion.objects.filter(
            start_date__lte=timezone.now(), end_date__gte=timezone.now(), is_use='T',
        )
        for promo in active_promos:
            if promo.use_mode == 2:  # 수강료차감
                is_member = PromotionMember.objects.filter(
                    coupon_uid=promo.uid, child_id=child_id,
                ).exists()
                if is_member:
                    promo_discount += promo.discount or 0

        bill_stats = 'PY' if is_auto else 'PP'
        EnrollmentBill.objects.create(
            enrollment=new_enrollment,
            bill_code='1001', bill_desc='수업료',
            bill_amt=total_lecture_price,
            pay_stats=bill_stats, insert_id=insert_id,
        )

        if multi_discount > 0:
            EnrollmentBill.objects.create(
                enrollment=new_enrollment,
                bill_code='1007', bill_desc=f'주{lec_cycle}회 할인',
                bill_amt=-multi_discount,
                pay_stats=bill_stats, insert_id=insert_id,
            )

        if promo_discount > 0:
            EnrollmentBill.objects.create(
                enrollment=new_enrollment,
                bill_code='1003', bill_desc='프로모션할인',
                bill_amt=-promo_discount,
                pay_stats=bill_stats, insert_id=insert_id,
            )

        pay_price = max(total_lecture_price - multi_discount - promo_discount, 0)
        new_enrollment.pay_price = pay_price
        new_enrollment.save()

    referer = request.META.get('HTTP_REFERER', '/ba_office/lfstudent/student/')
    return redirect(referer)


@office_login_required
@office_permission_required('H')
def batch_lms_proc(request):
    """결제안내 LMS 발송"""
    if request.method != 'POST':
        return redirect('/ba_office/lfstudent/student/')

    child_ids = request.POST.getlist('sel_chk')
    sch_ym = request.POST.get('sch_ym', '')
    insert_id = request.session.get('office_user', {}).get('office_id', '')

    if not child_ids or not sch_ym:
        return redirect('/ba_office/lfstudent/student/')

    try:
        course_ym_date = date(int(sch_ym[:4]), int(sch_ym[4:6]), 1)
    except (ValueError, IndexError):
        return redirect('/ba_office/lfstudent/student/')

    for child_id in child_ids:
        enrollments = Enrollment.objects.filter(
            child_id=child_id, pay_stats='PP', del_chk='N',
        )
        for enrollment in enrollments:
            has_course = EnrollmentCourse.objects.filter(
                enrollment=enrollment, course_ym=course_ym_date, bill_code='1001',
            ).exists()
            if not has_course:
                continue

            try:
                member = Member.objects.get(username=enrollment.member_id)
                phone = (member.phone or '').replace('-', '')
            except Member.DoesNotExist:
                continue
            if not phone:
                continue

            # 교육용품비 (bill_code 2xxx 합산)
            join_price = EnrollmentBill.objects.filter(
                enrollment=enrollment, bill_code__startswith='2',
            ).aggregate(total=Sum('bill_amt'))['total'] or 0

            # 수강료 (bill_code 1xxx 합산)
            lecture_price = EnrollmentBill.objects.filter(
                enrollment=enrollment, bill_code__startswith='1',
            ).aggregate(total=Sum('bill_amt'))['total'] or 0

            courses = EnrollmentCourse.objects.filter(
                enrollment=enrollment, bill_code='1001', course_ym=course_ym_date,
            )
            lec_codes = [c.lecture_code for c in courses]
            lectures = Lecture.objects.filter(
                lecture_code__in=lec_codes
            ).select_related('stadium', 'coach')

            # 첫 번째 강좌 기준 정보
            sta_name = ''
            lecture_title = ''
            coach_name = ''
            sta_phone = ''
            for lec in lectures:
                sta_name = lec.stadium.sta_name if lec.stadium else ''
                sta_phone = lec.stadium.sta_phone if lec.stadium else ''
                lecture_title = lec.lecture_title or ''
                coach_name = lec.coach.coach_name if lec.coach else ''
                break  # 첫 번째 강좌 정보만

            start_dt = enrollment.start_dt or ''
            end_dt = enrollment.end_dt or ''

            msg = (
                f'[AAFC]\n'
                f'안녕하세요. AAFC입니다.\n'
                f'수강료 금액 결제 요청 드립니다.\n\n'
                f'1. 홈페이지 결제 http://www.aafc.co.kr\n'
                f'[결제정보]\n'
                f'  ＊ 교육용품비 = {join_price:,}원\n'
                f'  ＊ 수강료 = {lecture_price:,}원\n'
                f'  ＊ 총 결제금액 = {enrollment.pay_price:,}원\n\n'
                f'[수업정보]\n'
                f'  ＊ 구장 = {sta_name}\n'
                f'  ＊ 클래스 = {lecture_title}\n'
                f'  ＊ 수강월 = {start_dt}~{end_dt}\n'
                f'  ＊ 담당 코치 = {coach_name}\n'
                f'  ＊ 야드 연락처 = {sta_phone}'
            )

            SMSLog.objects.create(
                service_type='3',
                recipient_num=phone,
                subject='수강료 결제 요청',
                content=msg,
                callback='18117909',
                date_client_req=timezone.now(),
                msg_status='1',
            )

    return redirect(request.META.get('HTTP_REFERER', '/ba_office/lfstudent/student/'))


# ============================================================
# 수강생관리 > AJAX (수강생등록/대기자용)
# ============================================================

@office_login_required
def ajax_child_search(request):
    """AJAX: 자녀 검색"""
    sch_field = request.GET.get('sch_field', 'member_id')
    sch_value = request.GET.get('sch_value', '')

    if not sch_value:
        return JsonResponse([], safe=False)

    qs = MemberChild.objects.filter(status='N')
    if sch_field == 'member_id':
        qs = qs.filter(parent__username__icontains=sch_value)
    elif sch_field == 'child_id':
        qs = qs.filter(child_id__icontains=sch_value)
    elif sch_field == 'child_name':
        qs = qs.filter(name__icontains=sch_value)

    qs = qs.select_related('parent').order_by('name')[:50]
    data = []
    for c in qs:
        data.append({
            'child_id': c.child_id,
            'name': c.name,
            'member_id': c.parent.username if c.parent else '',
            'member_name': c.parent.name if c.parent else '',
            'birth': c.birth or '',
            'sch_name': c.school or '',
            'sch_grade': c.grade or '',
            'sex': c.gender or '',
            'course_state': c.course_state or '',
        })
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_course_list(request):
    """AJAX: 구장+년월별 강좌목록"""
    sta_code = request.GET.get('sta_code', '')
    ym = request.GET.get('ym', '')

    if not sta_code:
        return JsonResponse([], safe=False)

    qs = Lecture.objects.filter(
        stadium__sta_code=int(sta_code), use_gbn='Y',
    ).select_related('stadium', 'coach').order_by('lecture_day', 'class_gbn', 'lecture_time')

    data = []
    for lec in qs:
        current_count = 0
        if ym:
            try:
                ym_date = date(int(ym[:4]), int(ym[4:6]), 1)
                current_count = EnrollmentCourse.objects.filter(
                    lecture_code=lec.lecture_code, course_ym=ym_date,
                    bill_code='1001', course_stats__in=['LY', 'LP'],
                ).values('enrollment__child_id').distinct().count()
            except (ValueError, IndexError):
                pass

        data.append({
            'lecture_code': lec.lecture_code,
            'lecture_title': lec.lecture_title,
            'lec_age': lec.lec_age or '',
            'class_gbn': lec.class_gbn or '',
            'lec_price': lec.lec_price,
            'lecture_day': lec.lecture_day or 0,
            'lecture_time': lec.lecture_time or '',
            'stu_cnt': lec.stu_cnt or 0,
            'current_count': current_count,
            'coach_name': lec.coach.coach_name if lec.coach else '',
            'dc_2': lec.dc_2 or 0,
            'dc_3': lec.dc_3 or 0,
            'dc_4': lec.dc_4 or 0,
        })
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_course_days(request):
    """AJAX: 강좌별 수업시작일"""
    lecture_code = request.GET.get('lecture_code', '')
    syear = request.GET.get('syear', '')
    smonth = request.GET.get('smonth', '')

    if not lecture_code or not syear or not smonth:
        return JsonResponse([], safe=False)

    days = LectureSelDay.objects.filter(
        lecture_code=int(lecture_code), syear=int(syear), smonth=int(smonth),
    ).order_by('sday')

    # LectureSelDay에 요일 필드 없음 → 날짜로 계산 (월=0 … 일=6)
    weekday_names = ['월', '화', '수', '목', '금', '토', '일']
    data = []
    for d in days:
        try:
            sweek = weekday_names[date(d.syear, d.smonth, d.sday).weekday()]
        except (ValueError, TypeError):
            sweek = ''
        data.append({'sday': d.sday, 'sweek': sweek})
    return JsonResponse(data, safe=False)


@office_login_required
def ajax_course_price(request):
    """AJAX: 선택 강좌 수강료 계산 (수업일수 기반, 원본 course_sel.asp 동일).
    수강료 = Σ(1회단가 × 그달 수업일수). 첫 달은 시작일 이후 잔여 세션만."""
    syear = int(request.GET.get('syear', '0') or '0')
    smonth = int(request.GET.get('smonth', '0') or '0')
    lec_period = int(request.GET.get('lec_period', '1') or '1')
    codes = [c for c in request.GET.get('lectures', '').split(',') if c]

    total = 0
    details = []
    for code in codes:
        try:
            lec = Lecture.objects.get(lecture_code=int(code))
        except (Lecture.DoesNotExist, ValueError):
            continue
        sday = int(request.GET.get(f'start_day_{code}', '1') or '1')
        price = 0
        for offset in range(lec_period):
            m = smonth + offset
            y = syear
            while m > 12:
                m -= 12
                y += 1
            if offset == 0:
                cnt = LectureSelDay.objects.filter(
                    lecture_code=lec.lecture_code, syear=y, smonth=m, sday__gte=sday
                ).count()
            else:
                cnt = LectureSelDay.objects.filter(
                    lecture_code=lec.lecture_code, syear=y, smonth=m
                ).count()
            price += lec.lec_price * cnt
        total += price
        details.append({'lecture_code': lec.lecture_code, 'price': price})
    return JsonResponse({'total_lec': total, 'details': details})


@office_login_required
def ajax_promotions(request):
    """AJAX: 프로모션 할인 조회"""
    child_id = request.GET.get('child_id', '')
    now = timezone.now()

    promotions = []
    if child_id:
        active_promos = Promotion.objects.filter(
            start_date__lte=now, end_date__gte=now, is_use='T',
        )
        for promo in active_promos:
            is_member = PromotionMember.objects.filter(
                coupon_uid=promo.uid, child_id=child_id,
            ).exists()
            if is_member:
                promotions.append({
                    'uid': promo.uid,
                    'title': promo.title,
                    'discount': promo.discount or 0,
                    'discount_unit': promo.discount_unit or '',
                    'use_mode': promo.use_mode or 0,
                })

    return JsonResponse(promotions, safe=False)


# ============================================================
# 수강생관리 > 수강생등록
# ============================================================

@office_login_required
@office_permission_required('H')
def student_add(request):
    """수강생등록 (관리자)"""
    if request.method == 'POST':
        return _student_add_proc(request)

    # [UX변경] 구장 전체 직접 표시. 원본: locd_list(권역) + AJAX cascade
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    setting = Setting.objects.first()
    join_price = setting.join_price if setting else 0

    default_ym = _get_default_ym()  # 'YYYYMM' (22일 이후면 익월)

    return render(request, 'ba_office/lfstudent/student_add.html', {
        'stadiums': stadiums,
        'join_price': join_price,
        'default_year': default_ym[:4],
        'default_month': default_ym[4:6],
    })


@transaction.atomic
def _student_add_proc(request):
    """수강생등록 처리 (POST)"""
    insert_id = request.session.get('office_user', {}).get('office_id', '')

    member_id = request.POST.get('member_id', '').strip()
    child_id = request.POST.get('child_id', '').strip()
    if not member_id or not child_id:
        return redirect('office_student_add')

    sta_code = int(request.POST.get('sta_code', '0') or '0')
    # [UX변경] 권역(locd_code) 직접입력 제거 → 선택 구장의 local_code에서 파생
    _sta = Stadium.objects.filter(sta_code=sta_code).first()
    local_code = _sta.local_code if _sta else 0
    syear = int(request.POST.get('syear', '0') or '0')
    smonth = int(request.POST.get('smonth', '0') or '0')
    lec_cycle = int(request.POST.get('lec_cycle', '1') or '1')
    lec_period = int(request.POST.get('lec_period', '1') or '1')

    lecture_codes = request.POST.getlist('sel_lecture')
    start_days = {}
    for code in lecture_codes:
        start_days[code] = int(request.POST.get(f'start_day_{code}', '1') or '1')

    pay_method = request.POST.get('pay_method', 'ACCT')
    pay_stats = request.POST.get('pay_stats', 'PP')
    lecture_stats = request.POST.get('lecture_stats', 'LP')
    join_price = int(request.POST.get('join_price', '0') or '0')
    pay_price = int(request.POST.get('pay_price', '0') or '0')
    shuttle_yn = request.POST.get('shuttle_yn', 'N')
    shuttle_amt = int(request.POST.get('shuttle_amt', '0') or '0')
    bigo = request.POST.get('bigo_content', '')
    recommend_id = request.POST.get('recommend_id', '').strip()

    eq_discount = int(request.POST.get('eq_discount', '0') or '0')
    tu_discount = int(request.POST.get('tu_discount', '0') or '0')
    py_discount = int(request.POST.get('py_discount', '0') or '0')
    multi_discount = int(request.POST.get('multi_discount', '0') or '0')
    recommend_discount = int(request.POST.get('recommend_discount', '0') or '0')

    # 종료년월 계산
    end_month = smonth + lec_period - 1
    end_year = syear
    while end_month > 12:
        end_month -= 12
        end_year += 1
    start_dt = f'{syear}{smonth:02d}'
    end_dt = f'{end_year}{end_month:02d}'

    # 신청구분
    try:
        child = MemberChild.objects.get(child_id=child_id)
        cs = child.course_state or 'CAN'
        if cs == 'ING':
            apply_gubun = 'AGAIN'
        elif cs in ('END', 'PAU'):
            apply_gubun = 'RENEW'
        else:
            apply_gubun = 'NEW'
    except MemberChild.DoesNotExist:
        apply_gubun = 'NEW'

    enrollment = Enrollment.objects.create(
        member_id=member_id,
        child_id=child_id,
        pay_stats=pay_stats,
        pay_method=pay_method,
        pay_price=pay_price,
        pay_dt=timezone.now() if pay_stats == 'PY' else None,
        lecture_stats=lecture_stats,
        lec_cycle=lec_cycle,
        lec_period=lec_period,
        start_dt=start_dt,
        end_dt=end_dt,
        apply_gubun=apply_gubun,
        source_gubun='02',
        shuttle_yn=shuttle_yn,
        recommend_id=recommend_id,
        bigo_content=bigo,
        del_chk='N',
        insert_id=insert_id,
    )

    total_lecture_price = 0
    first_month_price = 0  # 첫달 수업료 (1002 계산용)
    full_month_price = 0   # 한달 전체 수업료 (시작일 무관)
    bill_pay_stats = pay_stats

    for code in lecture_codes:
        try:
            lec = Lecture.objects.get(lecture_code=int(code))
        except Lecture.DoesNotExist:
            continue

        sday = start_days.get(code, 1)

        for offset in range(lec_period):
            m = smonth + offset
            y = syear
            while m > 12:
                m -= 12
                y += 1

            if offset == 0:
                count = LectureSelDay.objects.filter(
                    lecture_code=int(code), syear=y, smonth=m, sday__gte=sday
                ).count()
                full_count = LectureSelDay.objects.filter(
                    lecture_code=int(code), syear=y, smonth=m
                ).count()
                first_month_price += lec.lec_price * count
                full_month_price += lec.lec_price * full_count
            else:
                count = LectureSelDay.objects.filter(
                    lecture_code=int(code), syear=y, smonth=m
                ).count()

            course_amt = lec.lec_price * count
            total_lecture_price += course_amt

            start_ymd = date(y, m, sday if offset == 0 else 1)

            EnrollmentCourse.objects.create(
                enrollment=enrollment,
                bill_code='1001',
                course_ym=date(y, m, 1),
                course_ym_amt=course_amt,
                lecture_code=int(code),
                start_ymd=start_ymd,
                course_stats=lecture_stats,
            )

    # 수업료 (bill_code 1001)
    EnrollmentBill.objects.create(
        enrollment=enrollment,
        bill_code='1001', bill_desc='수업료',
        bill_amt=total_lecture_price,
        pay_stats=bill_pay_stats, insert_id=insert_id,
    )

    # 첫달수업료차감 (bill_code 1002) - 첫달 시작일로 인한 차액
    first_month_discount = full_month_price - first_month_price
    if first_month_discount != 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='1002', bill_desc='첫달수업료차감',
            bill_amt=-first_month_discount,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 교육용품비 (bill_code 2001)
    if join_price > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='2001', bill_desc='교육용품비',
            bill_amt=join_price,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 교육용품할인 (bill_code 2002)
    if eq_discount > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='2002', bill_desc='교육용품할인',
            bill_amt=-eq_discount,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 수강료할인 (bill_code 1003)
    if tu_discount > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='1003', bill_desc='수강료할인',
            bill_amt=-tu_discount,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 결제금액할인 (bill_code 1003)
    if py_discount > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='1003', bill_desc='결제금액할인',
            bill_amt=-py_discount,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 피추천인할인 (bill_code 1006)
    if recommend_discount > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='1006', bill_desc='피추천인할인',
            bill_amt=-recommend_discount,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 다회할인 (bill_code 1007)
    if multi_discount > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='1007', bill_desc=f'주{lec_cycle}회 할인',
            bill_amt=-multi_discount,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    # 차량이용료 (bill_code 1009)
    if shuttle_yn == 'Y' and shuttle_amt > 0:
        EnrollmentBill.objects.create(
            enrollment=enrollment,
            bill_code='1009', bill_desc='차량이용료',
            bill_amt=shuttle_amt,
            pay_stats=bill_pay_stats, insert_id=insert_id,
        )

    MemberChild.objects.filter(child_id=child_id).update(course_state='ING')

    WaitStudent.objects.filter(
        child_id=child_id, trans_gbn='N', del_chk='N'
    ).update(trans_gbn='Y')

    return redirect('office_student_list')


# ============================================================
# 과정관리 > 구장관리
# ============================================================

def _save_stadium_image(f, filename_prefix):
    """구장 이미지 업로드 헬퍼"""
    upload_dir = os.path.join(settings.MEDIA_ROOT, 'fcdata', 'stadium')
    os.makedirs(upload_dir, exist_ok=True)
    filename = filename_prefix + '_' + f.name
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, 'wb+') as dest:
        for chunk in f.chunks():
            dest.write(chunk)
    return filename


@office_login_required
@office_permission_required('L')
def stadium_list(request):
    """구장 목록"""
    use_gbn = request.GET.get('use_gbn', 'Y')
    stadiums = Stadium.objects.filter(use_gbn=use_gbn).order_by('sta_code')
    return render(request, 'ba_office/lfcourse/stadium_list.html', {
        'stadiums': stadiums,
        'use_gbn': use_gbn,
    })


@office_login_required
@office_permission_required('L')
def stadium_write(request):
    """구장 등록"""
    if request.method == 'POST':
        # sta_code 자동채번
        max_code = Stadium.objects.aggregate(m=Coalesce(Max('sta_code'), 0))['m']
        sta_code = max_code + 1

        stadium = Stadium.objects.create(
            sta_code=sta_code,
            sta_name=request.POST.get('sta_name', ''),
            sta_nickname=request.POST.get('sta_nickname', ''),
            sta_coach=request.POST.get('sta_coach', ''),
            use_gbn=request.POST.get('use_gbn', 'Y'),
            local_code=int(request.POST.get('local_code', '0') or '0'),
            sta_phone=request.POST.get('sta_phone', ''),
            three_lecyn=request.POST.get('three_lecyn', ''),
            sta_address=request.POST.get('sta_address', ''),
            sta_desc=request.POST.get('sta_desc', ''),
            kapa_tot=int(request.POST.get('kapa_tot', '0') or '0'),
            inve=request.POST.get('inve', ''),
            grou=request.POST.get('grou', ''),
            order_seq=int(request.POST.get('order_seq', '0') or '0'),
            insert_dt=timezone.now(),
        )

        # 이미지 저장
        prefix = str(sta_code)
        if request.FILES.get('sta_s_img'):
            stadium.sta_s_img = _save_stadium_image(request.FILES['sta_s_img'], prefix + '_s')
        if request.FILES.get('sta_l_img'):
            stadium.sta_l_img = _save_stadium_image(request.FILES['sta_l_img'], prefix + '_l')
        if request.FILES.get('sta_p_img'):
            stadium.sta_p_img = _save_stadium_image(request.FILES['sta_p_img'], prefix + '_p')
        if request.FILES.get('sta_m_img'):
            stadium.sta_m_img = _save_stadium_image(request.FILES['sta_m_img'], prefix + '_m')
        stadium.save()

        # StadiumCoach 일괄 생성
        coach_codes = request.POST.getlist('chk_coach')
        for cc in coach_codes:
            try:
                coach_obj = Coach.objects.get(coach_code=int(cc))
                StadiumCoach.objects.create(stadium=stadium, coach=coach_obj)
            except Coach.DoesNotExist:
                pass

        return redirect('office_stadium_list')

    # GET
    locd_list = CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N').order_by('code_order')
    inve_list = CodeValue.objects.filter(group__grpcode='INVE', del_chk='N').order_by('code_order')
    grou_list = CodeValue.objects.filter(group__grpcode='GROU', del_chk='N').order_by('code_order')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('order_seq', 'coach_name')
    return render(request, 'ba_office/lfcourse/stadium_write.html', {
        'locd_list': locd_list,
        'inve_list': inve_list,
        'grou_list': grou_list,
        'coaches': coaches,
    })


@office_login_required
@office_permission_required('L')
def stadium_modify(request, sta_code):
    """구장 수정 — 구장은 sta_code(구장코드)로 식별(list/goal/시간표 등 전 시스템 동일).
    주의: Django pk(id)와 sta_code 는 다른 값. pk로 조회하면 엉뚱한 구장이 열림."""
    stadium = get_object_or_404(Stadium, sta_code=sta_code)

    if request.method == 'POST':
        stadium.sta_name = request.POST.get('sta_name', '')
        stadium.sta_nickname = request.POST.get('sta_nickname', '')
        stadium.sta_coach = request.POST.get('sta_coach', '')
        stadium.use_gbn = request.POST.get('use_gbn', 'Y')
        stadium.local_code = int(request.POST.get('local_code', '0') or '0')
        stadium.sta_phone = request.POST.get('sta_phone', '')
        stadium.three_lecyn = request.POST.get('three_lecyn', '')
        stadium.sta_address = request.POST.get('sta_address', '')
        stadium.sta_desc = request.POST.get('sta_desc', '')
        stadium.kapa_tot = int(request.POST.get('kapa_tot', '0') or '0')
        stadium.inve = request.POST.get('inve', '')
        stadium.grou = request.POST.get('grou', '')
        stadium.order_seq = int(request.POST.get('order_seq', '0') or '0')

        # 이미지: 새 파일이 있으면 교체
        prefix = str(stadium.sta_code)
        if request.FILES.get('sta_s_img'):
            stadium.sta_s_img = _save_stadium_image(request.FILES['sta_s_img'], prefix + '_s')
        if request.FILES.get('sta_l_img'):
            stadium.sta_l_img = _save_stadium_image(request.FILES['sta_l_img'], prefix + '_l')
        if request.FILES.get('sta_p_img'):
            stadium.sta_p_img = _save_stadium_image(request.FILES['sta_p_img'], prefix + '_p')
        if request.FILES.get('sta_m_img'):
            stadium.sta_m_img = _save_stadium_image(request.FILES['sta_m_img'], prefix + '_m')
        stadium.save()

        # StadiumCoach 재생성 (기존 삭제 후 새로 생성)
        StadiumCoach.objects.filter(stadium=stadium).delete()
        coach_codes = request.POST.getlist('chk_coach')
        for cc in coach_codes:
            try:
                coach_obj = Coach.objects.get(coach_code=int(cc))
                StadiumCoach.objects.create(stadium=stadium, coach=coach_obj)
            except Coach.DoesNotExist:
                pass

        return redirect('office_stadium_list')

    # GET
    locd_list = CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N').order_by('code_order')
    inve_list = CodeValue.objects.filter(group__grpcode='INVE', del_chk='N').order_by('code_order')
    grou_list = CodeValue.objects.filter(group__grpcode='GROU', del_chk='N').order_by('code_order')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('order_seq', 'coach_name')

    # 기존 StadiumCoach 체크 상태
    sta_coaches = list(
        StadiumCoach.objects.filter(stadium=stadium).values_list('coach__coach_code', flat=True)
    )

    return render(request, 'ba_office/lfcourse/stadium_modify.html', {
        'stadium': stadium,
        'locd_list': locd_list,
        'inve_list': inve_list,
        'grou_list': grou_list,
        'coaches': coaches,
        'sta_coaches': sta_coaches,
    })


@office_login_required
@office_permission_required('L')
def stadium_goal(request, sta_code):
    """구장 목표 관리"""
    stadium = get_object_or_404(Stadium, sta_code=sta_code)

    if request.method == 'POST':
        sta_year = int(request.POST.get('sta_year', '0') or '0')
        sta_month = request.POST.get('sta_month', '')
        sta_goal = int(request.POST.get('sta_goal', '0') or '0')

        # 중복체크: 같은 구장 + 년도 + 월
        exists = StadiumGoal.objects.filter(
            stadium=stadium, sta_year=sta_year, sta_month=sta_month
        ).exists()
        if exists:
            goals = StadiumGoal.objects.filter(stadium=stadium).order_by('-sta_year', '-sta_month')
            return render(request, 'ba_office/lfcourse/stadium_goal.html', {
                'stadium': stadium,
                'goals': goals,
                'current_year': datetime.now().year,
                'error': '해당 년/월에 목표치가 존재합니다. 목표치를 삭제한 후 등록하여 주세요.',
            })

        StadiumGoal.objects.create(
            stadium=stadium,
            sta_year=sta_year,
            sta_month=sta_month,
            sta_goal=sta_goal,
        )
        return redirect('office_stadium_goal', sta_code=sta_code)

    # GET
    goals = StadiumGoal.objects.filter(stadium=stadium).order_by('-sta_year', '-sta_month')
    return render(request, 'ba_office/lfcourse/stadium_goal.html', {
        'stadium': stadium,
        'goals': goals,
        'current_year': datetime.now().year,
    })


@office_login_required
@office_permission_required('L')
def stadium_goal_del(request, sta_code):
    """구장 목표 삭제"""
    no_seq = request.GET.get('no_seq') or request.POST.get('no_seq')
    if no_seq:
        StadiumGoal.objects.filter(pk=no_seq).delete()
    return redirect('office_stadium_goal', sta_code=sta_code)


# ============================================================
# 과정관리 > 코치관리
# ============================================================

@office_login_required
@office_permission_required('L')
def coach_list(request):
    """코치 목록"""
    coaches = Coach.objects.filter(use_gbn='Y').order_by('order_seq', 'coach_name')

    # LEVL 코드그룹 조회하여 coach_level → code_name/code_desc 매핑
    levl_map = {}
    for cv in CodeValue.objects.filter(group__grpcode='LEVL', del_chk='N'):
        levl_map[str(cv.subcode)] = {
            'code_name': cv.code_name,
            'code_desc': cv.code_desc or '',
        }

    for c in coaches:
        info = levl_map.get(c.coach_level, {})
        c.level_name = info.get('code_name', '')
        c.level_desc = info.get('code_desc', '')

    return render(request, 'ba_office/lfcourse/coach_list.html', {
        'coaches': coaches,
    })


@office_login_required
@office_permission_required('L')
def coach_write(request):
    """코치 등록"""
    if request.method == 'POST':
        max_code = Coach.objects.aggregate(m=Coalesce(Max('coach_code'), 0))['m']
        coach_code = max_code + 1

        # 연락처: mhtel1)mhtel2-mhtel3 형식
        mhtel1 = request.POST.get('mhtel1', '')
        mhtel2 = request.POST.get('mhtel2', '')
        mhtel3 = request.POST.get('mhtel3', '')
        phone = f'{mhtel1}){mhtel2}-{mhtel3}' if mhtel1 else ''

        # 이미지 파일 업로드
        coach_s_img = ''
        if request.FILES.get('coach_s_img'):
            f = request.FILES['coach_s_img']
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'fcdata', 'coach')
            os.makedirs(upload_dir, exist_ok=True)
            filename = f'{coach_code}_{f.name}'
            filepath = os.path.join(upload_dir, filename)
            with open(filepath, 'wb+') as dest:
                for chunk in f.chunks():
                    dest.write(chunk)
            coach_s_img = filename

        Coach.objects.create(
            coach_code=coach_code,
            coach_name=request.POST.get('coach_name', ''),
            coach_level=request.POST.get('coach_level', ''),
            phone=phone,
            dpart=request.POST.get('dpart', ''),
            order_seq=int(request.POST.get('order_seq', '0') or '0'),
            coach_s_img=coach_s_img,
            use_gbn='Y',
            insert_dt=timezone.now(),
        )
        return redirect('office_coach_list')

    # GET
    levl_list = CodeValue.objects.filter(group__grpcode='LEVL', del_chk='N').order_by('code_order')
    return render(request, 'ba_office/lfcourse/coach_write.html', {
        'levl_list': levl_list,
    })


@office_login_required
@office_permission_required('L')
def coach_modify(request, coach_code):
    """코치 수정 — 코치는 coach_code(코치코드)로 식별(목록/삭제 링크 동일). pk(id)와 다른 값."""
    coach = get_object_or_404(Coach, coach_code=coach_code)

    if request.method == 'POST':
        coach.coach_name = request.POST.get('coach_name', '')
        coach.coach_level = request.POST.get('coach_level', '')

        # 연락처: mhtel1)mhtel2-mhtel3 형식
        mhtel1 = request.POST.get('mhtel1', '')
        mhtel2 = request.POST.get('mhtel2', '')
        mhtel3 = request.POST.get('mhtel3', '')
        coach.phone = f'{mhtel1}){mhtel2}-{mhtel3}' if mhtel1 else ''

        coach.dpart = request.POST.get('dpart', '')
        coach.order_seq = int(request.POST.get('order_seq', '0') or '0')

        # 이미지 파일 업로드 (새 파일이 있으면 교체)
        if request.FILES.get('coach_s_img'):
            f = request.FILES['coach_s_img']
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'fcdata', 'coach')
            os.makedirs(upload_dir, exist_ok=True)
            filename = f'{coach.coach_code}_{f.name}'
            filepath = os.path.join(upload_dir, filename)
            with open(filepath, 'wb+') as dest:
                for chunk in f.chunks():
                    dest.write(chunk)
            coach.coach_s_img = filename

        coach.save()
        return redirect('office_coach_list')

    # GET: phone → mhtel1/2/3 분리
    mhtel1, mhtel2, mhtel3 = '', '', ''
    if coach.phone and ')' in coach.phone:
        parts = coach.phone.split(')')
        mhtel1 = parts[0]
        rest = parts[1] if len(parts) > 1 else ''
        if '-' in rest:
            mhtel2, mhtel3 = rest.split('-', 1)
        else:
            mhtel2 = rest

    coach.mhtel1 = mhtel1
    coach.mhtel2 = mhtel2
    coach.mhtel3 = mhtel3

    levl_list = CodeValue.objects.filter(group__grpcode='LEVL', del_chk='N').order_by('code_order')
    return render(request, 'ba_office/lfcourse/coach_modify.html', {
        'coach': coach,
        'levl_list': levl_list,
    })


@office_login_required
@office_permission_required('L')
def coach_del(request, coach_code):
    """코치 삭제 (소프트)"""
    if request.method == 'POST':
        coach = get_object_or_404(Coach, coach_code=coach_code)
        coach.use_gbn = 'N'
        coach.save()
    return redirect('office_coach_list')


# ============================================================
# 과정관리 > 강좌관리
# ============================================================

@office_login_required
@office_permission_required('L')
def lecture_list(request):
    """강좌 목록"""
    use_gbn = request.GET.get('use_gbn', 'Y')
    # [UX변경] 원본: local_code = request.GET.get('local_code', '') (필드 제거)
    sel_sta_code = request.GET.get('sel_sta_code', '')
    class_gbn = request.GET.get('class_gbn', '')
    lecture_day = request.GET.get('lecture_day', '')

    # [UX변경] 구장 전체 직접 표시. 원본: locd_list + AJAX cascade
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    lectures = []
    total_count = 0
    txt_dt = ''

    # sta_code > 0 일 때만 목록 표시 (ASP 원본과 동일)
    if sel_sta_code:
        qs = Lecture.objects.filter(use_gbn=use_gbn)

        if sel_sta_code:
            qs = qs.filter(stadium__sta_code=int(sel_sta_code))
        if class_gbn:
            qs = qs.filter(class_gbn=class_gbn)
        if lecture_day:
            qs = qs.filter(lecture_day=int(lecture_day))

        qs = qs.select_related('stadium', 'coach', 't_coach').order_by('lecture_day', 'class_gbn', 'lecture_time')

        now = datetime.now()
        try:
            course_ym_dt = date(now.year, now.month, 1)
        except ValueError:
            course_ym_dt = date(now.year, 1, 1)

        txt_dt = f'{now.year}년 {now.month:02d}월'

        # 다음달 계산
        if now.month == 12:
            next_year = now.year + 1
            next_month = 1
        else:
            next_year = now.year
            next_month = now.month + 1

        for lec in qs:
            # 수강인원 통계
            base_filter = dict(
                lecture_code=lec.lecture_code,
                course_ym=course_ym_dt,
                bill_code='1001',
            )
            lec.c_student = EnrollmentCourse.objects.filter(
                **base_filter, course_stats='LY'
            ).count()
            lec.e_student = EnrollmentCourse.objects.filter(
                **base_filter, course_stats='LP'
            ).count()
            lec.f_student = EnrollmentCourse.objects.filter(
                **base_filter, course_stats__in=['LN', 'PN', 'LS']
            ).count()

            # 누적 통계
            acc_base_filter = dict(
                lecture_code=lec.lecture_code,
                bill_code='1001',
            )
            lec.acc_c_student = EnrollmentCourse.objects.filter(
                **acc_base_filter, course_stats='LY'
            ).count()
            lec.acc_e_student = EnrollmentCourse.objects.filter(
                **acc_base_filter, course_stats='LP'
            ).count()

            # 대기인원
            lec.dae_student = WaitStudent.objects.filter(
                lecture_code=lec.lecture_code, trans_gbn='N', del_chk='N'
            ).count()

            # 다음달 시간표 존재 여부
            lec.next_timetable = LectureSelDay.objects.filter(
                lecture_code=lec.lecture_code, syear=next_year, smonth=next_month
            ).count()

        lectures = list(qs)
        total_count = len(lectures)

    now_for_dt = datetime.now()
    cur_dt = f'{now_for_dt.year}{now_for_dt.month:02d}'

    return render(request, 'ba_office/lfcourse/lecture_list.html', {
        'lectures': lectures,
        'total_count': total_count,
        'txt_dt': txt_dt,
        'cur_dt': cur_dt,
        'use_gbn': use_gbn,
        'sel_sta_code': sel_sta_code,
        'class_gbn': class_gbn,
        'lecture_day': lecture_day,
        'stadiums': stadiums,
    })


@office_login_required
@office_permission_required('L')
def lecture_write(request):
    """강좌 등록"""
    if request.method == 'POST':
        max_code = Lecture.objects.aggregate(m=Coalesce(Max('lecture_code'), 0))['m']
        lecture_code = max_code + 1

        sta_code = int(request.POST.get('sta_code', '0') or '0')
        stadium_obj = Stadium.objects.filter(sta_code=sta_code).first()

        coach_code = request.POST.get('coach_code', '')
        coach_obj = Coach.objects.filter(coach_code=int(coach_code)).first() if coach_code else None

        t_coach_code = request.POST.get('t_coach_code', '')
        t_coach_obj = Coach.objects.filter(coach_code=int(t_coach_code)).first() if t_coach_code else None

        lec_day = int(request.POST.get('lecture_day', '0') or '0')
        lec_time = request.POST.get('lecture_time', '')
        cls_gbn = request.POST.get('class_gbn', '')

        # 강좌명 자동생성: "{구장닉네임}_{요일}_{시간}_{클래스}"
        day_names = {1: '월', 2: '화', 3: '수', 4: '목', 5: '금', 6: '토', 7: '일'}
        nickname = stadium_obj.sta_nickname if stadium_obj else ''
        day_str = day_names.get(lec_day, '')
        lecture_title = request.POST.get('lecture_title', '')
        if not lecture_title and nickname:
            lecture_title = f'{nickname}_{day_str}_{lec_time}_{cls_gbn}'

        Lecture.objects.create(
            lecture_code=lecture_code,
            # [UX변경] 필드(권역) 직접입력 제거 → 선택 구장의 local_code 사용
            local_code=(stadium_obj.local_code if stadium_obj else 0),
            stadium=stadium_obj,
            lecture_title=lecture_title,
            lec_age=request.POST.get('lec_age', ''),
            lecture_day=lec_day,
            lecture_time=lec_time,
            class_gbn=cls_gbn,
            class_gbn2=request.POST.get('class_gbn2', ''),
            lec_price=int(request.POST.get('lec_price', '0') or '0'),
            stu_cnt=int(request.POST.get('stu_cnt', '0') or '0'),
            coach=coach_obj,
            t_coach=t_coach_obj,
            sub_coach=request.POST.get('sub_coach', ''),
            dc_2=int(request.POST.get('dc_2', '0') or '0'),
            dc_3=int(request.POST.get('dc_3', '0') or '0'),
            dc_4=int(request.POST.get('dc_4', '0') or '0'),
            use_gbn='Y',
            insert_dt=timezone.now(),
        )
        return redirect('office_lecture_list')

    # GET
    locd_list = CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N').order_by('code_order')
    stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('order_seq', 'coach_name')
    return render(request, 'ba_office/lfcourse/lecture_write.html', {
        'locd_list': locd_list,
        'stadiums': stadiums,
        'coaches': coaches,
    })


@office_login_required
@office_permission_required('L')
def lecture_modify(request, lecture_code):
    """강좌 수정 — 강좌는 lecture_code(강좌코드)로 식별(목록/삭제/시간표 동일). pk(id)와 다른 값."""
    lecture = get_object_or_404(Lecture, lecture_code=lecture_code)

    if request.method == 'POST':
        sta_code = int(request.POST.get('sta_code', '0') or '0')
        stadium_obj = Stadium.objects.filter(sta_code=sta_code).first()

        coach_code = request.POST.get('coach_code', '')
        coach_obj = Coach.objects.filter(coach_code=int(coach_code)).first() if coach_code else None

        t_coach_code = request.POST.get('t_coach_code', '')
        t_coach_obj = Coach.objects.filter(coach_code=int(t_coach_code)).first() if t_coach_code else None

        # [UX변경] 필드(권역) 직접입력 제거 → 선택 구장의 local_code 사용
        lecture.local_code = stadium_obj.local_code if stadium_obj else 0
        lecture.stadium = stadium_obj
        lecture.lecture_title = request.POST.get('lecture_title', '')
        lecture.lec_age = request.POST.get('lec_age', '')
        lecture.lecture_day = int(request.POST.get('lecture_day', '0') or '0')
        lecture.lecture_time = request.POST.get('lecture_time', '')
        lecture.class_gbn = request.POST.get('class_gbn', '')
        lecture.class_gbn2 = request.POST.get('class_gbn2', '')
        lecture.lec_price = int(request.POST.get('lec_price', '0') or '0')
        lecture.stu_cnt = int(request.POST.get('stu_cnt', '0') or '0')
        lecture.coach = coach_obj
        lecture.t_coach = t_coach_obj
        lecture.sub_coach = request.POST.get('sub_coach', '')
        lecture.dc_2 = int(request.POST.get('dc_2', '0') or '0')
        lecture.dc_3 = int(request.POST.get('dc_3', '0') or '0')
        lecture.dc_4 = int(request.POST.get('dc_4', '0') or '0')
        lecture.save()
        return redirect('office_lecture_list')

    # GET
    locd_list = CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N').order_by('code_order')
    all_stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    coaches = Coach.objects.filter(use_gbn='Y').order_by('order_seq', 'coach_name')

    # 현재 강좌의 권역으로 필터된 구장 목록
    filtered_stadiums = all_stadiums.filter(local_code=lecture.local_code) if lecture.local_code else all_stadiums

    # FK 코드값을 템플릿에서 비교할 수 있도록 속성 추가
    lecture.sta_code = lecture.stadium.sta_code if lecture.stadium else 0
    lecture.coach_code = lecture.coach.coach_code if lecture.coach else 0
    lecture.t_coach_code = lecture.t_coach.coach_code if lecture.t_coach else 0

    return render(request, 'ba_office/lfcourse/lecture_modify.html', {
        'lecture': lecture,
        'locd_list': locd_list,
        'all_stadiums': all_stadiums,
        'filtered_stadiums': filtered_stadiums,
        'coaches': coaches,
    })


@office_login_required
@office_permission_required('L')
def lecture_del(request, lecture_code):
    """강좌 삭제 (소프트)"""
    if request.method == 'POST':
        lecture = get_object_or_404(Lecture, lecture_code=lecture_code)
        lecture.use_gbn = 'N'
        lecture.save()
    return redirect('office_lecture_list')


# ============================================================
# 과정관리 > 시간표 관리
# ============================================================

@office_login_required
@office_permission_required('L')
def _build_timetable_html(lecture_code, syear, smonth):
    """월별 시간표 HTML 문자열 생성 (AJAX 응답용)"""
    days = LectureSelDay.objects.filter(
        lecture_code=lecture_code, syear=syear, smonth=smonth
    ).order_by('sday')
    html = ''
    for d in days:
        html += (
            f"<span style='padding:5px 3px 5px 3px;width:50px;height:30px;"
            f"border:1px solid #d9d9d9;'>{d.sday}일 "
            f"<img src='/static/ba_office/images/ico_x.gif' "
            f"onclick=\"delDay({d.pk},{smonth});\" style='cursor:pointer;' />"
            f"</span>&nbsp;&nbsp;&nbsp;"
        )
    return html


def lecture_timetable(request, lecture_code):
    """시간표 관리"""
    lecture = get_object_or_404(Lecture, lecture_code=lecture_code)

    now = datetime.now()
    syear = int(request.GET.get('syear', str(now.year)))

    if request.method == 'POST':
        mode = request.POST.get('mode', '')
        admin_id = request.session.get('office_user', {}).get('office_id', '')

        if mode == 'add':
            add_year = int(request.POST.get('syear', str(syear)))
            add_month = int(request.POST.get('smonth', '1'))
            add_day = int(request.POST.get('sday', '1'))
            LectureSelDay.objects.get_or_create(
                lecture_code=lecture_code,
                syear=add_year,
                smonth=add_month,
                sday=add_day,
                defaults={'admin_id': admin_id},
            )
            html = _build_timetable_html(lecture_code, add_year, add_month)
            return JsonResponse({'status': 'ok', 'html': html})

        elif mode == 'del':
            uid = request.POST.get('uid', '')
            del_year = int(request.POST.get('syear', str(syear)))
            del_month = int(request.POST.get('smonth', '1'))
            if uid:
                LectureSelDay.objects.filter(pk=int(uid)).delete()
            html = _build_timetable_html(lecture_code, del_year, del_month)
            return JsonResponse({'status': 'ok', 'html': html})

        elif mode == 'bulk':
            post_year = int(request.POST.get('syear', str(syear)))
            # 해당 년도 전체 삭제 후 요일 기반 자동 생성
            LectureSelDay.objects.filter(
                lecture_code=lecture_code, syear=post_year
            ).delete()

            lec_day = lecture.lecture_day  # 1~7 (월~일)
            # Python weekday: 0=월, 6=일 → lec_day: 1=월, 7=일
            py_weekday = lec_day - 1  # 0~6

            for month in range(1, 13):
                cal = calendar.monthcalendar(post_year, month)
                for week in cal:
                    day = week[py_weekday]
                    if day != 0:
                        LectureSelDay.objects.get_or_create(
                            lecture_code=lecture_code,
                            syear=post_year,
                            smonth=month,
                            sday=day,
                            defaults={'admin_id': admin_id},
                        )
            return JsonResponse({'status': 'ok'})

    # GET: 12개월 시간표 로드
    seldays = LectureSelDay.objects.filter(
        lecture_code=lecture_code, syear=syear
    ).order_by('smonth', 'sday')

    # 월별로 그룹핑
    monthly_days = {}
    for sd in seldays:
        monthly_days.setdefault(sd.smonth, []).append(sd)

    # timetable_data: 1~12월 각각 days 리스트 포함
    timetable_data = []
    for m in range(1, 13):
        timetable_data.append({
            'month': m,
            'days': monthly_days.get(m, []),
        })

    # year_list 생성 (현재년도 ±2)
    year_list = list(range(now.year - 2, now.year + 3))

    # 권역명, 구장명 표시용
    locd_map = {cv.subcode: cv.code_name for cv in CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N')}
    lecture.local_code_name = locd_map.get(lecture.local_code, '')
    lecture.sta_code_name = lecture.stadium.sta_name if lecture.stadium else ''

    return render(request, 'ba_office/lfcourse/lecture_timetable.html', {
        'lecture': lecture,
        'syear': syear,
        'timetable_data': timetable_data,
        'year_list': year_list,
    })


# ============================================================
# 과정관리 > 훈련일정관리
# ============================================================

@office_login_required
@office_permission_required('L')
def train_list(request):
    """훈련일정 목록"""
    now = datetime.now()
    sel_year_code = request.GET.get('sel_year_code', '') or str(now.year)
    sel_month_code = request.GET.get('sel_month_code', '') or f'{now.month:02d}'
    # [UX변경] 원본: sch_locd_code = request.GET.get('sch_locd_code', '') (필드 제거)
    sch_sta_code = request.GET.get('sch_sta_code', '')

    # [UX변경] 구장 전체 직접 표시. 원본: locd_list + filtered_stadiums
    all_stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    # 연도 목록
    year_list = list(range(now.year - 2, now.year + 3))

    trains = []
    searched = False

    if sel_year_code and sel_month_code and sch_sta_code:
        searched = True
        try:
            y = int(sel_year_code)
            m = int(sel_month_code)
            start_date = date(y, m, 1)
            last_day = calendar.monthrange(y, m)[1]
            end_date = date(y, m, last_day)

            qs = LectureTraining.objects.filter(
                sta_code=int(sch_sta_code),
                training_dt__range=[start_date, end_date],
            ).order_by('training_dt')

            # 각 항목에 sta_name 추가
            sta_map = {s.sta_code: s.sta_name for s in Stadium.objects.all()}
            for t in qs:
                t.sta_name = sta_map.get(t.sta_code, '')
            trains = list(qs)
        except (ValueError, TypeError):
            pass

    return render(request, 'ba_office/lfcourse/train_list.html', {
        'trains': trains,
        'searched': searched,
        'sel_year_code': sel_year_code,
        'sel_month_code': sel_month_code,
        'sch_sta_code': sch_sta_code,
        'all_stadiums': all_stadiums,
        'year_list': year_list,
    })


@office_login_required
@office_permission_required('L')
def train_write(request):
    """훈련일정 등록"""
    if request.method == 'POST':
        sta_code = int(request.POST.get('sch_sta_code', '0') or '0')
        # [UX변경] 원본: local_code = int(request.POST.get('sch_locd_code', '0') or '0') (필드 제거)
        stadium_obj = Stadium.objects.filter(sta_code=sta_code).first()
        local_code = stadium_obj.local_code if stadium_obj else 0
        training_dt_str = request.POST.get('train_date', '')
        training_desc = request.POST.get('train_content', '')
        insert_id = request.session.get('office_user', {}).get('office_id', '')

        try:
            training_dt = datetime.strptime(training_dt_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            training_dt = None

        if training_dt:
            # 중복체크: 같은 sta_code + training_dt
            exists = LectureTraining.objects.filter(
                sta_code=sta_code, training_dt=training_dt
            ).exists()
            if exists:
                all_stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
                return render(request, 'ba_office/lfcourse/train_write.html', {
                    'error': '이미 동일한 날짜에 등록된 훈련일정이 있습니다.',
                    'all_stadiums': all_stadiums,
                })

            LectureTraining.objects.create(
                sta_code=sta_code,
                local_code=local_code,
                training_dt=training_dt,
                training_desc=training_desc,
                insert_dt=timezone.now(),
                insert_id=insert_id,
            )
        return redirect('office_train_list')

    # GET
    # [UX변경] 원본: locd_list = CodeValue.objects.filter(group__grpcode='LOCD'...) (필드 제거)
    all_stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    return render(request, 'ba_office/lfcourse/train_write.html', {
        'all_stadiums': all_stadiums,
    })


@office_login_required
@office_permission_required('L')
def train_modify(request, pk):
    """훈련일정 수정"""
    train = get_object_or_404(LectureTraining, pk=pk)

    if request.method == 'POST':
        train.sta_code = int(request.POST.get('sch_sta_code', '0') or '0')
        # [UX변경] 원본: train.local_code = int(request.POST.get('sch_locd_code', '0') or '0') (필드 제거)
        stadium_obj = Stadium.objects.filter(sta_code=train.sta_code).first()
        train.local_code = stadium_obj.local_code if stadium_obj else 0
        training_dt_str = request.POST.get('train_date', '')
        try:
            train.training_dt = datetime.strptime(training_dt_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
        train.training_desc = request.POST.get('train_content', '')
        train.save()
        return redirect('office_train_list')

    # GET
    # [UX변경] 원본: locd_list + filtered_stadiums (필드 제거)
    all_stadiums = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')
    return render(request, 'ba_office/lfcourse/train_write.html', {
        'train': train,
        'mode': 'edit',
        'all_stadiums': all_stadiums,
    })


@office_login_required
@office_permission_required('L')
def train_del(request, pk):
    """훈련일정 삭제 (하드)"""
    if request.method == 'POST':
        training = get_object_or_404(LectureTraining, pk=pk)
        training.delete()
    return redirect('office_train_list')


# ============================================================
# 과정관리 > 프로모션관리
# ============================================================

@office_login_required
@office_permission_required('L')
def promotion_list(request):
    """프로모션 목록"""
    useMode = request.GET.get('useMode', '0')
    stype = request.GET.get('stype', '')
    sword = request.GET.get('sword', '').strip()
    page = request.GET.get('page', '1')

    qs = Promotion.objects.all().order_by('-uid')

    # useMode 필터
    if useMode and useMode != '0':
        qs = qs.filter(use_mode=int(useMode))

    # stype/sword 검색
    if stype and sword:
        if stype == '1':
            # 제목 검색
            qs = qs.filter(title__icontains=sword)
        elif stype == '2':
            # member_id로 PromotionMember 조회
            coupon_uids = PromotionMember.objects.filter(
                member_id__icontains=sword
            ).values_list('coupon_uid', flat=True).distinct()
            qs = qs.filter(uid__in=coupon_uids)
        elif stype == '3':
            # 회원명으로 Member → member_id → PromotionMember
            member_ids = Member.objects.filter(
                name__icontains=sword
            ).values_list('username', flat=True)
            coupon_uids = PromotionMember.objects.filter(
                member_id__in=member_ids
            ).values_list('coupon_uid', flat=True).distinct()
            qs = qs.filter(uid__in=coupon_uids)
        elif stype == '4':
            # child_id로 PromotionMember 조회
            coupon_uids = PromotionMember.objects.filter(
                child_id__icontains=sword
            ).values_list('coupon_uid', flat=True).distinct()
            qs = qs.filter(uid__in=coupon_uids)
        elif stype == '5':
            # 자녀명으로 MemberChild → child_id → PromotionMember
            child_ids = MemberChild.objects.filter(
                name__icontains=sword
            ).values_list('child_id', flat=True)
            coupon_uids = PromotionMember.objects.filter(
                child_id__in=child_ids
            ).values_list('coupon_uid', flat=True).distinct()
            qs = qs.filter(uid__in=coupon_uids)

    # 표시용 문자열 추가
    use_mode_map = {0: '전체', 1: '교육용품비', 2: '수강료', 3: '결제금액'}
    issue_mode_map = {0: '-', 1: '회원별', 2: '야드별', 3: '구장별'}

    for p in qs:
        p.strUseMode = use_mode_map.get(p.use_mode, '')
        p.strIssueMode = issue_mode_map.get(p.issue_mode, '')
        p.strIsUse = '사용' if p.is_use == 'T' else '미사용'
        start_str = p.start_date.strftime('%Y-%m-%d') if p.start_date else ''
        end_str = p.end_date.strftime('%Y-%m-%d') if p.end_date else ''
        p.strEventPeriod = f'{start_str} ~ {end_str}' if start_str or end_str else ''

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(page)
    total_count = paginator.count

    return render(request, 'ba_office/lfcourse/promotion_list.html', {
        'promotions': page_obj,
        'total_count': total_count,
        'useMode': useMode,
        'stype': stype,
        'sword': sword,
        'page_obj': page_obj,
    })


@office_login_required
@office_permission_required('L')
def promotion_input(request):
    """프로모션 등록/수정"""
    uid = request.GET.get('uid', '0') if request.method == 'GET' else request.POST.get('uid', '0')
    try:
        uid = int(uid)
    except (ValueError, TypeError):
        uid = 0

    promotion = None
    if uid > 0:
        promotion = Promotion.objects.filter(uid=uid).first()

    if request.method == 'POST':
        # 등록/수정 구분 (회원 처리에서 수정모드는 기존 회원 전체삭제 후 재등록)
        is_edit = promotion is not None
        # 템플릿 폼 필드명(camelCase) 기준으로 읽기
        use_mode = int(request.POST.get('useMode', '0') or '0')
        issue_mode = int(request.POST.get('issueMode', '0') or '0')
        title = request.POST.get('title', '')
        summary = request.POST.get('summary', '')
        start_date_str = request.POST.get('startDate', '')
        end_date_str = request.POST.get('endDate', '')
        discount = int(request.POST.get('discount', '0') or '0')
        discount_unit = request.POST.get('discountUnit', '')
        is_period_limit = request.POST.get('isPeriodLimit', '')  # 체크시 'F' (제한없음)
        is_price_limit = request.POST.get('isPriceLimit', '')  # 체크시 'F' (제한없음)
        min_price = int(request.POST.get('minPrice', '0') or '0')
        max_price = int(request.POST.get('maxPrice', '0') or '0')
        is_use = request.POST.get('isUse', 'T')

        # issueMode별 local_code / sta_code
        local_code_val = ''
        sta_code_val = ''
        if issue_mode == 2:
            local_code_val = request.POST.get('local_code_2', '')
        elif issue_mode == 3:
            # [UX변경] mode3: 권역→구장 cascade 제거 → 구장 직접선택, local_code는 구장에서 파생
            sta_code_val = request.POST.get('sta_code3', '')
            local_code_val = request.POST.get('local_code_3', '')
            if sta_code_val and not local_code_val:
                _sta = Stadium.objects.filter(sta_code=int(sta_code_val)).first()
                local_code_val = str(_sta.local_code) if _sta and _sta.local_code else ''

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
        except ValueError:
            start_date = None
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        except ValueError:
            end_date = None

        # 기간 제한없음 체크 시 날짜 무시
        if is_period_limit == 'F':
            start_date = None
            end_date = None

        if promotion:
            # 수정
            promotion.use_mode = use_mode
            promotion.issue_mode = issue_mode
            promotion.title = title
            promotion.summary = summary
            promotion.start_date = start_date
            promotion.end_date = end_date
            promotion.discount = discount
            promotion.discount_unit = discount_unit
            promotion.is_price_limit = is_price_limit if is_price_limit else 'T'
            promotion.min_price = min_price
            promotion.max_price = max_price
            promotion.is_use = is_use
            promotion.local_code = int(local_code_val) if local_code_val else None
            promotion.sta_code = int(sta_code_val) if sta_code_val else None
            promotion.save()
        else:
            # 등록
            max_uid = Promotion.objects.aggregate(m=Coalesce(Max('uid'), 0))['m']
            new_uid = max_uid + 1

            promotion = Promotion.objects.create(
                uid=new_uid,
                use_mode=use_mode,
                issue_mode=issue_mode,
                title=title,
                summary=summary,
                start_date=start_date,
                end_date=end_date,
                discount=discount,
                discount_unit=discount_unit,
                is_price_limit=is_price_limit if is_price_limit else 'T',
                min_price=min_price,
                max_price=max_price,
                is_use=is_use,
                local_code=int(local_code_val) if local_code_val else None,
                sta_code=int(sta_code_val) if sta_code_val else None,
                reg_date=timezone.now(),
            )
            uid = promotion.uid

        # ===== 프로모션 선택 회원 처리 (원본 lfpromotion_input_proc.asp) =====
        # 회원별(issueMode=1) + 회원전체선택 해제 + items(child_code 콤마문자열)가 있을 때만.
        # 원본: items에 콤마가 있을 때만 처리 → 전체선택(items 공백) 시 기존 회원은 그대로 둠.
        items_str = request.POST.get('items', '').strip().rstrip(',')
        is_allmem = request.POST.get('isAllmem', '')  # 회원전체선택 체크 시 'T'
        dup_names = []
        if issue_mode == 1 and is_allmem != 'T' and items_str:
            child_codes = [c.strip() for c in items_str.split(',') if c.strip()]

            # 수정모드: 기존 선택회원 전체 삭제 후 재등록 (원본 동일)
            if is_edit:
                PromotionMember.objects.filter(coupon_uid=promotion.uid).delete()

            # 동일 useMode 활성 프로모션 보유 자녀 (중복 검사용)
            dup_uids = list(Promotion.objects.filter(
                use_mode=use_mode, is_use='T'
            ).values_list('uid', flat=True))

            for code in child_codes:
                try:
                    child = MemberChild.objects.filter(child_code=int(code)).first()
                except (ValueError, TypeError):
                    child = None
                if not child:
                    continue
                # 이미 같은 구분(useMode)의 활성 프로모션을 사용중이면 제외 (원본 child_dup_chk)
                dup = PromotionMember.objects.filter(
                    child_id=child.child_id, coupon_uid__in=dup_uids
                ).exists()
                if dup:
                    dup_names.append(child.name or child.child_id)
                    continue
                PromotionMember.objects.create(
                    coupon_uid=promotion.uid,
                    member_id=child.parent.username if child.parent else '',
                    child_id=child.child_id,
                    used='T', is_trash='T',
                )

        if dup_names:
            # 원본 AlertBack 대응: 중복 회원은 제외하고 나머지는 저장한 뒤 안내
            names = ', '.join(dup_names)
            return HttpResponse(
                '<script>alert("다음 회원은 기존 프로모션(동일 구분)을 사용중이라 제외되었습니다:\\n%s");'
                'location.href="/office/lfcourse/promotion/";</script>' % names
            )
        return redirect('office_promotion_list')

    # GET - 템플릿에서 사용하는 개별 변수로 전달
    locd_list = CodeValue.objects.filter(group__grpcode='LOCD', del_chk='N').order_by('code_order')
    stadiums_qs = Stadium.objects.filter(use_gbn='Y').order_by('sta_name')

    # 기본값 설정
    mode = 'add'
    mode_text = '등록'
    useMode = 0
    issueMode = 1
    p_title = ''
    p_summary = ''
    startDate = ''
    endDate = ''
    p_discount = ''
    discountUnit = 'WON'  # 기본 원 (DB 저장값: WON/PCT)
    isPeriodLimit = False
    isPriceLimit = False
    minPrice = ''
    maxPrice = ''
    isUse = 'T'
    isAllmem = True  # 원본 기본값: 회원전체선택 체크(전체회원 대상)
    local_code = ''
    sta_code = ''
    items = ''

    if promotion:
        mode = 'edit'
        mode_text = '수정'
        useMode = promotion.use_mode
        issueMode = promotion.issue_mode
        p_title = promotion.title
        p_summary = promotion.summary
        startDate = promotion.start_date.strftime('%Y-%m-%d') if promotion.start_date else ''
        endDate = promotion.end_date.strftime('%Y-%m-%d') if promotion.end_date else ''
        p_discount = promotion.discount
        discountUnit = promotion.discount_unit or 'WON'
        isPeriodLimit = not (promotion.start_date or promotion.end_date)
        isPriceLimit = (promotion.is_price_limit == 'F')
        minPrice = promotion.min_price
        maxPrice = promotion.max_price
        isUse = promotion.is_use or 'T'
        local_code = promotion.local_code or ''
        sta_code = getattr(promotion, 'sta_code', '') or ''

        # 프로모션 선택회원 → items(child_code 콤마문자열) 구성.
        # 화면 표시는 JS ViewMemberList()가 items로 sel_member 프래그먼트를 AJAX 로드(원본 동일).
        pm_qs = PromotionMember.objects.filter(coupon_uid=promotion.uid, used='T', is_trash='T')
        child_codes = []
        for pm in pm_qs:
            child = MemberChild.objects.filter(child_id=pm.child_id).first()
            if child and child.child_code:
                child_codes.append(str(child.child_code))
        if child_codes:
            items = ','.join(child_codes) + ','  # 원본과 동일하게 끝에 콤마
            isAllmem = False  # 선택회원이 있으면 전체선택 해제

    return render(request, 'ba_office/lfcourse/promotion_input.html', {
        'promotion': promotion,
        'uid': uid,
        'mode': mode,
        'mode_text': mode_text,
        'useMode': useMode,
        'issueMode': issueMode,
        'title': p_title,
        'summary': p_summary,
        'startDate': startDate,
        'endDate': endDate,
        'discount': p_discount,
        'discountUnit': discountUnit,
        'isPeriodLimit': isPeriodLimit,
        'isPriceLimit': isPriceLimit,
        'minPrice': minPrice,
        'maxPrice': maxPrice,
        'isUse': isUse,
        'isAllmem': isAllmem,
        'local_code': local_code,
        'sta_code': sta_code,
        'items': items,
        'locd_list': locd_list,
        'stadiums': stadiums_qs,
    })


@office_login_required
@office_permission_required('L')
def promotion_member_del(request):
    """프로모션 회원 즉시 삭제 (원본 lfpromotion_member_del_proc.asp)

    child_code(자녀코드)를 받아 child_id로 변환 후 해당 프로모션의 회원을 DB에서 삭제.
    원본 안내문대로 '수정처리를 하지 않아도 실제 삭제'된다.
    """
    uid = request.GET.get('uid', '0')
    child_code = request.GET.get('child_code', '')

    try:
        uid_int = int(uid)
    except (ValueError, TypeError):
        uid_int = 0

    if uid_int > 0 and child_code:
        try:
            child = MemberChild.objects.filter(child_code=int(child_code)).first()
        except (ValueError, TypeError):
            child = None
        if child:
            PromotionMember.objects.filter(
                coupon_uid=uid_int, child_id=child.child_id
            ).delete()

    return redirect(f'/office/lfcourse/promotion/input/?uid={uid}')


# ============================================================
# 과정관리 > AJAX
# ============================================================

@office_login_required
def ajax_course_stadium(request):
    """AJAX: 권역별 구장 목록"""
    local_code = request.GET.get('local_code', '')
    qs = Stadium.objects.filter(use_gbn='Y')
    if local_code:
        qs = qs.filter(local_code=int(local_code))
    qs = qs.order_by('sta_name')
    data = [{'sta_code': s.sta_code, 'sta_name': s.sta_name} for s in qs]
    return JsonResponse(data, safe=False)


# ============================================================
# 과정관리 > 수업시간표 팝업
# ============================================================

@office_login_required
def course_list_popup(request):
    """구장별 수업시간표 팝업"""
    sta_code = request.GET.get('sta_code', '')
    lecture_dt = request.GET.get('lecture_dt', '')

    now = timezone.now()
    # 날짜 옵션: 최근 12개월
    date_options = []
    for i in range(12):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        date_options.append(f'{y}-{m:02d}')
    if not lecture_dt:
        lecture_dt = date_options[0]

    class_list = []
    if sta_code:
        stadium = Stadium.objects.filter(sta_code=int(sta_code)).first()
        if stadium:
            for cls_gbn in ['A', 'B', 'C', 'D']:
                lectures = Lecture.objects.filter(
                    stadium=stadium,
                    class_gbn=cls_gbn,
                    use_gbn='Y'
                ).order_by('lecture_day', 'lecture_time', 'lec_age')

                if not lectures.exists():
                    continue

                rows = []
                rowspan_count = {}
                for lec in lectures:
                    key = (lec.lecture_day, lec.lecture_time)
                    if key not in rowspan_count:
                        rowspan_count[key] = 0
                    rowspan_count[key] += 1

                prev_key = None
                for lec in lectures:
                    key = (lec.lecture_day, lec.lecture_time)
                    days = {1: '월', 2: '화', 3: '수', 4: '목', 5: '금', 6: '토', 7: '일'}
                    try:
                        dt_parts = lecture_dt.split('-')
                        course_ym_dt = date(int(dt_parts[0]), int(dt_parts[1]), 1)
                    except Exception:
                        course_ym_dt = date(now.year, now.month, 1)
                    cur_cnt = EnrollmentCourse.objects.filter(
                        lecture_code=lec.lecture_code,
                        course_stats__in=['LY', 'LP'],
                        course_ym=course_ym_dt,
                        bill_code='1001',
                    ).count()

                    row = {
                        'lec_age': lec.lec_age or '-',
                        'stu_cnt': lec.stu_cnt,
                        'cur_cnt': cur_cnt,
                        'lecture_time': lec.lecture_time or '-',
                        'day_name': days.get(lec.lecture_day, ''),
                        'is_first': key != prev_key,
                        'rowspan': rowspan_count.get(key, 1),
                    }
                    rows.append(row)
                    prev_key = key

                class_list.append({
                    'class_gbn': cls_gbn,
                    'rows': rows,
                })

    return render(request, 'ba_office/lfcourse/course_list_popup.html', {
        'sta_code': sta_code,
        'lecture_dt': lecture_dt,
        'date_options': date_options,
        'class_list': class_list,
    })


# ============================================================
# 과정관리 > 프로모션 회원 팝업
# ============================================================

@office_login_required
@office_permission_required('L')
def promotion_member_popup(request):
    """프로모션 회원 검색/선택 팝업 (원본 member_list.asp)

    DB에 저장하지 않고 검색결과만 보여준다. 체크 후 [등록] 시 부모창의
    sel_member(child_code들)를 호출해 hidden items에 누적시키고 팝업을 닫는다.
    """
    uid = request.GET.get('uid', '')
    skey = request.GET.get('skey', 'm.member_id')
    sword = request.GET.get('sword', '').strip()
    child_code = request.GET.get('child_code', '')  # 이미 선택된 child_code 콤마문자열
    page = request.GET.get('page', '1')

    PAGE_SIZE = 15

    # 활성회원(부모 status='N')의 자녀 행
    qs = MemberChild.objects.select_related('parent').filter(parent__status='N')

    # 이미 선택된 회원 제외
    selected_codes = [c.strip() for c in child_code.split(',') if c.strip()]
    if selected_codes:
        try:
            qs = qs.exclude(child_code__in=[int(c) for c in selected_codes])
        except (ValueError, TypeError):
            pass

    # 검색
    if sword:
        skey_map = {
            'm.member_id': 'parent__username__icontains',
            'member_name': 'parent__name__icontains',
            'child_id': 'child_id__icontains',
            'child_name': 'name__icontains',
            'sch_name': 'school__icontains',
            'sch_grade': 'grade__icontains',
        }
        lookup = skey_map.get(skey, 'parent__name__icontains')
        qs = qs.filter(**{lookup: sword})

    qs = qs.order_by('-parent__insert_dt', 'parent__username')

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(page)

    # 동일 자녀의 현재 활성 프로모션 (구분 표시)
    use_mode_label = {1: '교육용품비', 2: '수강료', 3: '결제금액'}
    now = timezone.now()
    members = []
    for c in page_obj:
        parent = c.parent
        str_school = '-'
        if c.school:
            str_school = c.school
        if c.grade:
            str_school = (str_school if c.school else '') + f'[{c.grade}]'
        if not c.school and not c.grade:
            str_school = '-'

        # 보유 프로모션 (IsUse='T' + 기간내(종료일 없거나 시작<=오늘<=종료))
        prom_uids = PromotionMember.objects.filter(
            child_id=c.child_id
        ).values_list('coupon_uid', flat=True)
        prom_qs = Promotion.objects.filter(uid__in=list(prom_uids), is_use='T').filter(
            Q(end_date__isnull=True) | Q(start_date__lte=now, end_date__gte=now)
        )
        str_promotion = ' '.join(
            f'{p.title}({use_mode_label.get(p.use_mode, "")})' for p in prom_qs
        )

        members.append({
            'child_code': c.child_code,
            'member_id': parent.username if parent else '',
            'member_name': parent.name if parent else '',
            'child_id': c.child_id,
            'child_name': c.name,
            'str_school': str_school,
            'insert_dt': parent.insert_dt.strftime('%Y-%m-%d') if parent and parent.insert_dt else '',
            'str_promotion': str_promotion,
        })

    return render(request, 'ba_office/lfcourse/promotion_member_popup.html', {
        'uid': uid,
        'skey': skey,
        'sword': sword,
        'child_code': child_code,
        'members': members,
        'page_obj': page_obj,
        'total_count': paginator.count,
    })


@office_login_required
@office_permission_required('L')
def promotion_sel_member(request):
    """선택된 회원 목록 프래그먼트 (원본 sel_member.asp)

    items(child_code 콤마문자열)를 받아 부모창 #member_list에 AJAX로 렌더된다.
    """
    child_code = request.GET.get('childCode', '')
    uid = request.GET.get('uid', '0')
    codes = [c.strip() for c in child_code.split(',') if c.strip()]

    members = []
    if codes:
        try:
            int_codes = [int(c) for c in codes]
        except (ValueError, TypeError):
            int_codes = []
        # 입력 순서 유지
        child_map = {
            c.child_code: c for c in
            MemberChild.objects.select_related('parent').filter(child_code__in=int_codes)
        }
        for idx, code in enumerate(int_codes, start=1):
            c = child_map.get(code)
            if not c:
                continue
            parent = c.parent
            str_school = '-'
            if c.school:
                str_school = c.school
            if c.grade:
                str_school = (str_school if c.school else '') + f'[{c.grade}]'
            if not c.school and not c.grade:
                str_school = '-'
            members.append({
                'no': idx,
                'child_code': c.child_code,
                'member_id': parent.username if parent else '',
                'member_name': parent.name if parent else '',
                'child_id': c.child_id,
                'child_name': c.name,
                'str_school': str_school,
                'insert_dt': parent.insert_dt.strftime('%Y-%m-%d') if parent and parent.insert_dt else '',
            })

    return render(request, 'ba_office/lfcourse/promotion_sel_member.html', {
        'uid': uid,
        'members': members,
    })
